from __future__ import annotations

from datetime import datetime
from itertools import product
from pathlib import Path
import re
import subprocess
import math
import json
import shutil

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from runtime_config import (
    ROOT,
    INPUTS,
    RUNS,
    RESULTS,
    VIEWER_ASSETS,
    PHREEQC_EXE,
    DATABASE,
    phreeqc_configuration_help,
)

# =========================================================
# CONFIGURACION
# =========================================================

RHO_WATER_KG_L = 0.997
MW_H2O_KG_MOL = 0.01801528
MAX_EVAP_FRACTION_PER_PHREEQC_STEP = 0.05
RECOVERY_CONTROL_FILE = INPUTS / "recovery_control.txt"

RECOVERY_METHOD_TEXT = """Modulo auxiliar de recuperacion de agua

Este script compara dos ejecuciones completas del tren de piscinas:
1. Caso base sin recuperacion auxiliar.
2. Caso recovery con extraccion lateral, modulo PHREEQC y reinyeccion posterior.

El modulo se trata como una caja negra conceptual de recuperacion de agua. PHREEQC se usa
para representar la concentracion interna de la salmuera derivada, la precipitacion mineral
y la salmuera clarificada de retorno.

La extraccion se realiza sobre la corriente de entrada de extraction_stage_id antes de
aplicar la evaporacion de esa etapa. Si extraction_stage_id = PC1, la corriente del modulo
se toma directamente de la entrada de PC1 y la PC1 visible solo evapora la corriente
principal restante.

target_freshwater_L_s representa agua pura recuperada, no salmuera retirada. La tasa
de conversion del modulo no se lee del TXT: se calcula para cada escenario a partir
del caso base como la disminucion relativa de Q_in entre la etapa de extraccion y la
etapa de reinyeccion. Con esa conversion se calcula el caudal derivado al modulo:
Q_module_L_s = target_freshwater_L_s / module_conversion_fraction.

La reduccion de evaporacion de las piscinas saltadas se reparte uniformemente:
target_freshwater_L_s / numero_de_piscinas_saltadas. Para cada piscina, el porcentaje
de reduccion se calcula frente a la evaporacion del caso base de esa misma piscina.

La piscina auxiliar N+1 representa una decantacion interna de solidos dentro del modulo:
no se crea como etapa visible, no aparece en el tren principal y sus solidos precipitados
se registran exclusivamente en recovery_module_results.

distributed_effect_results es una visualizacion distribuida del efecto del modulo entre
la etapa de extraccion y la etapa de reinyeccion. No es una simulacion PHREEQC y no modifica
el balance real del caso recovery.

El modelo recovery evalua el impacto conceptual sobre el tren de piscinas. No constituye
validacion industrial ni experimental del modulo de recuperacion.
"""

DEFAULT_SCENARIO_VALUES = {
    "months": 1.0,
    "temperature": 25.0,
    "atmospheric_pressure": 1.0,
    "evap_factor": 1.0,
    "mg_removal": 0.0,
    "retention_r": 0.0,
}

DEFAULT_STEPS = {
    "months": 1.0,
    "temperature": 5.0,
    "atmospheric_pressure": 0.1,
    "evap_factor": 0.05,
    "mg_removal": 0.05,
    "retention_r": 0.05,
}

AQUEOUS_TOTALS = ["Li", "K", "Mg", "B", "Ca", "Na", "Cl", "S6"]
OUTPUT_SPECIES = ["Li", "K", "Mg", "B", "Ca", "Na", "Cl", "HCO3", "S6"]

MW_SO4_G_MOL = 96.06

MINERAL_FAMILIES = {
    "Halite": "chlorides",
    "Sylvite": "chlorides",
    "Carnallite": "chlorides",
    "Bischofite": "chlorides",
    "MgCl2_2H2O": "chlorides",
    "MgCl2_4H2O": "chlorides",
    "Glauberite": "sulfates",
    "Gypsum": "sulfates",
    "Anhydrite": "sulfates",
    "Thenardite": "sulfates",
    "Mirabilite": "sulfates",
    "Arcanite": "sulfates",
    "Epsomite": "sulfates",
    "Kieserite": "sulfates",
    "Bloedite": "sulfates",
    "Glaserite": "sulfates",
    "Goergeyite": "sulfates",
    "Hexahydrite": "sulfates",
    "Kainite": "sulfates",
    "Labile_S": "sulfates",
    "Leonhardite": "sulfates",
    "Leonite": "sulfates",
    "Pentahydrite": "sulfates",
    "Polyhalite": "sulfates",
    "Schoenite": "sulfates",
    "Syngenite": "sulfates",
    "Misenite": "sulfates",
    "Borax": "borates",
    "Boric_acid,s": "borates",
    "K2B4O7:4H2O": "borates",
    "KB5O8:4H2O": "borates",
    "NaB5O8:5H2O": "borates",
    "NaBO2:4H2O": "borates",
    "Teepleite": "borates",
    "Brucite": "hydroxides",
    "Portlandite": "hydroxides",
    "Calcite": "carbonates",
    "Aragonite": "carbonates",
    "Magnesite": "carbonates",
    "Dolomite": "carbonates",
    "Huntite": "carbonates",
}

PHASE_SCALING_PENALTIES = {
    "Halite": 30.0,
    "Sylvite": 35.0,
    "Borates": 85.0,
    "Gypsum": 65.0,
    "Anhydrite": 70.0,
    "Glauberite": 65.0,
    "Bloedite": 65.0,
    "Thenardite": 55.0,
    "Mirabilite": 55.0,
    "Calcite": 75.0,
    "Aragonite": 75.0,
    "Magnesite": 75.0,
    "Dolomite": 75.0,
    "Huntite": 75.0,
    "Brucite": 90.0,
    "Carnallite": 80.0,
    "Bischofite": 80.0,
    "Epsomite": 75.0,
}
BORATE_SCALING_PENALTY = 85.0
DEFAULT_UNKNOWN_PHASE_PENALTY = 50.0
SCALING_RISK_WEIGHTS = {
    "solid_water_score": 0.30 + (0.20 / 3.0),
    "phase_penalty_score": 0.25 + (0.20 / 3.0),
    "saturation_score": 0.15 + (0.20 / 3.0),
    "chemical_treatment_penalty": 1.0,
}

PHREEQC_PHASE_ALIASES = {
    "Astrakanite": "Bloedite",
}
_DATABASE_PHASE_NAMES_CACHE: set[str] | None = None

SCALING_RISK_METHOD_TEXT = """Scaling risk indicator

El indicador scaling_risk_indicator es un parametro compuesto entre 0 y 100 por etapa. No sustituye al calculo geoquimico de PHREEQC: integra resultados ya generados por el balance de proceso, la precipitacion mineral y los indices de saturacion.

Formula actual:
scaling_risk_indicator =
    0.366667 * Psa_solids_water_score
  + 0.316667 * Pf_phase_penalty_score
  + 0.216667 * Ps_saturation_score
    + chemical_treatment_penalty

1. Psa_solids_water_score
solids_water_ratio = total_solids_precipitated_kg_s / water_evaporated_kg_s
Si no hay agua evaporada, solids_water_ratio se guarda como 0 para evitar division por cero, pero Psa_solids_water_score = 100.
solid_water_score = 100 * solids_water_ratio / max_solids_water_ratio del conjunto exportado.

2. Pf_phase_penalty_score
Se calcula como media ponderada por masa precipitada de cada fase:
phase_penalty_score = sum(mass_phase_i * penalty_phase_i) / total_solids_precipitated
Penalizaciones base: Halite 30, Sylvite 35, Gypsum 65, Anhydrite 70, Glauberite 65, Bloedite 65, Thenardite 55, Mirabilite 55, Calcite 75, Aragonite 75, Magnesite 75, Dolomite 75, Huntite 75, Brucite 90, Carnallite 80, Bischofite 80, Epsomite 75. Las fases de la familia borates usan 85. Las fases no clasificadas usan 50 como valor provisional.

3. Ps_saturation_score
Se calcula a partir de los indices de saturacion leidos de PHREEQC:
SI >= 0 suma 20 puntos.
-0.2 <= SI < 0 suma 10 puntos.
SI < -0.2 no suma.
El resultado se limita a 100.

4. chemical_treatment_penalty
Se suma +10 si la etapa es LIM o chemical_treatment.
Se suma +10 si precipita Brucite.
Se suma +5 si precipita Gypsum o Anhydrite.

Clasificacion:
scaling_risk_indicator < 35: Bajo
35 <= scaling_risk_indicator < 60: Medio
scaling_risk_indicator >= 60: Alto

Nota metodologica:
El indicador es una herramienta de comparacion y cribado. Los umbrales y penalizaciones pueden recalibrarse cuando existan datos operativos o validacion experimental.
"""

ELEMENT_MW_G_MOL = {
    "Li": 6.941,
    "K": 39.0983,
    "Mg": 24.305,
    "B": 10.811,
    "Ca": 40.078,
    "Na": 22.98976928,
    "Cl": 35.453,
    "HCO3": 61.01684,
    "S6": 32.065,
}

PHASE_MW_KG_MOL = {
    "Halite": 0.058442769280,
    "Sylvite": 0.074551300000,
    "Glauberite": 0.278182738560,
    "Gypsum": 0.172171160000,
    "Anhydrite": 0.136140600000,
    "Thenardite": 0.142042138560,
    "Mirabilite": 0.322194938560,
    "Calcite": 0.100086900000,
    "Aragonite": 0.100086900000,
    "Magnesite": 0.084313900000,
    "Dolomite": 0.184400800000,
    "Huntite": 0.353028600000,
    "Brucite": 0.058319680000,
    "Portlandite": 0.074092680000,
    "Boric_acid,s": 0.061833020000,
    "Borax": 0.381372138560,
    "K2B4O7:4H2O": 0.305497520000,
    "KB5O8:4H2O": 0.293209620000,
    "NaB5O8:5H2O": 0.295116369280,
    "NaBO2:4H2O": 0.137860689280,
    "Teepleite": 0.160272898560,
    "Arcanite": 0.174259200000,
    "Bischofite": 0.203302680000,
    "Bloedite": 0.334470858560,
    "Carnallite": 0.277853980000,
    "Epsomite": 0.246474560000,
    "Glaserite": 0.332409869280,
    "Goergeyite": 0.872977480000,
    "Hexahydrite": 0.228459280000,
    "Kainite": 0.248964740000,
    "Kieserite": 0.138382880000,
    "Labile_S": 0.456255437120,
    "Leonhardite": 0.192428720000,
    "Leonite": 0.366687920000,
    "MgCl2_2H2O": 0.131241560000,
    "MgCl2_4H2O": 0.167272120000,
    "Pentahydrite": 0.210444000000,
    "Polyhalite": 0.602938560000,
    "Schoenite": 0.402718480000,
    "Syngenite": 0.328415080000,
    "Misenite": 0.991272240000,
}

MONTH_LABEL_TO_NUMBER = {
    "ene": 1,
    "enero": 1,
    "jan": 1,
    "january": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abr": 4,
    "abril": 4,
    "apr": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
    "dec": 12,
}


# =========================================================
# CARGA Y NORMALIZACION DE INPUTS TXT
# =========================================================

def read_semicolon_txt(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {path}")
    return pd.read_csv(path, sep=";")


def normalize_feed_columns(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "agua": "water_kg",
        "water": "water_kg",
        "h2o": "water_kg",
        "HCO3-": "HCO3",
        "HCO3": "HCO3",
        "S(6)": "S6",
        "SO4": "S6",
        "SO4/S(6)": "S6",
    }

    renamed = {}
    for col in df.columns:
        clean = str(col).strip()
        renamed[col] = aliases.get(clean, clean)
    return df.rename(columns=renamed)


def load_feed() -> dict:
    df = normalize_feed_columns(read_semicolon_txt(INPUTS / "feed_base.txt"))
    if df.empty:
        raise ValueError("feed_base.txt no contiene filas.")
    return df.iloc[0].to_dict()


def load_stages() -> pd.DataFrame:
    df = read_semicolon_txt(INPUTS / "stages_base.txt").copy()
    df.columns = [str(c).strip() for c in df.columns]

    if "surface_m2" not in df.columns and "evap_L_s" in df.columns:
        # El TXT nuevo conserva el encabezado antiguo, pero el valor representa superficie.
        df["surface_m2"] = df["evap_L_s"]

    if "reagent" not in df.columns:
        df["reagent"] = ""
    if "reagent_param" not in df.columns:
        df["reagent_param"] = 0.0

    df["reagent"] = df["reagent"].fillna("").astype(str).str.strip()
    df["reagent_param"] = pd.to_numeric(df["reagent_param"], errors="coerce").fillna(0.0)
    df["surface_m2"] = pd.to_numeric(df["surface_m2"], errors="coerce")

    return df


def normalize_month_column_name(col: str) -> str:
    raw = str(col).strip()
    norm = raw.lower()

    if norm in {"mes", "month", "months"}:
        return "month_label"
    if "f_m" in norm or norm == "fm":
        return "f_m"

    upper = raw.upper()
    if "TP" in upper and "PC" in upper:
        return "TP_PC"
    if "TP" in upper and "_H" in upper:
        return "TP_H"
    if "TP" in upper and "_K" in upper:
        return "TP_K"
    if "TP" in upper and "_C" in upper:
        return "TP_C"
    if "TP" in upper and "_L" in upper:
        return "TP_L"

    return raw


def parse_month_number(value) -> int:
    if pd.isna(value):
        raise ValueError("Mes vacio en months_control.txt.")

    text = str(value).strip()
    if is_base_month_label(text):
        return 0
    try:
        number = int(float(text))
        if 1 <= number <= 12:
            return number
    except ValueError:
        pass

    key = text.lower().replace(".", "")
    if key in MONTH_LABEL_TO_NUMBER:
        return MONTH_LABEL_TO_NUMBER[key]

    raise ValueError(f"No se pudo interpretar el mes '{value}' en months_control.txt.")


def is_base_month_label(value) -> bool:
    text = str(value).strip().lower().replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    return text in {"base", "base case", "case base"} or text.startswith("caso base")


def get_base_evaporation_row(months_control: pd.DataFrame) -> pd.Series:
    base_rows = months_control[months_control["month"].astype(int) == 0]
    if not base_rows.empty:
        return base_rows.iloc[0]
    raise ValueError("No se encontro la fila 'Caso base' en months_control.txt para fijar la evaporacion.")


def load_months_control() -> pd.DataFrame:
    df = read_semicolon_txt(INPUTS / "months_control.txt").copy()
    df = df.rename(columns={c: normalize_month_column_name(c) for c in df.columns})

    required_cols = ["month_label", "f_m", "TP_PC", "TP_H", "TP_K", "TP_C", "TP_L"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna '{col}' en months_control.txt")

    df["month"] = df["month_label"].apply(parse_month_number).astype(int)
    for col in ["f_m", "TP_PC", "TP_H", "TP_K", "TP_C", "TP_L"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        if df[col].isna().any():
            raise ValueError(f"Hay valores no numericos en '{col}' de months_control.txt")

    return df.sort_values("month").reset_index(drop=True)


def load_scenarios_control() -> pd.DataFrame:
    path = INPUTS / "scenarios_control.txt"
    if not path.exists():
        return pd.DataFrame(columns=["variable", "enabled", "min", "max", "step"])

    df = read_semicolon_txt(path).copy()
    df.columns = [str(c).strip() for c in df.columns]
    df["variable"] = df["variable"].astype(str).str.strip().str.lower()
    return df


# =========================================================
# VALIDACIONES
# =========================================================

def validate_paths() -> None:
    if PHREEQC_EXE is None:
        raise FileNotFoundError(f"No se ha configurado PHREEQC_EXE. {phreeqc_configuration_help()}")
    if DATABASE is None:
        raise FileNotFoundError(f"No se ha configurado PHREEQC_DATABASE. {phreeqc_configuration_help()}")
    if not PHREEQC_EXE.exists():
        raise FileNotFoundError(f"No existe PHREEQC_EXE: {PHREEQC_EXE}")
    if not DATABASE.exists():
        raise FileNotFoundError(f"No existe PHREEQC_DATABASE: {DATABASE}")


def validate_feed(feed: dict) -> None:
    required = ["pH", "water_kg", "Li", "K", "Mg", "B", "Ca", "Na", "Cl", "HCO3", "S6"]
    for key in required:
        if key not in feed:
            raise ValueError(f"Falta '{key}' en feed_base.txt")

    checks = {
        "pH": (0, 14),
        "water_kg": (1e-9, 1e9),
        "Li": (0, 50),
        "K": (0, 50),
        "Mg": (0, 50),
        "B": (0, 50),
        "Ca": (0, 50),
        "Na": (0, 50),
        "Cl": (0, 50),
        "HCO3": (0, 50),
        "S6": (0, 50),
    }

    for key, (vmin, vmax) in checks.items():
        val = float(feed[key])
        if not (vmin <= val <= vmax):
            raise ValueError(f"Valor fuera de rango en feed_base.txt: {key}={val}")


def validate_stages(stages: pd.DataFrame) -> None:
    required_cols = ["stage_id", "stage_type", "surface_m2", "reagent", "reagent_param", "phases"]
    for col in required_cols:
        if col not in stages.columns:
            raise ValueError(f"Falta la columna '{col}' en stages_base.txt")

    allowed_types = {"pond", "reactor"}
    bad_type_mask = ~stages["stage_type"].astype(str).str.lower().isin(allowed_types)
    if bad_type_mask.any():
        bad = stages.loc[bad_type_mask, "stage_type"].tolist()
        raise ValueError(f"Hay stage_type no validos en stages_base.txt: {bad}")

    if stages["surface_m2"].isna().any():
        raise ValueError("Hay superficies no numericas en stages_base.txt")
    if (stages["surface_m2"] < 0).any():
        raise ValueError("Hay superficies negativas en stages_base.txt")


def validate_months_control(months: pd.DataFrame) -> None:
    if months["month"].duplicated().any():
        dup = months.loc[months["month"].duplicated(), "month"].tolist()
        raise ValueError(f"Hay meses duplicados en months_control.txt: {dup}")
    missing = set(range(1, 13)) - set(months["month"].tolist())
    if missing:
        raise ValueError(f"Faltan meses en months_control.txt: {sorted(missing)}")


def validate_scenarios_control(df: pd.DataFrame) -> None:
    required_cols = ["variable", "enabled", "min", "max", "step"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Falta la columna '{col}' en scenarios_control.txt")

    expected_vars = {"months", "temperature", "evap_factor", "mg_removal", "retention_r"}
    found_vars = set(df["variable"].astype(str).str.lower().tolist())
    missing = expected_vars - found_vars
    if missing:
        raise ValueError(f"Faltan variables en scenarios_control.txt: {sorted(missing)}")

    for _, row in df.iterrows():
        var = str(row["variable"]).lower()
        min_v = float(row["min"])
        max_v = float(row["max"])
        step_v = float(row["step"])

        if max_v < min_v:
            raise ValueError(f"En scenarios_control.txt, {var}: max < min")
        if step_v <= 0:
            raise ValueError(f"En scenarios_control.txt, {var}: step debe ser > 0")

        if var == "months" and not (1 <= min_v <= 12 and 1 <= max_v <= 12):
            raise ValueError("Rango de months no razonable.")
        if var == "temperature" and not (-20 <= min_v <= 120 and -20 <= max_v <= 120):
            raise ValueError("Rango de temperature no razonable.")
        if var == "evap_factor" and not (0 < min_v <= 10 and 0 < max_v <= 10):
            raise ValueError("Rango de evap_factor no razonable.")
        if var == "mg_removal" and not (0 <= min_v <= 100 and 0 <= max_v <= 100):
            raise ValueError("Rango de mg_removal no razonable.")
        if var == "retention_r" and not (0 <= min_v <= 100 and 0 <= max_v <= 100):
            raise ValueError("Rango de retention_r no razonable.")


# =========================================================
# UTILIDADES
# =========================================================

def build_run_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def build_range(min_val: float, max_val: float, step: float) -> list[float]:
    values = []
    x = min_val
    n_guard = 0

    while x <= max_val + 1e-12 and n_guard < 10000:
        values.append(round(x, 10))
        x += step
        n_guard += 1

    if not values:
        values = [round(min_val, 10)]

    if abs(values[-1] - max_val) > 1e-9:
        values.append(round(max_val, 10))

    return sorted(set(values))


def normalize_fraction(value: float) -> float:
    value = float(value)
    if value > 1.0:
        return value / 100.0
    return value


def evap_l_s_to_h2o_mol(evap_l_s: float) -> float:
    return evap_l_s * RHO_WATER_KG_L / MW_H2O_KG_MOL


def parse_stage_phases(phases_text: str) -> list[str]:
    phases = []
    for p in str(phases_text).split("|"):
        p = p.strip().strip('"')
        if p and p.lower() != "nan":
            phase = normalize_stage_phase_name(p)
            if phase is not None and phase not in phases:
                phases.append(phase)
    return phases


def raw_stage_phases(phases_text: str) -> list[str]:
    phases = []
    for p in str(phases_text).split("|"):
        p = p.strip().strip('"')
        if p and p.lower() != "nan":
            phases.append(p)
    return phases


def database_phase_names() -> set[str]:
    global _DATABASE_PHASE_NAMES_CACHE
    if _DATABASE_PHASE_NAMES_CACHE is not None:
        return _DATABASE_PHASE_NAMES_CACHE

    phase_names = set()
    if DATABASE is None or not DATABASE.exists():
        _DATABASE_PHASE_NAMES_CACHE = phase_names
        return phase_names

    in_phases = False
    section_headers = {
        "END",
        "EXCHANGE_MASTER_SPECIES",
        "EXCHANGE_SPECIES",
        "KNOBS",
        "PITZER",
        "RATES",
        "SELECTED_OUTPUT",
        "SOLUTION_MASTER_SPECIES",
        "SOLUTION_SPECIES",
        "SURFACE_MASTER_SPECIES",
        "SURFACE_SPECIES",
    }
    for raw_line in DATABASE.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.upper() == "PHASES":
            in_phases = True
            continue
        if not in_phases:
            continue
        if raw_line[0].isspace():
            continue
        first_token = line.split()[0]
        if first_token.upper() in section_headers:
            break
        phase_names.add(first_token)

    _DATABASE_PHASE_NAMES_CACHE = phase_names
    return phase_names


def normalize_stage_phase_name(phase: str) -> str | None:
    phase = PHREEQC_PHASE_ALIASES.get(phase, phase)
    names = database_phase_names()
    if names and phase not in names:
        return None
    return phase


def stage_phase_warning(phases_text: str) -> str:
    notes = []
    names = database_phase_names()
    for phase in raw_stage_phases(phases_text):
        mapped = PHREEQC_PHASE_ALIASES.get(phase, phase)
        if mapped != phase:
            notes.append(f"{phase} se uso como {mapped}")
        if names and mapped not in names:
            database_name = DATABASE.name if DATABASE is not None else "PHREEQC database"
            notes.append(f"{phase} omitida: no existe en {database_name}")
    return " | ".join(dict.fromkeys(notes))


def safe_ratio(num, den):
    if num is None or den is None or pd.isna(num) or pd.isna(den) or den == 0:
        return None
    return num / den


def extract_last_value(text: str, label: str):
    pattern = rf"{re.escape(label)}\s*=\s*([\-+0-9.Ee]+)"
    matches = re.findall(pattern, text)
    if not matches:
        return None
    try:
        return float(matches[-1])
    except Exception:
        return None


def parse_pqo_properties(output_file: Path) -> dict:
    text = output_file.read_text(encoding="utf-8", errors="ignore")
    density = extract_last_value(text, "Density (g/cm3)")
    volume = extract_last_value(text, "Volume (L)")
    mass_water = extract_last_value(text, "Mass of water (kg)")
    water_activity = extract_last_value(text, "Activity of water")
    return {
        "density_g_cm3": density,
        "volume_L": volume,
        "mass_H2O_kg": mass_water,
        "water_activity": water_activity,
    }


def compact_phreeqc_message(text: str) -> str:
    if not text:
        return ""

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    important = []
    patterns = [
        "ERROR:",
        "WARNING: The program has failed",
        "Maximum iterations exceeded",
        "Numerical method failed",
        "failed to converge",
    ]

    for line in lines:
        if any(pattern in line for pattern in patterns):
            important.append(line)

    if not important:
        important = lines[-8:]

    compact = []
    for line in important:
        if line not in compact:
            compact.append(line)

    return " | ".join(compact[-8:])


def get_solution_molality(solution: dict, species: str) -> float:
    return float(solution.get(species, 0.0))


def get_row_molality(row: pd.Series, species: str) -> float:
    if species in row.index and pd.notna(row[species]):
        return float(row[species])
    if species == "HCO3":
        for candidate in ["C(4)", "C4", "C", "alkalinity", "alk", "Alk", "ALK", "HCO3-"]:
            if candidate in row.index and pd.notna(row[candidate]):
                return float(row[candidate])
    raise ValueError(f"No se pudo leer la molalidad de {species} en selected_output.")


def estimate_liquid_mass_kg(mass_H2O_kg: float, molalities: dict) -> float:
    total_kg = float(mass_H2O_kg)
    for sp in OUTPUT_SPECIES:
        if sp not in molalities:
            continue
        moles = float(molalities[sp]) * float(mass_H2O_kg)
        total_kg += moles * ELEMENT_MW_G_MOL[sp] / 1000.0
    return total_kg


def estimate_tds_mg_l(mass_H2O_kg: float, liquid_mass_kg: float, volume_l: float) -> float | None:
    if volume_l is None or pd.isna(volume_l) or volume_l <= 0:
        return None
    solutes_kg = max(float(liquid_mass_kg) - float(mass_H2O_kg), 0.0)
    return solutes_kg * 1_000_000.0 / float(volume_l)


def major_ion_tds_g_l(row: pd.Series | dict) -> float | None:
    total_mg_l = 0.0
    any_value = False

    for sp in ["Li", "K", "Mg", "B", "Ca", "Na", "Cl", "HCO3"]:
        key = f"{sp}_mg_L_est"
        value = row.get(key, None)
        if value is not None and pd.notna(value):
            total_mg_l += float(value)
            any_value = True

    so4_value = sulfate_mg_l(row)
    if so4_value is not None and pd.notna(so4_value):
        total_mg_l += float(so4_value)
        any_value = True

    return total_mg_l / 1000.0 if any_value else None


def sulfate_mg_l(row: pd.Series | dict) -> float | None:
    s6_molality = row.get("S6", row.get("S6_molality_out", None))
    water_kg = row.get("mass_H2O_out_kg", None)
    volume_l = row.get("volume_out_L", None)

    if (
        s6_molality is None or pd.isna(s6_molality) or
        water_kg is None or pd.isna(water_kg) or
        volume_l is None or pd.isna(volume_l) or float(volume_l) <= 0
    ):
        return None

    return float(s6_molality) * float(water_kg) * MW_SO4_G_MOL * 1000.0 / float(volume_l)


def build_main_saturation_indexes(react_row: pd.Series) -> str:
    pairs = []
    for col in react_row.index:
        col_text = str(col)
        if not col_text.startswith("si_"):
            continue
        value = react_row.get(col, None)
        if value is None or pd.isna(value):
            continue
        phase = col_text[3:]
        pairs.append((phase, float(value)))

    pairs.sort(key=lambda item: abs(item[1]), reverse=True)
    return "|".join(f"{phase}:{value:.4g}" for phase, value in pairs)


def normalize_output_df(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "S(6)": "S6",
        "C(4)": "HCO3",
        "C4": "HCO3",
        "C": "HCO3",
        "alk": "HCO3",
        "Alk": "HCO3",
        "ALK": "HCO3",
        "alkalinity": "HCO3",
        "HCO3-": "HCO3",
        "mass_H2O": "mass_H2O",
    }
    return df.rename(columns={c: rename_map.get(str(c), c) for c in df.columns})


def as_posix_or_none(path: Path | None) -> str | None:
    if path is None:
        return None
    return str(path)


# =========================================================
# ESCENARIOS Y EVAPORACION
# =========================================================

def get_control_values(df: pd.DataFrame, variable: str) -> list[float]:
    row = df[df["variable"] == variable]
    if row.empty:
        return [DEFAULT_SCENARIO_VALUES[variable]]

    row = row.iloc[0]
    enabled = int(float(row["enabled"])) if pd.notna(row["enabled"]) else 0
    if enabled != 1:
        return [DEFAULT_SCENARIO_VALUES[variable]]

    values = build_range(
        float(row["min"]),
        float(row["max"]),
        float(row["step"]) if pd.notna(row["step"]) else DEFAULT_STEPS[variable],
    )

    if variable == "months":
        return [float(int(round(v))) for v in values]
    if variable == "mg_removal":
        return [normalize_fraction(v) for v in values]
    return values


def expand_scenarios_control(df: pd.DataFrame, months_control: pd.DataFrame) -> pd.DataFrame:
    values_map = {
        "months": get_control_values(df, "months"),
        "temperature": get_control_values(df, "temperature"),
        "evap_factor": get_control_values(df, "evap_factor"),
        "mg_removal": get_control_values(df, "mg_removal"),
        "retention_r": get_control_values(df, "retention_r"),
    }

    rows = []
    available_months = set(months_control["month"].astype(int).tolist())

    for month, temp, evap_factor, mg_removal, retention_r in product(
        values_map["months"],
        values_map["temperature"],
        values_map["evap_factor"],
        values_map["mg_removal"],
        values_map["retention_r"],
    ):
        month = int(round(month))
        if month not in available_months:
            raise ValueError(f"El mes {month} esta en scenarios_control.txt, pero no en months_control.txt")

        month_row = months_control.loc[months_control["month"] == month].iloc[0]
        scenario_id = (
            f"M{month:02d}"
            f"_T{int(round(float(temp))):03d}"
            f"_E{int(round(float(evap_factor) * 100)):03d}"
            f"_MG{int(round(float(mg_removal) * 100)):03d}"
            f"_R{int(round(float(retention_r) * 100)):03d}"
        )

        rows.append({
            "scenario_id": scenario_id,
            "enabled": 1,
            "month": month,
            "month_label": month_row["month_label"],
            "month_f_m": float(month_row["f_m"]),
            "bypass_velocity_factor": float(month_row["f_m"]),
            "temperature": float(temp),
            "evap_factor": float(evap_factor),
            "mg_removal": float(mg_removal),
            "retention_r": float(retention_r),
        })

    return pd.DataFrame(rows)


def get_stage_group(stage_id: str) -> str | None:
    sid = str(stage_id).upper().strip()
    if sid.startswith("LIM"):
        return None
    for prefix in ["PC", "H", "K", "C", "L"]:
        if sid.startswith(prefix):
            return prefix
    return None


def get_evap_mm_d(month_row: pd.Series, stage_id: str, stage_type: str) -> tuple[str | None, float]:
    if str(stage_type).lower() != "pond":
        return None, 0.0

    group = get_stage_group(stage_id)
    if group is None:
        return None, 0.0

    col = f"TP_{group}"
    if col not in month_row.index:
        raise ValueError(f"No existe '{col}' para calcular evaporacion de {stage_id}.")

    return group, float(month_row[col])


def apply_scenario(
    feed: dict,
    stages: pd.DataFrame,
    months_control: pd.DataFrame,
    scenario: pd.Series,
) -> tuple[dict, pd.DataFrame]:
    feed_mod = feed.copy()
    stages_mod = stages.copy()

    month = int(scenario["month"])
    month_row = months_control.loc[months_control["month"] == month].iloc[0]
    base_evaporation_row = get_base_evaporation_row(months_control)
    evap_factor = float(scenario["evap_factor"])
    original_water_kg = float(feed_mod["water_kg"])

    feed_mod["temp_C"] = float(scenario["temperature"])
    feed_mod["original_water_kg"] = original_water_kg

    stages_mod["stage_type"] = stages_mod["stage_type"].astype(str).str.lower()
    stages_mod["month"] = month
    stages_mod["month_label"] = scenario["month_label"]
    stages_mod["month_f_m"] = float(scenario.get("month_f_m", month_row["f_m"]))
    stages_mod["bypass_velocity_factor"] = float(scenario.get("bypass_velocity_factor", month_row["f_m"]))
    stages_mod["temperature"] = float(scenario["temperature"])
    stages_mod["temp_C"] = float(scenario["temperature"])
    stages_mod["evap_factor"] = evap_factor
    stages_mod["mg_removal"] = float(scenario["mg_removal"])
    stages_mod["retention_r"] = float(scenario["retention_r"])
    stages_mod["evap_group"] = None
    stages_mod["E_mm_d"] = 0.0
    stages_mod["Q_evap"] = 0.0
    stages_mod["Q_evap_kg_s"] = 0.0
    stages_mod["reagent_moles_calc"] = 0.0

    for idx, row in stages_mod.iterrows():
        group, e_mm_d_base = get_evap_mm_d(base_evaporation_row, row["stage_id"], row["stage_type"])
        e_mm_d = e_mm_d_base * evap_factor
        q_evap_l_s = e_mm_d * float(row["surface_m2"]) / 86400.0

        stages_mod.at[idx, "evap_group"] = group
        stages_mod.at[idx, "E_mm_d"] = e_mm_d
        stages_mod.at[idx, "Q_evap"] = q_evap_l_s
        stages_mod.at[idx, "Q_evap_kg_s"] = q_evap_l_s * RHO_WATER_KG_L

    design_total_evap_l_s = float(stages_mod["Q_evap"].sum())
    feed_mod["design_total_evap_L_s"] = design_total_evap_l_s

    for col in [
        "design_total_evap_L_s",
    ]:
        stages_mod[col] = feed_mod[col]

    return feed_mod, stages_mod


# =========================================================
# CONSTRUCCION DEL INPUT PHREEQC
# =========================================================

def build_selected_output(stage_phases: list[str], selected_filename: str) -> str:
    phases_text = " ".join(stage_phases)
    phase_lines = ""
    if phases_text:
        phase_lines = (
            f"    -equilibrium_phases       {phases_text}\n"
            f"    -saturation_indices       {phases_text}\n"
        )

    return f"""SELECTED_OUTPUT
    -file                     {selected_filename}
    -reset                    false
    -high_precision           true
    -simulation               true
    -state                    true
    -solution                 true
    -pH                       true
    -ionic_strength           true
    -water                    true
    -charge_balance           true
    -percent_error            true
    -totals                   Li K Mg B Ca Na Cl C(4) S(6)
{phase_lines}"""


def build_solution_block(solution_id: int, solution_dict: dict, title: str) -> str:
    return f"""TITLE {title}

SOLUTION {solution_id} {title}
    temp      {solution_dict['temp_C']}
    pH        {solution_dict['pH']}
    units     mol/kgw
    water     {solution_dict['water_kg']}
    Li        {solution_dict['Li']}
    K         {solution_dict['K']}
    Mg        {solution_dict['Mg']}
    B         {solution_dict['B']}
    Ca        {solution_dict['Ca']}
    Na        {solution_dict['Na']}
    Cl        {solution_dict['Cl']}
    C(4)      {solution_dict['HCO3']}
    S(6)      {solution_dict['S6']}
END
"""


def build_stage_block(prev_solution_id: int, stage_solution_id: int, stage_row: pd.Series) -> str:
    stage_id = stage_row["stage_id"]
    stage_type = str(stage_row["stage_type"]).lower()
    reagent = "" if pd.isna(stage_row["reagent"]) else str(stage_row["reagent"]).strip()

    phases = parse_stage_phases(stage_row["phases"])
    phases_block = "\n".join([f"    {phase}      0.0 0.0" for phase in phases])

    common_start = f"""TITLE {stage_id}

USE solution {prev_solution_id}

EQUILIBRIUM_PHASES {stage_solution_id}
{phases_block}
"""

    if stage_type == "pond":
        h2o_mol = evap_l_s_to_h2o_mol(float(stage_row["Q_evap"]))
        reaction_steps = int(stage_row.get("reaction_steps", 1))
        return f"""{common_start}
REACTION {stage_solution_id}
    H2O   -{h2o_mol:.10f}
    1 in {reaction_steps} steps

SAVE solution {stage_solution_id}
END
"""

    if stage_type == "reactor":
        reagent_moles_calc = float(stage_row.get("reagent_moles_calc", 0.0))
        if reagent and reagent_moles_calc > 0.0:
            reaction = f"""
REACTION {stage_solution_id}
    {reagent} 1
    {reagent_moles_calc:.10f}
"""
        else:
            reaction = ""

        return f"""{common_start}{reaction}
SAVE solution {stage_solution_id}
END
"""

    raise ValueError(f"Tipo de etapa no reconocido: {stage_type}")


def build_stage_input_text(solution_in: dict, stage_row: pd.Series, selected_filename: str) -> str:
    stage_phases = parse_stage_phases(stage_row["phases"])
    parts = [
        build_selected_output(stage_phases, selected_filename),
        build_solution_block(1, solution_in, "CURRENT_IN"),
        build_stage_block(1, 2, stage_row),
    ]
    return "\n\n".join(parts)


# =========================================================
# EJECUCION Y PARSE PHREEQC
# =========================================================

def parse_selected_output(selected_path: Path) -> pd.DataFrame:
    return normalize_output_df(pd.read_csv(selected_path, sep=r"\s+", engine="python"))


def extract_react_row(raw_df: pd.DataFrame) -> pd.Series:
    df = normalize_output_df(raw_df.copy())
    if "state" in df.columns:
        react_df = df[df["state"].astype(str).str.lower() == "react"]
        if not react_df.empty:
            return react_df.iloc[-1]
    return df.iloc[-1]


def run_phreeqc(input_file: Path, output_file: Path, cwd: Path) -> subprocess.CompletedProcess:
    cmd = [str(PHREEQC_EXE), str(input_file), str(output_file), str(DATABASE)]
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=cwd)

    if proc.returncode != 0:
        compact_stderr = compact_phreeqc_message(proc.stderr)
        compact_stdout = compact_phreeqc_message(proc.stdout)
        details = "\n".join([
            f"PHREEQC fallo con codigo {proc.returncode}",
            f"input={input_file}",
            f"output={output_file}",
            f"stdout={compact_stdout}",
            f"stderr={compact_stderr}",
        ])
        raise RuntimeError(details)

    return proc


def execute_stage_phreeqc(
    stage_row: pd.Series,
    solution_in: dict,
    stage_folder: Path,
    input_name: str,
    output_name: str,
    selected_name: str,
) -> tuple[pd.Series, dict, Path, Path, Path]:
    input_file = stage_folder / input_name
    output_file = stage_folder / output_name
    selected_file = stage_folder / selected_name

    input_text = build_stage_input_text(solution_in, stage_row, selected_file.name)
    input_file.write_text(input_text, encoding="utf-8")

    run_phreeqc(input_file, output_file, stage_folder)

    if not selected_file.exists():
        raise FileNotFoundError(f"No se encontro {selected_file}")

    raw_df = parse_selected_output(selected_file)
    react_row = extract_react_row(raw_df)
    pqo_props = parse_pqo_properties(output_file)

    return react_row, pqo_props, input_file, output_file, selected_file


# =========================================================
# RESULTADOS POR ETAPA
# =========================================================

def build_stage_phase_summary(
    react_row: pd.Series,
    stage_row: pd.Series,
    stage_order: int,
    scenario_id: str,
) -> pd.DataFrame:
    stage_id = stage_row["stage_id"]
    stage_type = stage_row["stage_type"]
    phases = parse_stage_phases(stage_row["phases"])

    rows = []
    for phase in phases:
        if phase not in react_row.index:
            continue

        value = react_row.get(phase, None)
        if pd.isna(value):
            continue

        phase_moles = float(value)
        if phase_moles <= 0:
            continue

        phase_mass_kg = None
        if phase in PHASE_MW_KG_MOL:
            phase_mass_kg = phase_moles * PHASE_MW_KG_MOL[phase]

        rows.append({
            "scenario_id": scenario_id,
            "stage_order": stage_order,
            "stage_id": stage_id,
            "stage_type": stage_type,
            "phase_name": phase,
            "phase_moles": phase_moles,
            "phase_mass_kg": phase_mass_kg,
        })

    return pd.DataFrame(rows)


def solution_molalities(solution: dict) -> dict:
    return {sp: get_solution_molality(solution, sp) for sp in OUTPUT_SPECIES}


def row_molalities(react_row: pd.Series) -> dict:
    return {sp: get_row_molality(react_row, sp) for sp in OUTPUT_SPECIES}


def build_next_solution_after_retention(
    react_row: pd.Series,
    current_temp: float,
    retained_fraction_liquid: float,
) -> dict:
    water_out_kg = float(react_row["mass_H2O"])
    water_to_next_kg = water_out_kg * (1.0 - retained_fraction_liquid)

    if water_to_next_kg <= 1e-12:
        raise ValueError(f"La masa de agua que pasa a la siguiente etapa es <= 0 ({water_to_next_kg:.6e} kg).")

    next_solution = {
        "temp_C": float(current_temp),
        "pH": float(react_row["pH"]),
        "water_kg": water_to_next_kg,
    }

    for sp in OUTPUT_SPECIES:
        next_solution[sp] = get_row_molality(react_row, sp)

    return next_solution


def build_species_outputs(row: dict, molalities: dict, water_kg: float, volume_l: float) -> None:
    for sp in OUTPUT_SPECIES:
        molality = float(molalities[sp])
        moles = molality * water_kg
        row[sp] = molality
        row[f"{sp}_molality_out"] = molality
        row[f"{sp}_mol_out_est"] = moles
        if volume_l and volume_l > 0:
            row[f"{sp}_mg_L_est"] = moles * ELEMENT_MW_G_MOL[sp] * 1000.0 / volume_l
        else:
            row[f"{sp}_mg_L_est"] = None

    row["SO4/S(6)"] = row["S6"]
    row["Mg/Li"] = safe_ratio(row["Mg"], row["Li"])
    row["K/Li"] = safe_ratio(row["K"], row["Li"])
    row["SO4/Li"] = safe_ratio(row["S6"], row["Li"])
    row["Ca/Li"] = safe_ratio(row["Ca"], row["Li"])


def build_feed_row(solution_dict: dict, scenario_id: str) -> dict:
    water_kg = float(solution_dict["water_kg"])
    molalities = solution_molalities(solution_dict)
    liquid_mass_kg = estimate_liquid_mass_kg(water_kg, molalities)
    volume_l = solution_dict.get("feed_phreeqc_volume_L_s")
    if volume_l is None or pd.isna(volume_l) or float(volume_l) <= 0:
        volume_l = liquid_mass_kg
        feed_density_kg_l = None
        volume_basis = "estimated_liquid_mass"
    else:
        volume_l = float(volume_l)
        feed_density_kg_l = liquid_mass_kg / volume_l if volume_l > 0 else None
        volume_basis = "phreeqc_volume_check"
    tds_mg_l = estimate_tds_mg_l(water_kg, liquid_mass_kg, volume_l)

    row = {
        "scenario_id": scenario_id,
        "stage_order": 0,
        "stage_id": "FEED",
        "stage_type": "feed",
        "temp_C": float(solution_dict["temp_C"]),
        "pH_out": float(solution_dict["pH"]),
        "Q_in": None,
        "Q_evap": 0.0,
        "Q_evap_external_L_s": 0.0,
        "evap_recovered_by_module_L_s": 0.0,
        "evap_reduction_percent": 0.0,
        "Q_module_recovered_water_L_s": 0.0,
        "Q_module_extra_recovered_water_L_s": 0.0,
        "Q_evap_kg_s": 0.0,
        "Q_retained": 0.0,
        "Q_out": liquid_mass_kg,
        "original_water_kg": solution_dict.get("original_water_kg"),
        "design_total_evap_L_s": solution_dict.get("design_total_evap_L_s"),
        "E_mm_d": 0.0,
        "surface_m2": 0.0,
        "density_out_kg_L": feed_density_kg_l,
        "water_activity": None,
        "main_saturation_indexes": "",
        "volume_out_L": volume_l,
        "feed_volume_basis": volume_basis,
        "feed_phreeqc_volume_L_s": solution_dict.get("feed_phreeqc_volume_L_s"),
        "mass_H2O_out_kg": water_kg,
        "liquid_mass_out_kg": liquid_mass_kg,
        "TDS": tds_mg_l,
        "retention_r": 0.0,
        "phases_precipitated": "",
        "phase_masses_kg": "",
        "total_precipitated_mass_kg": 0.0,
        "cumulative_solids_kg": 0.0,
        "retained_brine_kg": 0.0,
        "salmuera_retenida": 0.0,
        "retained_fraction_liquid": 0.0,
        "retention_clamped": 0,
        "mass_H2O_to_next_kg": water_kg,
        "liquid_to_next_kg": liquid_mass_kg,
        "phreeqc_input_file": None,
        "phreeqc_output_file": None,
        "selected_output_file": None,
        "error": "",
        "warning": "",
    }
    build_species_outputs(row, molalities, water_kg, volume_l)
    for sp in OUTPUT_SPECIES:
        row[f"{sp}_mol_to_next_est"] = row[f"{sp}_mol_out_est"]
    return row


def build_stage_clean_row(
    react_row: pd.Series,
    stage_row: pd.Series,
    stage_order: int,
    scenario_id: str,
    pqo_props: dict,
    phase_df: pd.DataFrame,
    solution_in: dict,
    input_file: Path,
    output_file: Path,
    selected_file: Path,
) -> tuple[dict, dict]:
    stage_id = stage_row["stage_id"]
    stage_type = stage_row["stage_type"]
    retention_r = float(stage_row.get("retention_r", 0.0))

    mass_H2O_out_kg = float(react_row["mass_H2O"])
    molalities = row_molalities(react_row)
    density_out_kg_L = pqo_props.get("density_g_cm3")
    volume_out_l = pqo_props.get("volume_L")

    liquid_mass_out_kg = estimate_liquid_mass_kg(mass_H2O_out_kg, molalities)
    if volume_out_l is None or pd.isna(volume_out_l) or float(volume_out_l) <= 0:
        if density_out_kg_L is not None and pd.notna(density_out_kg_L) and float(density_out_kg_L) > 0:
            volume_out_l = liquid_mass_out_kg / float(density_out_kg_L)
        else:
            volume_out_l = liquid_mass_out_kg

    input_liquid_mass_kg = estimate_liquid_mass_kg(float(solution_in["water_kg"]), solution_molalities(solution_in))
    tds_mg_l = estimate_tds_mg_l(mass_H2O_out_kg, liquid_mass_out_kg, float(volume_out_l))

    solids_precip_kg = 0.0
    phase_names = []
    phase_mass_pairs = []
    if not phase_df.empty:
        if "phase_mass_kg" in phase_df.columns:
            solids_precip_kg = float(phase_df["phase_mass_kg"].fillna(0.0).sum())
        phase_names = phase_df["phase_name"].astype(str).tolist()
        for _, phase_row in phase_df.iterrows():
            mass = phase_row.get("phase_mass_kg", None)
            mass_text = "" if mass is None or pd.isna(mass) else f"{float(mass):.10g}"
            phase_mass_pairs.append(f"{phase_row['phase_name']}:{mass_text}")

    retained_brine_kg_raw = retention_r * solids_precip_kg
    max_retained_kg = max(liquid_mass_out_kg * 0.999999, 0.0)
    retained_brine_kg = min(retained_brine_kg_raw, max_retained_kg)
    retention_clamped = int(retained_brine_kg_raw > max_retained_kg)
    retained_fraction_liquid = retained_brine_kg / liquid_mass_out_kg if liquid_mass_out_kg > 0 else 0.0

    mass_H2O_to_next_kg = mass_H2O_out_kg * (1.0 - retained_fraction_liquid)
    liquid_to_next_kg = liquid_mass_out_kg * (1.0 - retained_fraction_liquid)

    row = {
        "scenario_id": scenario_id,
        "stage_order": stage_order,
        "stage_id": stage_id,
        "stage_type": stage_type,
        "temp_C": float(stage_row.get("temp_C", stage_row.get("temperature", 25.0))),
        "pH_out": float(react_row["pH"]),
        "mu_out": float(react_row["mu"]) if "mu" in react_row.index and pd.notna(react_row["mu"]) else None,
        "Q_in": input_liquid_mass_kg,
        "Q_evap": float(stage_row.get("Q_evap", 0.0)),
        "Q_evap_original_L_s": float(stage_row.get("Q_evap_original_L_s", stage_row.get("Q_evap", 0.0))),
        "Q_evap_external_L_s": float(stage_row.get("Q_evap_external_L_s", 0.0)),
        "evap_recovered_by_module_L_s": float(stage_row.get("evap_recovered_by_module_L_s", 0.0)),
        "evap_reduction_percent": stage_row.get("evap_reduction_percent", None),
        "Q_module_recovered_water_L_s": float(stage_row.get("Q_module_recovered_water_L_s", 0.0)),
        "Q_module_extra_recovered_water_L_s": float(stage_row.get("Q_module_extra_recovered_water_L_s", 0.0)),
        "Q_evap_kg_s": float(stage_row.get("Q_evap_kg_s", 0.0)),
        "recovery_evap_scale_factor": float(stage_row.get("recovery_evap_scale_factor", 1.0)),
        "bypass_fraction_percent": stage_row.get("bypass_fraction_percent", None),
        "Q_retained": retained_brine_kg,
        "Q_out": liquid_to_next_kg,
        "design_total_evap_L_s": stage_row.get("design_total_evap_L_s", None),
        "E_mm_d": float(stage_row.get("E_mm_d", 0.0)),
        "surface_m2": float(stage_row.get("surface_m2", 0.0)),
        "evap_group": stage_row.get("evap_group", None),
        "reagent": stage_row.get("reagent", ""),
        "reagent_param": float(stage_row.get("reagent_param", 0.0)),
        "lime_dose_mol": float(stage_row.get("reagent_moles_calc", 0.0)),
        "mg_removal_target": float(stage_row.get("mg_removal", 0.0)),
        "mg_removal_achieved": stage_row.get("mg_removal_achieved", None),
        "liming_iterations": stage_row.get("liming_iterations", None),
        "density_out_kg_L": density_out_kg_L,
        "water_activity": pqo_props.get("water_activity"),
        "main_saturation_indexes": build_main_saturation_indexes(react_row),
        "volume_out_L": volume_out_l,
        "mass_H2O_out_kg": mass_H2O_out_kg,
        "liquid_mass_out_kg": liquid_mass_out_kg,
        "TDS": tds_mg_l,
        "retention_r": retention_r,
        "phases_precipitated": "|".join(phase_names),
        "phase_masses_kg": ";".join(phase_mass_pairs),
        "total_precipitated_mass_kg": solids_precip_kg,
        "retained_brine_kg": retained_brine_kg,
        "salmuera_retenida": retained_brine_kg,
        "retained_fraction_liquid": retained_fraction_liquid,
        "retention_clamped": retention_clamped,
        "mass_H2O_to_next_kg": mass_H2O_to_next_kg,
        "liquid_to_next_kg": liquid_to_next_kg,
        "phreeqc_input_file": as_posix_or_none(input_file),
        "phreeqc_output_file": as_posix_or_none(output_file),
        "selected_output_file": as_posix_or_none(selected_file),
        "error": "",
        "warning": " | ".join(
            part for part in [
                str(stage_row.get("warning", "") or "").strip(),
                stage_phase_warning(stage_row.get("phases", "")),
            ]
            if part
        ),
    }
    build_species_outputs(row, molalities, mass_H2O_out_kg, float(volume_out_l))

    for sp in OUTPUT_SPECIES:
        row[f"{sp}_mol_to_next_est"] = row[f"{sp}_mol_out_est"] * (1.0 - retained_fraction_liquid)

    next_solution = build_next_solution_after_retention(
        react_row=react_row,
        current_temp=float(stage_row.get("temp_C", stage_row.get("temperature", 25.0))),
        retained_fraction_liquid=retained_fraction_liquid,
    )

    return row, next_solution


def build_error_row(
    stage_order: int,
    stage_row: pd.Series,
    scenario_id: str,
    solution_in: dict,
    error: Exception,
) -> dict:
    water_kg = float(solution_in["water_kg"])
    molalities = solution_molalities(solution_in)
    input_liquid_mass_kg = estimate_liquid_mass_kg(water_kg, molalities)
    tds_mg_l = estimate_tds_mg_l(water_kg, input_liquid_mass_kg, input_liquid_mass_kg)
    row = {
        "scenario_id": scenario_id,
        "stage_order": stage_order,
        "stage_id": stage_row.get("stage_id", None),
        "stage_type": stage_row.get("stage_type", None),
        "temp_C": float(stage_row.get("temp_C", stage_row.get("temperature", 25.0))),
        "pH_out": float(solution_in.get("pH", 0.0)),
        "Q_in": input_liquid_mass_kg,
        "Q_evap": float(stage_row.get("Q_evap", 0.0)),
        "Q_evap_original_L_s": float(stage_row.get("Q_evap_original_L_s", stage_row.get("Q_evap", 0.0))),
        "Q_evap_external_L_s": float(stage_row.get("Q_evap_external_L_s", 0.0)),
        "evap_recovered_by_module_L_s": float(stage_row.get("evap_recovered_by_module_L_s", 0.0)),
        "evap_reduction_percent": stage_row.get("evap_reduction_percent", None),
        "Q_module_recovered_water_L_s": float(stage_row.get("Q_module_recovered_water_L_s", 0.0)),
        "Q_module_extra_recovered_water_L_s": float(stage_row.get("Q_module_extra_recovered_water_L_s", 0.0)),
        "Q_evap_kg_s": float(stage_row.get("Q_evap_kg_s", 0.0)),
        "recovery_evap_scale_factor": float(stage_row.get("recovery_evap_scale_factor", 1.0)),
        "bypass_fraction_percent": stage_row.get("bypass_fraction_percent", None),
        "Q_retained": None,
        "Q_out": None,
        "design_total_evap_L_s": stage_row.get("design_total_evap_L_s", None),
        "E_mm_d": float(stage_row.get("E_mm_d", 0.0)),
        "surface_m2": float(stage_row.get("surface_m2", 0.0)),
        "volume_out_L": input_liquid_mass_kg,
        "density_out_kg_L": None,
        "water_activity": None,
        "main_saturation_indexes": "",
        "mass_H2O_out_kg": water_kg,
        "liquid_mass_out_kg": input_liquid_mass_kg,
        "TDS": tds_mg_l,
        "retention_r": float(stage_row.get("retention_r", 0.0)),
        "error": str(error),
        "warning": "",
    }
    build_species_outputs(row, molalities, water_kg, input_liquid_mass_kg)
    for sp in OUTPUT_SPECIES:
        row[f"{sp}_mol_to_next_est"] = row[f"{sp}_mol_out_est"]
    return row


# =========================================================
# EJECUCION DE ETAPAS
# =========================================================

def prepare_stage_row_for_run(stage_row: pd.Series, solution_in: dict) -> pd.Series:
    stage_row = stage_row.copy()
    stage_row["temp_C"] = solution_in["temp_C"]

    if str(stage_row["stage_type"]).lower() == "pond":
        evap_kg_s = float(stage_row.get("Q_evap_kg_s", 0.0))
        water_in = float(solution_in["water_kg"])
        if evap_kg_s >= water_in:
            raise ValueError(
                f"Evaporacion mayor o igual que el agua disponible en {stage_row['stage_id']}: "
                f"Q_evap_kg_s={evap_kg_s:.6g}, water_in_kg_s={water_in:.6g}"
            )
        evap_fraction = evap_kg_s / water_in if water_in > 0 else 0.0
        reaction_steps = max(1, int(math.ceil(evap_fraction / MAX_EVAP_FRACTION_PER_PHREEQC_STEP)))
        stage_row["reaction_steps"] = reaction_steps

    return stage_row


def run_regular_stage(
    stage_order: int,
    stage_row: pd.Series,
    solution_in: dict,
    scenario_id: str,
    scenario_run_dir: Path,
) -> tuple[pd.DataFrame, dict, dict]:
    stage_id = stage_row["stage_id"]
    stage_folder = scenario_run_dir / f"{stage_order:02d}_{stage_id}"
    stage_folder.mkdir(parents=True, exist_ok=True)

    stage_row = prepare_stage_row_for_run(stage_row, solution_in)
    react_row, pqo_props, input_file, output_file, selected_file = execute_stage_phreeqc(
        stage_row=stage_row,
        solution_in=solution_in,
        stage_folder=stage_folder,
        input_name="input.pqi",
        output_name="output.pqo",
        selected_name="selected_output.txt",
    )

    phase_df = build_stage_phase_summary(react_row, stage_row, stage_order, scenario_id)
    clean_row, next_solution = build_stage_clean_row(
        react_row=react_row,
        stage_row=stage_row,
        stage_order=stage_order,
        scenario_id=scenario_id,
        pqo_props=pqo_props,
        phase_df=phase_df,
        solution_in=solution_in,
        input_file=input_file,
        output_file=output_file,
        selected_file=selected_file,
    )

    return phase_df, clean_row, next_solution


def evaluate_liming_dose(
    dose_mol: float,
    iteration_label: str,
    stage_order: int,
    stage_row: pd.Series,
    solution_in: dict,
    scenario_id: str,
    stage_folder: Path,
) -> dict:
    trial_row = stage_row.copy()
    trial_row["reagent_moles_calc"] = float(dose_mol)
    trial_row = prepare_stage_row_for_run(trial_row, solution_in)

    react_row, pqo_props, input_file, output_file, selected_file = execute_stage_phreeqc(
        stage_row=trial_row,
        solution_in=solution_in,
        stage_folder=stage_folder,
        input_name=f"input_{iteration_label}.pqi",
        output_name=f"output_{iteration_label}.pqo",
        selected_name=f"selected_output_{iteration_label}.txt",
    )

    mg_in_mol = get_solution_molality(solution_in, "Mg") * float(solution_in["water_kg"])
    mg_out_mol = get_row_molality(react_row, "Mg") * float(react_row["mass_H2O"])
    removal = (mg_in_mol - mg_out_mol) / mg_in_mol if mg_in_mol > 0 else 0.0

    phase_df = build_stage_phase_summary(react_row, trial_row, stage_order, scenario_id)

    return {
        "dose_mol": float(dose_mol),
        "removal": float(removal),
        "react_row": react_row,
        "pqo_props": pqo_props,
        "phase_df": phase_df,
        "stage_row": trial_row,
        "input_file": input_file,
        "output_file": output_file,
        "selected_file": selected_file,
    }


def run_liming_stage(
    stage_order: int,
    stage_row: pd.Series,
    solution_in: dict,
    scenario_id: str,
    scenario_run_dir: Path,
) -> tuple[pd.DataFrame, dict, dict]:
    stage_id = stage_row["stage_id"]
    stage_folder = scenario_run_dir / f"{stage_order:02d}_{stage_id}"
    stage_folder.mkdir(parents=True, exist_ok=True)

    target = normalize_fraction(float(stage_row.get("mg_removal", 0.0)))
    mg_in_mol = get_solution_molality(solution_in, "Mg") * float(solution_in["water_kg"])

    if mg_in_mol <= 0 or target <= 0:
        return run_regular_stage(stage_order, stage_row, solution_in, scenario_id, scenario_run_dir)

    evaluations = []
    low = evaluate_liming_dose(0.0, "iter_000_low", stage_order, stage_row, solution_in, scenario_id, stage_folder)
    evaluations.append(low)

    if low["removal"] >= target:
        chosen = low
        warning = "La eliminacion objetivo de Mg se cumple sin Ca(OH)2 adicional."
    else:
        high_dose = max(mg_in_mol * 1.05, 1e-6)
        high = None
        warning = ""

        for i in range(1, 18):
            high = evaluate_liming_dose(
                high_dose,
                f"iter_{i:03d}_high",
                stage_order,
                stage_row,
                solution_in,
                scenario_id,
                stage_folder,
            )
            evaluations.append(high)
            if high["removal"] >= target:
                break
            high_dose *= 2.0

        if high is None:
            raise RuntimeError("No se pudo evaluar la dosis alta de Ca(OH)2.")

        if high["removal"] < target:
            chosen = max(evaluations, key=lambda item: item["removal"])
            warning = (
                f"No se alcanzo mg_removal={target:.6g}; "
                f"mejor={chosen['removal']:.6g} con dosis={chosen['dose_mol']:.6g} mol."
            )
        else:
            low_dose = 0.0
            high_dose = high["dose_mol"]
            best = high

            for i in range(18, 43):
                mid_dose = 0.5 * (low_dose + high_dose)
                mid = evaluate_liming_dose(
                    mid_dose,
                    f"iter_{i:03d}_bis",
                    stage_order,
                    stage_row,
                    solution_in,
                    scenario_id,
                    stage_folder,
                )
                evaluations.append(mid)

                if mid["removal"] >= target:
                    best = mid
                    high_dose = mid_dose
                else:
                    low_dose = mid_dose

            chosen = best

    final_row = stage_row.copy()
    final_row["reagent_moles_calc"] = chosen["dose_mol"]
    final_row["mg_removal_target"] = target
    final_row["mg_removal_achieved"] = chosen["removal"]
    final_row["liming_iterations"] = len(evaluations)
    final_row["warning"] = warning

    final_row = prepare_stage_row_for_run(final_row, solution_in)
    react_row, pqo_props, input_file, output_file, selected_file = execute_stage_phreeqc(
        stage_row=final_row,
        solution_in=solution_in,
        stage_folder=stage_folder,
        input_name="input.pqi",
        output_name="output.pqo",
        selected_name="selected_output.txt",
    )

    final_row["mg_removal_achieved"] = (
        mg_in_mol - get_row_molality(react_row, "Mg") * float(react_row["mass_H2O"])
    ) / mg_in_mol

    phase_df = build_stage_phase_summary(react_row, final_row, stage_order, scenario_id)
    clean_row, next_solution = build_stage_clean_row(
        react_row=react_row,
        stage_row=final_row,
        stage_order=stage_order,
        scenario_id=scenario_id,
        pqo_props=pqo_props,
        phase_df=phase_df,
        solution_in=solution_in,
        input_file=input_file,
        output_file=output_file,
        selected_file=selected_file,
    )

    return phase_df, clean_row, next_solution


def run_stage(
    stage_order: int,
    stage_row: pd.Series,
    solution_in: dict,
    scenario_id: str,
    scenario_run_dir: Path,
) -> tuple[pd.DataFrame, dict, dict]:
    stage_id = str(stage_row["stage_id"]).upper()
    reagent = str(stage_row.get("reagent", "")).strip().lower()

    if stage_id == "LIM1" and reagent in {"ca(oh)2", "caoh2"}:
        return run_liming_stage(stage_order, stage_row, solution_in, scenario_id, scenario_run_dir)

    return run_regular_stage(stage_order, stage_row, solution_in, scenario_id, scenario_run_dir)


# =========================================================
# VALIDACION DE TENDENCIAS
# =========================================================

def build_validation_summary(stage_trace: pd.DataFrame, scenarios_generated: pd.DataFrame) -> pd.DataFrame:
    rows = []

    expected_scenarios = len(scenarios_generated)
    observed_scenarios = stage_trace["scenario_id"].nunique() if not stage_trace.empty else 0
    rows.append({
        "check": "scenario_count",
        "status": "OK" if observed_scenarios == expected_scenarios else "WARN",
        "details": f"expected={expected_scenarios}; observed={observed_scenarios}",
    })

    if not scenarios_generated.empty and "month" in scenarios_generated.columns:
        rows.append({
            "check": "months_generated",
            "status": "OK",
            "details": ",".join(str(int(m)) for m in sorted(scenarios_generated["month"].unique())),
        })

    if not stage_trace.empty:
        pond_rows = stage_trace[(stage_trace["stage_type"] == "pond") & (stage_trace["error"].fillna("") == "")]
        bad_evap = pond_rows[pond_rows["Q_evap"].fillna(0.0) < 0]
        rows.append({
            "check": "evaporation_non_negative",
            "status": "OK" if bad_evap.empty else "FAIL",
            "details": f"bad_rows={len(bad_evap)}",
        })

        completed = stage_trace[(stage_trace["stage_id"] != "FEED") & (stage_trace["error"].fillna("") == "")]
        if not completed.empty:
            trends = []
            for sid, df in completed.groupby("scenario_id"):
                df = df.sort_values("stage_order")
                if (df["Q_out"].dropna().diff().dropna() > 1e-9).any():
                    trends.append(sid)
            rows.append({
                "check": "flow_generally_decreases",
                "status": "OK" if not trends else "WARN",
                "details": ",".join(trends[:10]),
            })

        lim_rows = stage_trace[(stage_trace["stage_id"] == "LIM1") & (stage_trace["error"].fillna("") == "")]
        if not lim_rows.empty:
            diff = (lim_rows["mg_removal_achieved"] - lim_rows["mg_removal"]).abs()
            ok = diff.dropna().le(0.005).all()
            rows.append({
                "check": "liming_target_tolerance",
                "status": "OK" if ok else "WARN",
                "details": f"max_abs_diff={diff.max()}",
            })

        qout_rows = stage_trace[(stage_trace["stage_id"] != "FEED") & (stage_trace["error"].fillna("") == "")]
        bad_qout = qout_rows[pd.to_numeric(qout_rows["Q_out"], errors="coerce") < -1e-12]
        rows.append({
            "check": "Q_out_non_negative",
            "status": "OK" if bad_qout.empty else "FAIL",
            "details": f"bad_rows={len(bad_qout)}",
        })

        lim_evap = stage_trace[stage_trace["stage_id"].astype(str).str.upper() == "LIM1"]
        bad_lim_evap = lim_evap[pd.to_numeric(lim_evap["Q_evap"], errors="coerce").fillna(0.0).abs() > 1e-12]
        rows.append({
            "check": "LIM1_evaporation_zero",
            "status": "OK" if bad_lim_evap.empty else "FAIL",
            "details": f"bad_rows={len(bad_lim_evap)}",
        })

        decreasing_solids = []
        for sid, df in qout_rows.groupby("scenario_id"):
            diffs = pd.to_numeric(df.sort_values("stage_order")["cumulative_solids_kg"], errors="coerce").diff()
            if (diffs.dropna() < -1e-12).any():
                decreasing_solids.append(sid)
        rows.append({
            "check": "accumulated_solids_non_decreasing",
            "status": "OK" if not decreasing_solids else "FAIL",
            "details": ",".join(decreasing_solids[:10]),
        })

        missing_scenario = stage_trace["scenario_id"].isna().sum()
        rows.append({
            "check": "scenario_linkage_present",
            "status": "OK" if missing_scenario == 0 else "FAIL",
            "details": f"missing_scenario_id_rows={missing_scenario}",
        })

    return pd.DataFrame(rows)


# =========================================================
# SALIDAS METODOLOGICAS TFG
# =========================================================

def kg_s_to_t_d(value):
    if value is None or pd.isna(value):
        return None
    return float(value) * 86.4


def pct_delta(current, previous):
    if current is None or previous is None or pd.isna(current) or pd.isna(previous) or float(previous) == 0:
        return None
    return 100.0 * (float(current) - float(previous)) / float(previous)


def stage_group_type(stage_id: str, stage_type: str | None = None) -> str:
    sid = str(stage_id).upper().strip()
    if sid == "FEED":
        return "FEED"
    if sid.startswith("LIM"):
        return "LIM"
    for prefix in ["PC", "H", "K", "C", "L"]:
        if sid.startswith(prefix):
            return prefix
    return str(stage_type or "unknown")


def row_density_kg_l(row: pd.Series | dict) -> float | None:
    density = row.get("density_out_kg_L", None)
    if density is not None and pd.notna(density) and float(density) > 0:
        return float(density)

    liquid_mass = row.get("liquid_mass_out_kg", None)
    volume_l = row.get("volume_out_L", None)
    if (
        liquid_mass is not None and pd.notna(liquid_mass) and
        volume_l is not None and pd.notna(volume_l) and float(volume_l) > 0
    ):
        return float(liquid_mass) / float(volume_l)
    return None


def retained_l_s(row: pd.Series | dict) -> float | None:
    retained_kg_s = row.get("retained_brine_kg", row.get("Q_retained", None))
    if retained_kg_s is None or pd.isna(retained_kg_s):
        return None

    density = row_density_kg_l(row)
    if density is not None and density > 0:
        return float(retained_kg_s) / density

    volume_l = row.get("volume_out_L", None)
    frac = row.get("retained_fraction_liquid", None)
    if volume_l is not None and pd.notna(volume_l) and frac is not None and pd.notna(frac):
        return float(volume_l) * float(frac)

    return float(retained_kg_s)


def brine_remaining_l_s(row: pd.Series | dict) -> float | None:
    volume_l = row.get("volume_out_L", None)
    frac = row.get("retained_fraction_liquid", 0.0)
    if volume_l is not None and pd.notna(volume_l):
        frac = 0.0 if frac is None or pd.isna(frac) else float(frac)
        return float(volume_l) * (1.0 - frac)

    liquid_to_next = row.get("liquid_to_next_kg", row.get("Q_out", None))
    density = row_density_kg_l(row)
    if liquid_to_next is not None and pd.notna(liquid_to_next) and density is not None and density > 0:
        return float(liquid_to_next) / density
    return None


def phase_mass_from_pairs(phase_masses_text, phase_filter) -> float:
    if phase_masses_text is None or pd.isna(phase_masses_text):
        return 0.0

    total = 0.0
    for part in str(phase_masses_text).split(";"):
        part = part.strip()
        if not part or ":" not in part:
            continue
        phase, mass_text = part.rsplit(":", 1)
        try:
            mass = float(mass_text)
        except ValueError:
            continue
        if phase_filter(phase):
            total += mass
    return total


def percent_of_total(value, total) -> float | None:
    if value is None or total is None or pd.isna(value) or pd.isna(total) or float(total) == 0.0:
        return None
    return 100.0 * float(value) / float(total)


def valid_stage_trace(stage_trace: pd.DataFrame) -> pd.DataFrame:
    if stage_trace.empty:
        return stage_trace.copy()
    return stage_trace[
        (stage_trace["stage_id"].astype(str) != "FEED") &
        (stage_trace["error"].fillna("").astype(str) == "")
    ].copy()


def build_summary_results(stage_trace: pd.DataFrame) -> pd.DataFrame:
    output_cols = [
        "scenario_id", "month", "month_label", "month_f_m", "bypass_velocity_factor",
        "evap_factor", "mg_removal", "retention_r",
        "last_valid_stage_id", "actual_last_valid_Q_out_L_s", "design_total_evap_L_s",
        "stage_id", "stage_type", "group_type",
        "Q_in_L_s", "Q_out_L_s", "Q_evap_L_s", "Q_evap_original_L_s",
        "Q_evap_external_L_s", "Q_evap_accum_visible_L_s",
        "evap_recovered_by_module_L_s", "evap_reduction_percent",
        "Q_module_recovered_water_L_s", "Q_module_extra_recovered_water_L_s",
        "recovery_evap_scale_factor", "bypass_fraction_percent",
        "Q_evap_accum_L_s",
        "accumulated_evaporated_water_percent",
        "Q_retained_L_s", "Q_retained_accum_L_s",
        "Q_brine_remaining_L_s",
        "Li_mg_L", "Li_g_L", "TDS_g_L",
        "total_solids_precipitated_kg_s", "total_solids_precipitated_t_d",
        "accumulated_solids_kg_s", "accumulated_solids_t_d",
        "accumulated_halite_precipitated_percent",
        "accumulated_sulfates_precipitated_percent",
    ]
    valid = valid_stage_trace(stage_trace)
    if valid.empty:
        return pd.DataFrame(columns=output_cols)

    feed = stage_trace[stage_trace["stage_id"].astype(str) == "FEED"].copy()
    feed["Q_brine_remaining_L_s_calc"] = feed.apply(brine_remaining_l_s, axis=1)

    valid["group_type"] = valid.apply(lambda r: stage_group_type(r["stage_id"], r["stage_type"]), axis=1)
    valid["Q_brine_remaining_L_s"] = valid.apply(brine_remaining_l_s, axis=1)
    valid["Q_retained_L_s"] = valid.apply(retained_l_s, axis=1)
    valid["TDS_g_L"] = valid.apply(major_ion_tds_g_l, axis=1)
    valid["Li_mg_L"] = pd.to_numeric(valid["Li_mg_L_est"], errors="coerce")
    valid["Li_g_L"] = valid["Li_mg_L"] / 1000.0
    valid["total_solids_precipitated_kg_s"] = pd.to_numeric(valid["total_precipitated_mass_kg"], errors="coerce").fillna(0.0)
    valid["total_solids_precipitated_t_d"] = valid["total_solids_precipitated_kg_s"].apply(kg_s_to_t_d)
    valid["accumulated_solids_kg_s"] = pd.to_numeric(valid["cumulative_solids_kg"], errors="coerce").fillna(0.0)
    valid["accumulated_solids_t_d"] = valid["accumulated_solids_kg_s"].apply(kg_s_to_t_d)

    valid = valid.sort_values(["scenario_id", "stage_order"]).reset_index(drop=True)
    valid["Q_evap_external_L_s"] = pd.to_numeric(
        valid.get("Q_evap_external_L_s", 0.0),
        errors="coerce",
    ).fillna(0.0)
    valid["evap_recovered_by_module_L_s"] = pd.to_numeric(
        valid.get("evap_recovered_by_module_L_s", 0.0),
        errors="coerce",
    ).fillna(0.0)
    valid["evap_reduction_percent"] = pd.to_numeric(
        valid.get("evap_reduction_percent", 0.0),
        errors="coerce",
    ).fillna(0.0)
    valid["Q_module_recovered_water_L_s"] = pd.to_numeric(
        valid.get("Q_module_recovered_water_L_s", 0.0),
        errors="coerce",
    ).fillna(0.0)
    valid["Q_module_extra_recovered_water_L_s"] = pd.to_numeric(
        valid.get("Q_module_extra_recovered_water_L_s", 0.0),
        errors="coerce",
    ).fillna(0.0)
    valid["Q_evap_accum_visible_L_s"] = valid.groupby("scenario_id")["Q_evap"].cumsum()
    valid["Q_evap_total_for_accum_L_s"] = valid["Q_evap"] + valid["Q_evap_external_L_s"]
    valid["Q_evap_accum_L_s"] = valid.groupby("scenario_id")["Q_evap_total_for_accum_L_s"].cumsum()
    valid["Q_retained_accum_L_s"] = valid.groupby("scenario_id")["Q_retained_L_s"].cumsum()
    valid["halite_precipitated_kg_s_calc"] = valid["phase_masses_kg"].apply(
        lambda value: phase_mass_from_pairs(value, lambda phase: phase == "Halite")
    )
    valid["sulfates_precipitated_kg_s_calc"] = valid["phase_masses_kg"].apply(
        lambda value: phase_mass_from_pairs(value, lambda phase: mineral_family(phase) == "sulfates")
    )
    valid["accumulated_halite_precipitated_kg_s_calc"] = valid.groupby("scenario_id")["halite_precipitated_kg_s_calc"].cumsum()
    valid["accumulated_sulfates_precipitated_kg_s_calc"] = valid.groupby("scenario_id")["sulfates_precipitated_kg_s_calc"].cumsum()

    halite_total_map = valid.groupby("scenario_id")["halite_precipitated_kg_s_calc"].sum().to_dict()
    sulfates_total_map = valid.groupby("scenario_id")["sulfates_precipitated_kg_s_calc"].sum().to_dict()
    valid["accumulated_halite_precipitated_percent"] = valid.apply(
        lambda r: percent_of_total(
            r["accumulated_halite_precipitated_kg_s_calc"],
            halite_total_map.get(r["scenario_id"]),
        ),
        axis=1,
    )
    valid["accumulated_sulfates_precipitated_percent"] = valid.apply(
        lambda r: percent_of_total(
            r["accumulated_sulfates_precipitated_kg_s_calc"],
            sulfates_total_map.get(r["scenario_id"]),
        ),
        axis=1,
    )

    q_in_values = []
    feed_q_map = dict(zip(feed["scenario_id"], feed["Q_brine_remaining_L_s_calc"]))
    previous_by_scenario = {}
    for _, row in valid.iterrows():
        sid = row["scenario_id"]
        q_in_override = pd.to_numeric(row.get("Q_in_L_s_override", None), errors="coerce")
        if pd.notna(q_in_override):
            q_in = float(q_in_override)
        else:
            q_in = previous_by_scenario.get(sid, feed_q_map.get(sid, None))
        q_in_values.append(q_in)
        previous_by_scenario[sid] = row["Q_brine_remaining_L_s"]
    valid["Q_in_L_s"] = q_in_values
    valid["Q_out_L_s"] = valid["Q_brine_remaining_L_s"]

    final_rows = valid.sort_values(["scenario_id", "stage_order"]).groupby("scenario_id", as_index=False).tail(1)
    target_records = {}
    for _, final_row in final_rows.iterrows():
        target_records[final_row["scenario_id"]] = {
            "last_valid_stage_id": final_row["stage_id"],
            "actual_last_valid_Q_out_L_s": final_row["Q_out_L_s"],
            "design_total_evap_L_s": final_row.get("design_total_evap_L_s", final_row["Q_evap_accum_L_s"]),
        }

    for key in [
        "last_valid_stage_id",
        "actual_last_valid_Q_out_L_s",
        "design_total_evap_L_s",
    ]:
        valid[key] = valid["scenario_id"].map(lambda sid: target_records.get(sid, {}).get(key))

    total_evap_map = valid.groupby("scenario_id")["Q_evap_total_for_accum_L_s"].sum().to_dict()
    valid["evap_percent_denominator_L_s_calc"] = pd.to_numeric(
        valid["design_total_evap_L_s"],
        errors="coerce",
    )
    valid["evap_percent_denominator_L_s_calc"] = valid.apply(
        lambda r: total_evap_map.get(r["scenario_id"])
        if pd.isna(r["evap_percent_denominator_L_s_calc"]) or float(r["evap_percent_denominator_L_s_calc"]) == 0.0
        else r["evap_percent_denominator_L_s_calc"],
        axis=1,
    )
    valid["accumulated_evaporated_water_percent"] = valid.apply(
        lambda r: percent_of_total(
            r["Q_evap_accum_L_s"],
            r["evap_percent_denominator_L_s_calc"],
        ),
        axis=1,
    )

    cols = [
        "scenario_id", "month", "month_label", "month_f_m", "bypass_velocity_factor",
        "evap_factor", "mg_removal", "retention_r",
        "last_valid_stage_id", "actual_last_valid_Q_out_L_s", "design_total_evap_L_s",
        "stage_id", "stage_type", "group_type",
        "Q_in_L_s", "Q_out_L_s", "Q_evap", "Q_evap_original_L_s",
        "Q_evap_external_L_s", "Q_evap_accum_visible_L_s",
        "evap_recovered_by_module_L_s", "evap_reduction_percent",
        "Q_module_recovered_water_L_s", "Q_module_extra_recovered_water_L_s",
        "recovery_evap_scale_factor", "bypass_fraction_percent",
        "Q_evap_accum_L_s",
        "accumulated_evaporated_water_percent",
        "Q_retained_L_s", "Q_retained_accum_L_s",
        "Q_brine_remaining_L_s",
        "Li_mg_L", "Li_g_L", "TDS_g_L",
        "total_solids_precipitated_kg_s", "total_solids_precipitated_t_d",
        "accumulated_solids_kg_s", "accumulated_solids_t_d",
        "accumulated_halite_precipitated_percent",
        "accumulated_sulfates_precipitated_percent",
    ]
    out = valid[cols].rename(columns={"Q_evap": "Q_evap_L_s"})
    return out


def mineral_family(phase: str) -> str:
    return MINERAL_FAMILIES.get(str(phase), "other")


def build_precipitation_by_phase(phases_df: pd.DataFrame) -> pd.DataFrame:
    if phases_df.empty:
        return pd.DataFrame(columns=[
            "scenario_id", "month", "stage_id", "stage_type", "mineral_phase",
            "mineral_family", "precipitated_mass_kg_s", "precipitated_mass_t_d",
            "accumulated_precipitated_mass_kg_s", "accumulated_precipitated_mass_t_d",
        ])

    df = phases_df.copy()
    df["mineral_phase"] = df["phase_name"]
    df["mineral_family"] = df["mineral_phase"].apply(mineral_family)
    df["precipitated_mass_kg_s"] = pd.to_numeric(df["phase_mass_kg"], errors="coerce").fillna(0.0)
    df["precipitated_mass_t_d"] = df["precipitated_mass_kg_s"].apply(kg_s_to_t_d)
    df = df.sort_values(["scenario_id", "mineral_phase", "stage_order"]).reset_index(drop=True)
    df["accumulated_precipitated_mass_kg_s"] = df.groupby(["scenario_id", "mineral_phase"])["precipitated_mass_kg_s"].cumsum()
    df["accumulated_precipitated_mass_t_d"] = df["accumulated_precipitated_mass_kg_s"].apply(kg_s_to_t_d)

    cols = [
        "scenario_id", "month", "stage_id", "stage_type", "mineral_phase",
        "mineral_family", "precipitated_mass_kg_s", "precipitated_mass_t_d",
        "accumulated_precipitated_mass_kg_s", "accumulated_precipitated_mass_t_d",
    ]
    return df[cols]


def build_family_percent_lookup(phases_df: pd.DataFrame) -> dict[tuple[str, int], str]:
    if phases_df.empty:
        return {}

    df = phases_df.copy()
    df["mineral_family"] = df["phase_name"].apply(mineral_family)
    df["mass"] = pd.to_numeric(df["phase_mass_kg"], errors="coerce").fillna(0.0)
    total_by_scenario = df.groupby("scenario_id")["mass"].sum().to_dict()

    lookup = {}
    for sid, sdf in df.groupby("scenario_id"):
        total = float(total_by_scenario.get(sid, 0.0))
        if total <= 0:
            continue
        stage_orders = sorted(sdf["stage_order"].unique())
        for stage_order in stage_orders:
            upto = sdf[sdf["stage_order"] <= stage_order]
            family_masses = upto.groupby("mineral_family")["mass"].sum()
            parts = [f"{family}:{100.0 * mass / total:.3f}" for family, mass in family_masses.sort_index().items()]
            lookup[(sid, int(stage_order))] = "|".join(parts)
    return lookup


def build_evaluation_results(summary_df: pd.DataFrame, stage_trace: pd.DataFrame, phases_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame()

    df = summary_df.copy()
    feed_li_map = (
        stage_trace[stage_trace["stage_id"].astype(str) == "FEED"]
        .set_index("scenario_id")["Li_mg_L_est"]
        .to_dict()
    )
    total_solids_map = df.groupby("scenario_id")["accumulated_solids_kg_s"].max().to_dict()
    family_lookup = build_family_percent_lookup(phases_df)
    stage_order_map = (
        stage_trace[["scenario_id", "stage_id", "stage_order"]]
        .drop_duplicates()
        .set_index(["scenario_id", "stage_id"])["stage_order"]
        .to_dict()
    )

    valid = valid_stage_trace(stage_trace)
    for col in ["Mg/Li", "K/Li", "SO4/Li", "Ca/Li", "TDS"]:
        if col not in valid.columns:
            valid[col] = None

    df["Li_concentration_factor"] = df.apply(
        lambda r: safe_ratio(r["Li_mg_L"], feed_li_map.get(r["scenario_id"])), axis=1
    )
    df["solids_per_evaporated_water_kg_kg"] = df.apply(
        lambda r: safe_ratio(r["total_solids_precipitated_kg_s"], r["Q_evap_L_s"] * RHO_WATER_KG_L), axis=1
    )
    df["accumulated_precipitation_percent"] = df.apply(
        lambda r: None if total_solids_map.get(r["scenario_id"], 0.0) == 0 else
        100.0 * r["accumulated_solids_kg_s"] / total_solids_map[r["scenario_id"]],
        axis=1,
    )
    df["stage_order"] = df.apply(lambda r: stage_order_map.get((r["scenario_id"], r["stage_id"])), axis=1)
    df["accumulated_precipitation_by_family_percent"] = df.apply(
        lambda r: family_lookup.get((r["scenario_id"], int(r["stage_order"])), ""), axis=1
    )

    ratio_cols = valid[["scenario_id", "stage_id", "Mg/Li", "K/Li", "SO4/Li", "Ca/Li", "TDS"]].copy()
    ratio_cols["TDS_g_L_trace"] = ratio_cols.apply(major_ion_tds_g_l, axis=1)
    df = df.merge(ratio_cols, on=["scenario_id", "stage_id"], how="left")
    df = df.sort_values(["scenario_id", "stage_order"]).reset_index(drop=True)

    for source_col, target_col in [
        ("Li_mg_L", "delta_Li_between_stages_percent"),
        ("TDS_g_L", "delta_TDS_between_stages_percent"),
        ("Mg/Li", "delta_Mg_Li_between_stages_percent"),
        ("SO4/Li", "delta_SO4_Li_between_stages_percent"),
    ]:
        values = []
        previous = {}
        for _, row in df.iterrows():
            sid = row["scenario_id"]
            values.append(pct_delta(row.get(source_col), previous.get(sid)))
            previous[sid] = row.get(source_col)
        df[target_col] = values

    cols = [
        "scenario_id", "month", "stage_id", "stage_type",
        "Li_concentration_factor",
        "solids_per_evaporated_water_kg_kg",
        "accumulated_precipitation_percent",
        "accumulated_precipitation_by_family_percent",
        "Mg/Li", "K/Li", "SO4/Li", "Ca/Li",
        "delta_Li_between_stages_percent",
        "delta_TDS_between_stages_percent",
        "delta_Mg_Li_between_stages_percent",
        "delta_SO4_Li_between_stages_percent",
    ]
    return df[cols].rename(columns={
        "Mg/Li": "Mg_Li_ratio",
        "K/Li": "K_Li_ratio",
        "SO4/Li": "SO4_Li_ratio",
        "Ca/Li": "Ca_Li_ratio",
    })


def build_geochemical_results(stage_trace: pd.DataFrame) -> pd.DataFrame:
    valid = valid_stage_trace(stage_trace)
    if valid.empty:
        return pd.DataFrame()

    rows = []
    for _, row in valid.sort_values(["scenario_id", "stage_order"]).iterrows():
        out = {
            "scenario_id": row["scenario_id"],
            "month": row["month"],
            "stage_id": row["stage_id"],
            "density_kg_L": row_density_kg_l(row),
            "pH": row.get("pH_out"),
            "Na_mg_L": row.get("Na_mg_L_est"),
            "K_mg_L": row.get("K_mg_L_est"),
            "Mg_mg_L": row.get("Mg_mg_L_est"),
            "Ca_mg_L": row.get("Ca_mg_L_est"),
            "Li_mg_L": row.get("Li_mg_L_est"),
            "Cl_mg_L": row.get("Cl_mg_L_est"),
            "SO4_mg_L": sulfate_mg_l(row),
            "B_mg_L": row.get("B_mg_L_est"),
            "HCO3_mg_L": row.get("HCO3_mg_L_est"),
            "Na_mol_kgw": row.get("Na"),
            "K_mol_kgw": row.get("K"),
            "Mg_mol_kgw": row.get("Mg"),
            "Ca_mol_kgw": row.get("Ca"),
            "Li_mol_kgw": row.get("Li"),
            "Cl_mol_kgw": row.get("Cl"),
            "SO4_mol_kgw": row.get("S6"),
            "B_mol_kgw": row.get("B"),
            "HCO3_mol_kgw": row.get("HCO3"),
            "Mg_Li_ratio": row.get("Mg/Li"),
            "Ca_Li_ratio": row.get("Ca/Li"),
            "K_Li_ratio": row.get("K/Li"),
            "SO4_Li_ratio": row.get("SO4/Li"),
            "precipitation_sequence": row.get("phases_precipitated", ""),
            "main_saturation_indexes": row.get("main_saturation_indexes", ""),
            "water_activity": row.get("water_activity", None),
            "ionic_strength_mol_kgw": row.get("mu_out", None),
        }
        rows.append(out)

    return pd.DataFrame(rows)


def phase_scaling_penalty(phase: str) -> float:
    phase_name = str(phase)
    if phase_name in PHASE_SCALING_PENALTIES:
        return PHASE_SCALING_PENALTIES[phase_name]
    if mineral_family(phase_name) == "borates":
        return BORATE_SCALING_PENALTY
    return DEFAULT_UNKNOWN_PHASE_PENALTY


def parse_saturation_indexes(si_text) -> dict[str, float]:
    if si_text is None or pd.isna(si_text):
        return {}

    parsed = {}
    for part in str(si_text).split("|"):
        part = part.strip()
        if not part or ":" not in part:
            continue
        phase, value = part.rsplit(":", 1)
        try:
            parsed[phase.strip()] = float(value)
        except ValueError:
            continue
    return parsed


def saturation_score_from_indexes(si_text) -> float:
    score = 0.0
    for si in parse_saturation_indexes(si_text).values():
        if si >= 0.0:
            score += 20.0
        elif si >= -0.2:
            score += 10.0
    return min(score, 100.0)


def scaling_risk_class(value) -> str:
    if value is None or pd.isna(value):
        return ""
    value = float(value)
    if value < 35.0:
        return "Bajo"
    if value < 60.0:
        return "Medio"
    return "Alto"


def build_phase_mass_lookup(precipitation_df: pd.DataFrame) -> dict[tuple[str, str], dict[str, float]]:
    if precipitation_df.empty:
        return {}

    work = precipitation_df.copy()
    work["precipitated_mass_kg_s"] = pd.to_numeric(
        work["precipitated_mass_kg_s"],
        errors="coerce",
    ).fillna(0.0)

    lookup = {}
    for (sid, stage_id), group in work.groupby(["scenario_id", "stage_id"], sort=False):
        masses = (
            group.groupby("mineral_phase")["precipitated_mass_kg_s"]
            .sum()
            .to_dict()
        )
        lookup[(sid, stage_id)] = {str(k): float(v) for k, v in masses.items() if float(v) > 0.0}
    return lookup


def build_scaling_risk_results(
    summary_df: pd.DataFrame,
    precipitation_df: pd.DataFrame,
    geochemical_df: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "scenario_id", "month", "stage_id", "stage_type",
        "water_evaporated_kg_s", "total_solids_precipitated_kg_s",
        "solids_water_ratio_kg_kg",
        "solid_water_score", "phase_penalty_score",
        "saturation_score", "chemical_treatment_penalty",
        "Psa_solids_water_score", "Pf_phase_penalty_score",
        "Ps_saturation_score", "Ptq_chemical_treatment_penalty",
        "solid_water_weight", "phase_penalty_weight",
        "saturation_weight", "chemical_treatment_penalty_weight",
        "solid_water_contribution", "phase_penalty_contribution",
        "saturation_contribution", "chemical_treatment_contribution",
        "scaling_risk_indicator", "scaling_risk_class",
        "precipitated_phases", "saturation_indexes_by_phase",
    ]
    if summary_df.empty:
        return pd.DataFrame(columns=columns)

    df = summary_df.copy()
    df["water_evaporated_kg_s"] = pd.to_numeric(df["Q_evap_L_s"], errors="coerce").fillna(0.0) * RHO_WATER_KG_L
    df["total_solids_precipitated_kg_s"] = pd.to_numeric(
        df["total_solids_precipitated_kg_s"],
        errors="coerce",
    ).fillna(0.0)
    df["solids_water_ratio_kg_kg"] = df.apply(
        lambda r: 0.0 if float(r["water_evaporated_kg_s"]) <= 0.0 else
        float(r["total_solids_precipitated_kg_s"]) / float(r["water_evaporated_kg_s"]),
        axis=1,
    )

    max_ratio = float(df["solids_water_ratio_kg_kg"].max() or 0.0)
    df["solid_water_score"] = df.apply(
        lambda r: 100.0 if float(r["water_evaporated_kg_s"]) <= 0.0 else
        (0.0 if max_ratio <= 0.0 else 100.0 * float(r["solids_water_ratio_kg_kg"]) / max_ratio),
        axis=1,
    )

    si_lookup = {}
    if not geochemical_df.empty:
        si_lookup = (
            geochemical_df[["scenario_id", "stage_id", "main_saturation_indexes"]]
            .drop_duplicates()
            .set_index(["scenario_id", "stage_id"])["main_saturation_indexes"]
            .to_dict()
        )
    phase_lookup = build_phase_mass_lookup(precipitation_df)

    phase_scores = []
    saturation_scores = []
    treatment_penalties = []
    phase_strings = []
    si_strings = []

    for _, row in df.iterrows():
        key = (row["scenario_id"], row["stage_id"])
        masses = phase_lookup.get(key, {})
        total_phase_mass = sum(masses.values())
        if total_phase_mass > 0.0:
            phase_score = sum(mass * phase_scaling_penalty(phase) for phase, mass in masses.items()) / total_phase_mass
        else:
            phase_score = 0.0

        si_text = si_lookup.get(key, "")
        saturation_score = saturation_score_from_indexes(si_text)

        stage_id = str(row["stage_id"]).upper()
        stage_type = str(row["stage_type"]).lower()
        penalty = 0.0
        if "LIM" in stage_id or stage_type == "chemical_treatment":
            penalty += 10.0
        if any(phase.lower() == "brucite" for phase in masses):
            penalty += 10.0
        if any(phase in {"Gypsum", "Anhydrite"} for phase in masses):
            penalty += 5.0

        phase_scores.append(phase_score)
        saturation_scores.append(saturation_score)
        treatment_penalties.append(penalty)
        phase_strings.append("|".join(f"{phase}:{mass:.6g}" for phase, mass in sorted(masses.items())))
        si_strings.append(si_text)

    df["phase_penalty_score"] = phase_scores
    df["saturation_score"] = saturation_scores
    df["chemical_treatment_penalty"] = treatment_penalties
    df["Psa_solids_water_score"] = df["solid_water_score"]
    df["Pf_phase_penalty_score"] = df["phase_penalty_score"]
    df["Ps_saturation_score"] = df["saturation_score"]
    df["Ptq_chemical_treatment_penalty"] = df["chemical_treatment_penalty"]
    df["solid_water_weight"] = SCALING_RISK_WEIGHTS["solid_water_score"]
    df["phase_penalty_weight"] = SCALING_RISK_WEIGHTS["phase_penalty_score"]
    df["saturation_weight"] = SCALING_RISK_WEIGHTS["saturation_score"]
    df["chemical_treatment_penalty_weight"] = SCALING_RISK_WEIGHTS["chemical_treatment_penalty"]
    df["solid_water_contribution"] = df["solid_water_score"] * df["solid_water_weight"]
    df["phase_penalty_contribution"] = df["phase_penalty_score"] * df["phase_penalty_weight"]
    df["saturation_contribution"] = df["saturation_score"] * df["saturation_weight"]
    df["chemical_treatment_contribution"] = (
        df["chemical_treatment_penalty"] * df["chemical_treatment_penalty_weight"]
    )
    df["precipitated_phases"] = phase_strings
    df["saturation_indexes_by_phase"] = si_strings
    df["scaling_risk_indicator"] = df.apply(
        lambda r: min(
            100.0,
            float(r["solid_water_contribution"]) +
            float(r["phase_penalty_contribution"]) +
            float(r["saturation_contribution"]) +
            float(r["chemical_treatment_contribution"]),
        ),
        axis=1,
    )
    df["scaling_risk_class"] = df["scaling_risk_indicator"].apply(scaling_risk_class)
    return df[columns]


def build_scenario_metadata(
    stage_trace: pd.DataFrame,
    scenarios_generated: pd.DataFrame,
    validation_df: pd.DataFrame,
    run_stamp: str,
) -> dict:
    scenario_rows = []
    for _, scenario in scenarios_generated.iterrows():
        sid = scenario["scenario_id"]
        sdf = stage_trace[stage_trace["scenario_id"] == sid].sort_values("stage_order")
        valid = sdf[(sdf["stage_id"].astype(str) != "FEED") & (sdf["error"].fillna("").astype(str) == "")]
        errors = sdf[sdf["error"].fillna("").astype(str) != ""]

        last_valid = valid.iloc[-1] if not valid.empty else None
        first_error = errors.iloc[0] if not errors.empty else None
        total_evap_l_s = 0.0
        total_retained_l_s = 0.0
        design_total_evap_l_s = None
        actual_last_q_out_l_s = None

        if last_valid is not None:
            valid = valid.copy()
            valid["Q_retained_L_s_calc"] = valid.apply(retained_l_s, axis=1)
            valid["Q_brine_remaining_L_s_calc"] = valid.apply(brine_remaining_l_s, axis=1)
            total_evap_l_s = float(pd.to_numeric(valid["Q_evap"], errors="coerce").fillna(0.0).sum())
            total_retained_l_s = float(pd.to_numeric(valid["Q_retained_L_s_calc"], errors="coerce").fillna(0.0).sum())
            design_total_evap_l_s = last_valid.get("design_total_evap_L_s", None)
            if design_total_evap_l_s is not None and pd.notna(design_total_evap_l_s):
                design_total_evap_l_s = float(design_total_evap_l_s)
            actual_last_q_out_l_s = brine_remaining_l_s(last_valid)

        scenario_rows.append({
            "scenario_id": sid,
            "month": int(scenario["month"]),
            "month_label": scenario["month_label"],
            "month_f_m": float(scenario["month_f_m"]),
            "bypass_velocity_factor": float(scenario["bypass_velocity_factor"]),
            "temperature": float(scenario["temperature"]),
            "evap_factor": float(scenario["evap_factor"]),
            "mg_removal": float(scenario["mg_removal"]),
            "retention_r": float(scenario["retention_r"]),
            "scenario_status": "FAIL" if first_error is not None else "OK",
            "last_valid_stage_id": None if last_valid is None else last_valid["stage_id"],
            "Li_last_valid_mg_L": None if last_valid is None else float(last_valid["Li_mg_L_est"]),
            "Li_last_valid_ppm_approx": None if last_valid is None else float(last_valid["Li_mg_L_est"]),
            "actual_last_valid_Q_out_L_s": actual_last_q_out_l_s,
            "design_total_evap_L_s": design_total_evap_l_s,
            "actual_total_retained_L_s": total_retained_l_s,
            "failed_stage_id": None if first_error is None else first_error["stage_id"],
            "error": None if first_error is None else first_error["error"],
        })

    return {
        "run_stamp": run_stamp,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "root": str(ROOT),
        "input_dir": str(INPUTS),
        "phreeqc_exe": str(PHREEQC_EXE),
        "phreeqc_database": str(DATABASE),
        "exports": [
            "simulator_results.xlsx",
            "results_viewer.html",
            "viewer_assets/plotly-2.35.2.min.js",
            "summary_results.csv",
            "evaluation_results.csv",
            "geochemical_results.csv",
            "precipitation_by_phase.csv",
            "scaling_risk_results.csv",
            "scaling_risk_methodology.txt",
            "scenario_metadata.json",
        ],
        "validation": validation_df.to_dict(orient="records"),
        "scenarios": scenario_rows,
    }


def build_methodological_outputs(
    stage_trace: pd.DataFrame,
    phases_df: pd.DataFrame,
    scenarios_generated: pd.DataFrame,
    validation_df: pd.DataFrame,
    run_stamp: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    summary_df = build_summary_results(stage_trace)
    precipitation_df = build_precipitation_by_phase(phases_df)
    evaluation_df = build_evaluation_results(summary_df, stage_trace, phases_df)
    geochemical_df = build_geochemical_results(stage_trace)
    scaling_risk_df = build_scaling_risk_results(summary_df, precipitation_df, geochemical_df)
    metadata = build_scenario_metadata(stage_trace, scenarios_generated, validation_df, run_stamp)
    return summary_df, evaluation_df, geochemical_df, precipitation_df, scaling_risk_df, metadata


# =========================================================
# FORMATO EXCEL
# =========================================================

def build_units_row(df: pd.DataFrame) -> list[str]:
    units_map = {
        "scenario_id": "-",
        "scenario_status": "-",
        "month": "-",
        "month_label": "-",
        "month_f_m": "-",
        "bypass_velocity_factor": "-",
        "temperature": "degC",
        "evap_factor": "-",
        "mg_removal": "fraction",
        "retention_r": "kg brine/kg solid",
        "stage_order": "-",
        "stage_id": "-",
        "stage_type": "-",
        "temp_C": "degC",
        "pH_out": "-",
        "mu_out": "mol/kgw",
        "Q_in": "kg/s basis",
        "Q_in_L_s": "L/s",
        "Q_evap": "L/s",
        "Q_evap_L_s": "L/s",
        "Q_evap_original_L_s": "L/s",
        "Q_evap_external_L_s": "L/s",
        "Q_evap_accum_visible_L_s": "L/s",
        "evap_recovered_by_module_L_s": "L/s",
        "evap_reduction_percent": "%",
        "Q_module_recovered_water_L_s": "L/s",
        "Q_module_extra_recovered_water_L_s": "L/s",
        "Q_evap_kg_s": "kg/s",
        "recovery_evap_scale_factor": "-",
        "bypass_fraction_percent": "%",
        "Q_evap_accum_L_s": "L/s",
        "accumulated_evaporated_water_percent": "%",
        "Q_retained": "kg/s basis",
        "Q_retained_L_s": "L/s",
        "Q_retained_accum_L_s": "L/s",
        "Q_out": "kg/s basis",
        "Q_out_L_s": "L/s",
        "Q_brine_remaining_L_s": "L/s",
        "target_freshwater_L_s": "L/s",
        "actual_last_valid_Q_out_L_s": "L/s",
        "design_total_evap_L_s": "L/s",
        "actual_total_retained_L_s": "L/s",
        "module_conversion_fraction": "-",
        "module_conversion_percent": "%",
        "base_Q_in_extraction_L_s": "L/s",
        "base_Q_in_reinjection_L_s": "L/s",
        "base_flow_drop_extraction_to_reinjection_L_s": "L/s",
        "Q_available_at_extraction_L_s": "L/s",
        "Q_module_L_s": "L/s",
        "Q_main_L_s": "L/s",
        "Q_return_L_s": "L/s",
        "Q_return_design_L_s": "L/s",
        "Q_return_actual_L_s": "L/s",
        "recovered_water_L_s": "L/s",
        "skipped_evap_stage_ids": "-",
        "skipped_evap_stage_count": "-",
        "evap_reduction_per_skipped_stage_L_s": "L/s",
        "module_input_volume_L_s": "L/s",
        "module_volume_check_base_L_s": "L/s",
        "module_volume_calibration_factor": "-",
        "module_input_water_kg_s": "kg/s",
        "last_valid_stage_id": "-",
        "E_mm_d": "mm/d",
        "surface_m2": "m2",
        "evap_group": "-",
        "group_type": "-",
        "density_out_kg_L": "kg/L",
        "density_kg_L": "kg/L",
        "volume_out_L": "L",
        "mass_H2O_out_kg": "kg/s basis",
        "liquid_mass_out_kg": "kg/s basis",
        "TDS": "mg/L",
        "TDS_g_L": "g/L",
        "Li_mg_L": "mg/L",
        "Li_g_L": "g/L",
        "phases_precipitated": "-",
        "phase_masses_kg": "kg",
        "total_precipitated_mass_kg": "kg/s basis",
        "total_solids_precipitated_kg_s": "kg/s",
        "total_solids_precipitated_t_d": "t/d",
        "cumulative_solids_kg": "kg/s basis",
        "accumulated_solids_kg_s": "kg/s",
        "accumulated_solids_t_d": "t/d",
        "retained_brine_kg": "kg/s basis",
        "salmuera_retenida": "kg/s basis",
        "retained_fraction_liquid": "-",
        "retention_clamped": "-",
        "mass_H2O_to_next_kg": "kg/s basis",
        "liquid_to_next_kg": "kg/s basis",
        "phase_name": "-",
        "mineral_phase": "-",
        "mineral_family": "-",
        "phase_moles": "mol/s basis",
        "phase_mass_kg": "kg/s basis",
        "precipitated_mass_kg_s": "kg/s",
        "precipitated_mass_t_d": "t/d",
        "accumulated_precipitated_mass_kg_s": "kg/s",
        "accumulated_precipitated_mass_t_d": "t/d",
        "enabled": "-",
        "variable": "-",
        "min": "-",
        "max": "-",
        "step": "-",
        "f_m": "-",
        "TP_PC": "mm/d",
        "TP_H": "mm/d",
        "TP_K": "mm/d",
        "TP_C": "mm/d",
        "TP_L": "mm/d",
        "reagent": "-",
        "reagent_param": "-",
        "lime_dose_mol": "mol/s basis",
        "mg_removal_target": "fraction",
        "mg_removal_achieved": "fraction",
        "liming_iterations": "-",
        "phreeqc_input_file": "-",
        "phreeqc_output_file": "-",
        "selected_output_file": "-",
        "error": "-",
        "warning": "-",
        "Li_final_mg_L_est": "mg/L",
        "Li_concentration_factor": "-",
        "Li_concent_F": "-",
        "Li_recovery_liquid_pct": "%",
        "water_removed_pct": "%",
        "solids_total_kg": "kg/s basis",
        "halite_PC_kg": "kg/s basis",
        "halite_total_kg": "kg/s basis",
        "sulfates_PC_kg": "kg/s basis",
        "sulfates_total_kg": "kg/s basis",
        "Mg_removed_LIM1_pct": "%",
        "Mg_Li_final_molar_ratio": "mol/mol",
        "K_Li_final_molar_ratio": "mol/mol",
        "Mg_Li_ratio": "mol/mol",
        "K_Li_ratio": "mol/mol",
        "SO4_Li_ratio": "mol/mol",
        "Ca_Li_ratio": "mol/mol",
        "solids_per_evaporated_water_kg_kg": "kg/kg",
        "accumulated_precipitation_percent": "%",
        "accumulated_precipitation_by_family_percent": "%",
        "accumulated_halite_precipitated_percent": "%",
        "accumulated_sulfates_precipitated_percent": "%",
        "delta_Li_between_stages_percent": "%",
        "delta_TDS_between_stages_percent": "%",
        "delta_Mg_Li_between_stages_percent": "%",
        "delta_SO4_Li_between_stages_percent": "%",
        "water_evaporated_kg_s": "kg/s",
        "solids_water_ratio_kg_kg": "kg/kg",
        "solid_water_score": "0-100",
        "phase_penalty_score": "0-100",
        "saturation_score": "0-100",
        "chemical_treatment_penalty": "points",
        "Psa_solids_water_score": "0-100",
        "Pf_phase_penalty_score": "0-100",
        "Ps_saturation_score": "0-100",
        "Ptq_chemical_treatment_penalty": "points",
        "solid_water_weight": "-",
        "phase_penalty_weight": "-",
        "saturation_weight": "-",
        "chemical_treatment_penalty_weight": "-",
        "solid_water_contribution": "points",
        "phase_penalty_contribution": "points",
        "saturation_contribution": "points",
        "chemical_treatment_contribution": "points",
        "scaling_risk_indicator": "0-100",
        "scaling_risk_class": "-",
        "precipitated_phases": "phase:kg/s",
        "saturation_indexes_by_phase": "-",
        "max_error": "-",
        "warnings": "-",
        "check": "-",
        "status": "-",
        "details": "-",
        "pH": "-",
        "water_activity": "-",
        "ionic_strength_mol_kgw": "mol/kgw",
        "main_saturation_indexes": "-",
        "precipitation_sequence": "-",
    }

    row = []
    for col in df.columns:
        if col in OUTPUT_SPECIES or col == "SO4/S(6)":
            row.append("mol/kgw")
        elif col.endswith("_molality_out"):
            row.append("mol/kgw")
        elif col.endswith("_mol_out_est") or col.endswith("_mol_to_next_est"):
            row.append("mol/s basis")
        elif col.endswith("_mg_L_est"):
            row.append("mg/L")
        elif col.endswith("_mg_L"):
            row.append("mg/L")
        elif col.endswith("_mol_kgw"):
            row.append("mol/kgw")
        elif col in {"Mg/Li", "K/Li", "SO4/Li"}:
            row.append("mol/mol")
        else:
            row.append(units_map.get(col, "-"))
    return row


def write_sheet_with_units(
    writer,
    df: pd.DataFrame,
    sheet_name: str,
    tab_color: str | None = None,
    blue_recovery_rows: bool = False,
):
    if df is None:
        df = pd.DataFrame()

    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
    ws = writer.sheets[sheet_name]
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    units_row = build_units_row(df)
    for col_idx, unit in enumerate(units_row, start=1):
        ws.cell(row=2, column=col_idx, value=unit)

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), len(str(units_row[col_idx - 1])))
        for value in df.iloc[:, col_idx - 1].tolist():
            value_str = "" if pd.isna(value) else str(value)
            max_len = max(max_len, len(value_str))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    if blue_recovery_rows and "simulation_case" in df.columns:
        case_col = list(df.columns).index("simulation_case")
        fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        font = Font(color="1F4E79")
        for row_idx, value in enumerate(df.iloc[:, case_col].tolist(), start=3):
            if str(value).lower() == "recovery":
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.font = font

    if sheet_name == "scaling_risk_results" and "scaling_risk_class" in df.columns:
        class_col = list(df.columns).index("scaling_risk_class") + 1
        indicator_col = (
            list(df.columns).index("scaling_risk_indicator") + 1
            if "scaling_risk_indicator" in df.columns else None
        )
        risk_styles = {
            "bajo": (PatternFill(fill_type="solid", fgColor="C6EFCE"), Font(color="006100")),
            "medio": (PatternFill(fill_type="solid", fgColor="FFF2CC"), Font(color="7F6000")),
            "alto": (PatternFill(fill_type="solid", fgColor="F4CCCC"), Font(color="9C0006")),
        }
        for row_idx, value in enumerate(df["scaling_risk_class"].tolist(), start=3):
            style = risk_styles.get(str(value).lower())
            if not style:
                continue
            fill, font = style
            for col_idx in [indicator_col, class_col]:
                if col_idx is None:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.font = font

    ws.freeze_panes = "A3"


def add_blank_rows_between_scenarios(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "scenario_id" not in df.columns:
        return df

    parts = []
    grouped = list(df.groupby("scenario_id", sort=False, dropna=False))
    blank = {col: None for col in df.columns}

    for idx, (_, group) in enumerate(grouped):
        parts.append(group)
        if idx < len(grouped) - 1:
            parts.append(pd.DataFrame([blank], columns=df.columns))

    return pd.concat(parts, ignore_index=True) if parts else df


def df_to_json_records(df: pd.DataFrame) -> list[dict]:
    if df is None or df.empty:
        return []
    clean = df.copy()
    clean = clean.where(pd.notna(clean), None)
    return clean.to_dict(orient="records")


def make_json_safe(value):
    if isinstance(value, dict):
        return {str(k): make_json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [make_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [make_json_safe(v) for v in value]
    if pd.isna(value):
        return None
    return value


def copy_plotly_asset(batch_results_dir: Path) -> str | None:
    source = VIEWER_ASSETS / "plotly-2.35.2.min.js"
    if not source.exists():
        return None

    assets_dir = batch_results_dir / "viewer_assets"
    assets_dir.mkdir(parents=True, exist_ok=True)
    target = assets_dir / source.name
    shutil.copy2(source, target)
    return f"viewer_assets/{source.name}"


def write_results_viewer(
    summary_df: pd.DataFrame,
    evaluation_df: pd.DataFrame,
    geochemical_df: pd.DataFrame,
    precipitation_df: pd.DataFrame,
    scaling_risk_df: pd.DataFrame,
    metadata: dict,
    batch_results_dir: Path,
) -> None:
    plotly_src = copy_plotly_asset(batch_results_dir)
    plotly_tag = (
        f'<script src="{plotly_src}"></script>'
        if plotly_src
        else '<script>window.PLOTLY_MISSING = true;</script>'
    )

    payload = {
        "tables": {
            "summary_results": df_to_json_records(summary_df),
            "evaluation_results": df_to_json_records(evaluation_df),
            "geochemical_results": df_to_json_records(geochemical_df),
            "precipitation_by_phase": df_to_json_records(precipitation_df),
            "scaling_risk_results": df_to_json_records(scaling_risk_df),
        },
        "metadata": metadata,
        "scaling_risk_methodology": SCALING_RISK_METHOD_TEXT,
    }
    payload_json = json.dumps(make_json_safe(payload), ensure_ascii=False, allow_nan=False)

    html_template = r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Visor de resultados - Simulador de salmueras</title>
  __PLOTLY_TAG__
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #17202a;
      --muted: #667085;
      --line: #d8dee8;
      --accent: #0969da;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      background: var(--bg);
      color: var(--ink);
    }
    header {
      padding: 16px 22px;
      border-bottom: 1px solid var(--line);
      background: var(--panel);
      position: sticky;
      top: 0;
      z-index: 10;
    }
    h1 { margin: 0 0 4px; font-size: 20px; }
    .sub { color: var(--muted); font-size: 13px; }
    main {
      display: grid;
      grid-template-columns: 320px 1fr;
      min-height: calc(100vh - 72px);
    }
    aside {
      padding: 16px;
      border-right: 1px solid var(--line);
      background: var(--panel);
      overflow: auto;
    }
    section { padding: 16px; overflow: auto; }
    label {
      display: block;
      margin: 12px 0 5px;
      font-size: 12px;
      font-weight: 700;
      color: #344054;
    }
    select, button, input {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px;
      background: #fff;
      color: var(--ink);
      font-size: 13px;
    }
    select[multiple] { min-height: 116px; }
    .checkbox-list {
      width: 100%;
      max-height: 180px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 6px;
      background: #fff;
    }
    .checkbox-list.compact {
      max-height: 130px;
    }
    .check-option {
      display: flex;
      align-items: center;
      gap: 8px;
      margin: 0;
      padding: 4px 2px;
      font-size: 12px;
      font-weight: 400;
      line-height: 1.2;
      color: var(--ink);
    }
    .check-option input {
      width: auto;
      margin: 0;
      flex: 0 0 auto;
    }
    .filter-actions {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 8px;
      margin-top: 6px;
    }
    .filter-actions button {
      margin-top: 0;
      padding: 6px 8px;
      font-size: 12px;
    }
    button {
      cursor: pointer;
      font-weight: 700;
      margin-top: 10px;
    }
    button.primary {
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }
    button.danger {
      color: #b42318;
      background: #fff;
    }
    .hint {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
      margin-top: 10px;
    }
    .stats {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
      gap: 10px;
      margin-bottom: 14px;
    }
    .stat {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
    }
    .stat .k { color: var(--muted); font-size: 12px; }
    .stat .v { font-size: 18px; font-weight: 700; margin-top: 4px; }
    .chart-card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-bottom: 14px;
      overflow: hidden;
    }
    .chart-head {
      display: grid;
      grid-template-columns: repeat(6, minmax(130px, 1fr)) 86px;
      gap: 8px;
      padding: 12px;
      border-bottom: 1px solid var(--line);
      align-items: end;
    }
    .chart-head label { margin-top: 0; }
    .plot { min-height: 430px; }
    .table-wrap {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: auto;
      max-height: 440px;
      margin-top: 14px;
    }
    .methodology {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-top: 14px;
      padding: 12px;
    }
    .methodology summary {
      cursor: pointer;
      font-weight: 700;
    }
    .methodology pre {
      white-space: pre-wrap;
      margin: 12px 0 0;
      color: #344054;
      font-family: Arial, Helvetica, sans-serif;
      font-size: 12px;
      line-height: 1.45;
    }
    table { border-collapse: collapse; width: 100%; font-size: 12px; }
    th, td { border-bottom: 1px solid #edf0f5; padding: 6px 8px; text-align: right; white-space: nowrap; }
    th { position: sticky; top: 0; background: #f9fafb; z-index: 1; color: #344054; }
    th:first-child, td:first-child { text-align: left; }
    .missing {
      padding: 14px;
      margin-bottom: 12px;
      border: 1px solid #fecdca;
      color: #b42318;
      background: #fffbfa;
      border-radius: 8px;
      display: none;
    }
  </style>
</head>
<body>
  <header>
    <h1>Visor de resultados del simulador</h1>
    <div class="sub" id="runInfo"></div>
  </header>
  <main>
    <aside>
      <button class="primary" id="addChartBtn">Añadir gráfico</button>
      <label>Tabla para vista rápida</label>
      <select id="previewTable"></select>
      <label>Filtrar meses</label>
      <div id="globalMonths" class="checkbox-list compact"></div>
      <div class="filter-actions">
        <button type="button" data-check-target="globalMonths" data-check-value="true">Todos</button>
        <button type="button" data-check-target="globalMonths" data-check-value="false">Ninguno</button>
      </div>
      <label>Casos en gráficos</label>
      <div id="globalScenarios" class="checkbox-list"></div>
      <div class="filter-actions">
        <button type="button" data-check-target="globalScenarios" data-check-value="true">Todos</button>
        <button type="button" data-check-target="globalScenarios" data-check-value="false">Ninguno</button>
      </div>
      <button id="applyPreviewBtn">Actualizar tabla</button>
      <p class="hint">
        En los gráficos, el eje X son las etapas del tren y el eje Y es el parámetro elegido.
        La serie "media anual" agrupa los 12 meses con el mismo barrido de temperatura,
        evap_factor, mg_removal y retention_r.
      </p>
    </aside>
    <section>
      <div class="missing" id="plotlyMissing">
        No se encontró Plotly local. El visor se generó, pero los gráficos no podrán dibujarse hasta que exista viewer_assets/plotly-2.35.2.min.js.
      </div>
      <div class="stats" id="stats"></div>
      <div id="charts"></div>
      <details class="methodology">
        <summary>Metodología scaling_risk_indicator</summary>
        <pre id="scalingRiskMethodology"></pre>
      </details>
      <div class="table-wrap"><table id="preview"></table></div>
    </section>
  </main>
  <script id="payload" type="application/json">__PAYLOAD_JSON__</script>
  <script>
    const STAGE_ORDER = ["PC1","PC2","PC3","PC4","PC5","PC6","PC7","PC8","LIM1","H1","H2","H3","K1","C1","L1"];
    const stageRank = Object.fromEntries(STAGE_ORDER.map((s, i) => [s, i]));
    const payload = JSON.parse(document.getElementById("payload").textContent);
    const tables = payload.tables;
    const scenarios = payload.metadata.scenarios || [];
    const scenarioMap = Object.fromEntries(scenarios.map(s => [s.scenario_id, s]));
    const hiddenPreviewColumns = new Set(["stage_type"]);
    let chartCounter = 0;

    function num(v) {
      if (v === null || v === undefined || v === "") return null;
      const n = Number(String(v).replace(",", "."));
      return Number.isFinite(n) ? n : null;
    }
    function fmt(v) {
      const n = num(v);
      if (n === null) return "";
      return Math.abs(n) >= 1000 ? n.toFixed(2) : n.toPrecision(5);
    }
    function annualGroup(sid) {
      const s = scenarioMap[sid] || {};
      return `ANNUAL_MEAN | T=${s.temperature ?? "-"} | E=${s.evap_factor ?? "-"} | Mg=${s.mg_removal ?? "-"} | R=${s.retention_r ?? "-"}`;
    }
    function numericColumns(rows) {
      if (!rows || !rows.length) return [];
      const keys = Object.keys(rows[0]);
      return keys.filter(k => rows.some(r => num(r[k]) !== null));
    }
    function uniqueValues(rows, key) {
      return [...new Set(rows.map(r => r[key]).filter(v => v !== null && v !== undefined && v !== ""))];
    }
    function setOptions(select, values, selectedAll = false) {
      select.innerHTML = "";
      values.forEach(v => {
        const opt = document.createElement("option");
        opt.value = v;
        opt.textContent = v;
        opt.selected = selectedAll;
        select.appendChild(opt);
      });
    }
    function selectedValues(select) {
      return [...select.selectedOptions].map(o => o.value);
    }
    function checkedValues(containerId) {
      return [...document.querySelectorAll(`#${containerId} input[type="checkbox"]:checked`)].map(i => i.value);
    }
    function refreshOutputs() {
      renderPreview();
      document.querySelectorAll(".chart-card").forEach(renderChart);
    }
    function renderCheckboxGroup(containerId, values) {
      const container = document.getElementById(containerId);
      container.innerHTML = "";
      values.forEach(v => {
        const label = document.createElement("label");
        label.className = "check-option";
        const input = document.createElement("input");
        input.type = "checkbox";
        input.value = v;
        input.checked = true;
        input.addEventListener("change", refreshOutputs);
        const text = document.createElement("span");
        text.textContent = v;
        label.appendChild(input);
        label.appendChild(text);
        container.appendChild(label);
      });
    }
    function setAllCheckboxes(containerId, checked) {
      document.querySelectorAll(`#${containerId} input[type="checkbox"]`).forEach(input => {
        input.checked = checked;
      });
      refreshOutputs();
    }
    function filteredRows(tableName) {
      const rows = tables[tableName] || [];
      const months = checkedValues("globalMonths");
      const sids = checkedValues("globalScenarios");
      return rows.filter(r => {
        const hasMonth = r.month !== null && r.month !== undefined && r.month !== "";
        const hasScenario = r.scenario_id !== null && r.scenario_id !== undefined && r.scenario_id !== "";
        const mOk = !hasMonth || months.includes(String(r.month));
        const sOk = !hasScenario || sids.includes(String(r.scenario_id));
        return mOk && sOk;
      });
    }
    function sortByStage(rows) {
      return [...rows].sort((a, b) => {
        const ar = stageRank[a.stage_id] ?? 999;
        const br = stageRank[b.stage_id] ?? 999;
        if (ar !== br) return ar - br;
        return String(a.stage_id).localeCompare(String(b.stage_id));
      });
    }
    function groupRows(rows, yCol, mode) {
      if (mode === "annual_mean") {
        const buckets = new Map();
        rows.forEach(r => {
          if (!r.stage_id) return;
          const y = num(r[yCol]);
          if (y === null) return;
          const key = annualGroup(r.scenario_id) + "||" + r.stage_id;
          if (!buckets.has(key)) buckets.set(key, { series: annualGroup(r.scenario_id), stage_id: r.stage_id, values: [] });
          buckets.get(key).values.push(y);
        });
        const bySeries = new Map();
        buckets.forEach(b => {
          if (!bySeries.has(b.series)) bySeries.set(b.series, []);
          bySeries.get(b.series).push({ stage_id: b.stage_id, value: b.values.reduce((a, c) => a + c, 0) / b.values.length });
        });
        return bySeries;
      }

      const seriesKey = mode;
      const bySeries = new Map();
      rows.forEach(r => {
        if (!r.stage_id) return;
        const y = num(r[yCol]);
        if (y === null) return;
        const series = String(r[seriesKey] ?? "serie");
        if (!bySeries.has(series)) bySeries.set(series, []);
        bySeries.get(series).push({ stage_id: r.stage_id, value: y });
      });
      return bySeries;
    }
    function renderChart(card) {
      const tableName = card.querySelector(".table-select").value;
      const yCol = card.querySelector(".y-select").value;
      const chartType = card.querySelector(".type-select").value;
      const mode = card.querySelector(".series-select").value;
      const rows = filteredRows(tableName);
      const grouped = groupRows(rows, yCol, mode);
      const traces = [];
      grouped.forEach((items, series) => {
        const sorted = sortByStage(items);
        traces.push({
          x: sorted.map(i => i.stage_id),
          y: sorted.map(i => i.value),
          type: chartType === "bar" ? "bar" : "scatter",
          mode: chartType === "bar" ? undefined : "lines+markers",
          name: series,
        });
      });
      const layout = {
        margin: { l: 70, r: 20, t: 40, b: 70 },
        title: `${yCol} por etapa`,
        xaxis: { title: "Etapa", categoryorder: "array", categoryarray: STAGE_ORDER },
        yaxis: { title: yCol },
        barmode: "group",
        legend: { orientation: "h", y: -0.25 },
      };
      Plotly.newPlot(card.querySelector(".plot"), traces, layout, { responsive: true, displaylogo: false });
    }
    function updateYOptions(card) {
      const tableName = card.querySelector(".table-select").value;
      const cols = numericColumns(tables[tableName] || []);
      const y = card.querySelector(".y-select");
      const preferred = cols.includes("Li_mg_L") ? "Li_mg_L" : cols[0];
      setOptions(y, cols);
      y.value = preferred || "";
    }
    function addChart() {
      chartCounter += 1;
      const card = document.createElement("div");
      card.className = "chart-card";
      card.innerHTML = `
        <div class="chart-head">
          <div><label>Tabla</label><select class="table-select"></select></div>
          <div><label>Parámetro Y</label><select class="y-select"></select></div>
          <div><label>Tipo</label><select class="type-select"><option value="line">Puntos + líneas</option><option value="bar">Barras</option></select></div>
          <div><label>Series</label><select class="series-select"><option value="scenario_id">Escenario</option><option value="month">Mes</option><option value="annual_mean">Media anual</option><option value="mineral_phase">Fase mineral</option></select></div>
          <div><label>Actualizar</label><button class="update-btn">Redibujar</button></div>
          <div><label>Duplicar</label><button class="duplicate-btn">Duplicar</button></div>
          <div><button class="danger delete-btn">Quitar</button></div>
        </div>
        <div class="plot" id="plot-${chartCounter}"></div>`;
      document.getElementById("charts").appendChild(card);
      setOptions(card.querySelector(".table-select"), Object.keys(tables));
      card.querySelector(".table-select").value = "summary_results";
      updateYOptions(card);
      card.querySelector(".table-select").addEventListener("change", () => { updateYOptions(card); renderChart(card); });
      card.querySelector(".update-btn").addEventListener("click", () => renderChart(card));
      card.querySelector(".duplicate-btn").addEventListener("click", addChart);
      card.querySelector(".delete-btn").addEventListener("click", () => card.remove());
      card.querySelectorAll("select").forEach(s => s.addEventListener("change", () => renderChart(card)));
      renderChart(card);
    }
    function renderPreview() {
      const tableName = document.getElementById("previewTable").value;
      const rows = filteredRows(tableName).slice(0, 300);
      const table = document.getElementById("preview");
      if (!rows.length) { table.innerHTML = ""; return; }
      const cols = Object.keys(rows[0]).filter(c => !hiddenPreviewColumns.has(c));
      table.innerHTML = `<thead><tr>${cols.map(c => `<th>${c}</th>`).join("")}</tr></thead>` +
        `<tbody>${rows.map(r => `<tr>${cols.map(c => `<td>${num(r[c]) === null ? (r[c] ?? "") : fmt(r[c])}</td>`).join("")}</tr>`).join("")}</tbody>`;
    }
    function renderStats() {
      const meta = payload.metadata || {};
      const l1 = (tables.summary_results || []).filter(r => r.stage_id === "L1").map(r => num(r.Li_mg_L)).filter(v => v !== null);
      const avgLi = l1.length ? l1.reduce((a, c) => a + c, 0) / l1.length : null;
      const risks = (tables.scaling_risk_results || []).map(r => num(r.scaling_risk_indicator)).filter(v => v !== null);
      const maxRisk = risks.length ? Math.max(...risks) : null;
      const cards = [
        ["Run", meta.run_stamp || ""],
        ["Escenarios", scenarios.length],
        ["Li medio en L1", avgLi === null ? "" : `${avgLi.toFixed(1)} mg/L`],
        ["Riesgo max.", maxRisk === null ? "" : maxRisk.toFixed(1)],
      ];
      document.getElementById("stats").innerHTML = cards.map(([k, v]) => `<div class="stat"><div class="k">${k}</div><div class="v">${v}</div></div>`).join("");
    }
    function init() {
      if (window.PLOTLY_MISSING || typeof Plotly === "undefined") {
        document.getElementById("plotlyMissing").style.display = "block";
      }
      document.getElementById("runInfo").textContent = `${payload.metadata.run_stamp || ""} · PHREEQC: ${payload.metadata.phreeqc_database || ""}`;
      setOptions(document.getElementById("previewTable"), Object.keys(tables));
      const allMonths = [...new Set(scenarios.map(s => String(s.month)))].sort((a, b) => Number(a) - Number(b));
      const allScenarios = scenarios.map(s => s.scenario_id);
      renderCheckboxGroup("globalMonths", allMonths);
      renderCheckboxGroup("globalScenarios", allScenarios);
      document.getElementById("scalingRiskMethodology").textContent = payload.scaling_risk_methodology || "";
      document.getElementById("addChartBtn").addEventListener("click", addChart);
      document.getElementById("applyPreviewBtn").addEventListener("click", refreshOutputs);
      document.querySelectorAll("[data-check-target]").forEach(button => {
        button.addEventListener("click", () => {
          setAllCheckboxes(button.dataset.checkTarget, button.dataset.checkValue === "true");
        });
      });
      renderStats();
      addChart();
      renderPreview();
    }
    init();
  </script>
</body>
</html>
"""

    html = (
        html_template
        .replace("__PLOTLY_TAG__", plotly_tag)
        .replace("__PAYLOAD_JSON__", payload_json.replace("</", "<\\/"))
    )
    (batch_results_dir / "results_viewer.html").write_text(html, encoding="utf-8")


def methodology_text_df() -> pd.DataFrame:
    rows = [
        {
            "section": "formula",
            "item": "scaling_risk_indicator",
            "value": (
                "Psa*solid_water_weight + Pf*phase_penalty_weight + "
                "Ps*saturation_weight + Ptq"
            ),
            "description": "Indicador compuesto de riesgo de incrustacion, limitado entre 0 y 100.",
        },
        {
            "section": "weight",
            "item": "Psa_solids_water_score",
            "value": f"{SCALING_RISK_WEIGHTS['solid_water_score']:.6f}",
            "description": "Peso aplicado a la puntuacion solidos/agua evaporada.",
        },
        {
            "section": "Psa_rule",
            "item": "water_evaporated_kg_s <= 0",
            "value": "Psa = 100",
            "description": "Si una etapa no evapora agua, la puntuacion solidos/agua se fuerza al maximo.",
        },
        {
            "section": "weight",
            "item": "Pf_phase_penalty_score",
            "value": f"{SCALING_RISK_WEIGHTS['phase_penalty_score']:.6f}",
            "description": "Peso aplicado a la puntuacion por tipo de fase mineral.",
        },
        {
            "section": "weight",
            "item": "Ps_saturation_score",
            "value": f"{SCALING_RISK_WEIGHTS['saturation_score']:.6f}",
            "description": "Peso aplicado a la puntuacion de saturacion.",
        },
        {
            "section": "penalty",
            "item": "Ptq_chemical_treatment_penalty",
            "value": "direct addition",
            "description": "Penalizacion adicional por condiciones quimicas u operativas criticas.",
        },
        {
            "section": "classification",
            "item": "Bajo",
            "value": "< 35",
            "description": "Riesgo bajo de incrustacion.",
        },
        {
            "section": "classification",
            "item": "Medio",
            "value": "35 <= indicator < 60",
            "description": "Riesgo medio de incrustacion.",
        },
        {
            "section": "classification",
            "item": "Alto",
            "value": ">= 60",
            "description": "Riesgo alto de incrustacion.",
        },
        {
            "section": "saturation",
            "item": "SI >= 0",
            "value": "+20",
            "description": "Fase saturada o con precipitacion posible.",
        },
        {
            "section": "saturation",
            "item": "-0.2 <= SI < 0",
            "value": "+10",
            "description": "Fase proxima a saturacion.",
        },
        {
            "section": "saturation",
            "item": "SI < -0.2",
            "value": "0",
            "description": "Sin penalizacion relevante.",
        },
        {
            "section": "chemical_penalty",
            "item": "LIM or chemical_treatment",
            "value": "+10",
            "description": "Etapa de liming o tratamiento quimico.",
        },
        {
            "section": "chemical_penalty",
            "item": "Brucite precipitated",
            "value": "+10",
            "description": "Precipitacion de brucita.",
        },
        {
            "section": "chemical_penalty",
            "item": "Gypsum or Anhydrite precipitated",
            "value": "+5",
            "description": "Precipitacion de yeso o anhidrita.",
        },
    ]
    rows.extend(
        {
            "section": "phase_penalty",
            "item": phase,
            "value": penalty,
            "description": "Penalizacion base por fase mineral.",
        }
        for phase, penalty in sorted(PHASE_SCALING_PENALTIES.items())
    )
    rows.append({
        "section": "phase_penalty",
        "item": "Unknown phase",
        "value": DEFAULT_UNKNOWN_PHASE_PENALTY,
        "description": "Valor provisional para fases no clasificadas.",
    })
    return pd.DataFrame(rows)


def export_methodological_outputs(
    summary_df: pd.DataFrame,
    evaluation_df: pd.DataFrame,
    geochemical_df: pd.DataFrame,
    precipitation_df: pd.DataFrame,
    scaling_risk_df: pd.DataFrame,
    metadata: dict,
    validation_df: pd.DataFrame,
    batch_results_dir: Path,
) -> None:
    tables = {
        "summary_results": summary_df,
        "evaluation_results": evaluation_df,
        "geochemical_results": geochemical_df,
        "precipitation_by_phase": precipitation_df,
        "scaling_risk_results": scaling_risk_df,
        "validation": validation_df,
    }

    for name, df in tables.items():
        csv_path = batch_results_dir / f"{name}.csv"
        df.to_csv(csv_path, index=False)

    scenario_metadata_df = pd.DataFrame(metadata.get("scenarios", []))
    scaling_risk_method_df = methodology_text_df()
    run_info_df = pd.DataFrame([{
        "run_stamp": metadata.get("run_stamp"),
        "generated_at": metadata.get("generated_at"),
        "root": metadata.get("root"),
        "input_dir": metadata.get("input_dir"),
        "phreeqc_exe": metadata.get("phreeqc_exe"),
        "phreeqc_database": metadata.get("phreeqc_database"),
    }])

    workbook_path = batch_results_dir / "simulator_results.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(summary_df), "summary_results")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(evaluation_df), "evaluation_results")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(geochemical_df), "geochemical_results")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(precipitation_df), "precipitation_by_phase")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(scaling_risk_df), "scaling_risk_results")
        write_sheet_with_units(writer, scaling_risk_method_df, "scaling_risk_method")
        write_sheet_with_units(writer, validation_df, "validation")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(scenario_metadata_df), "scenario_metadata")
        write_sheet_with_units(writer, run_info_df, "run_info")

    (batch_results_dir / "scenario_metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    (batch_results_dir / "scaling_risk_methodology.txt").write_text(
        SCALING_RISK_METHOD_TEXT,
        encoding="utf-8",
    )
    write_results_viewer(
        summary_df,
        evaluation_df,
        geochemical_df,
        precipitation_df,
        scaling_risk_df,
        metadata,
        batch_results_dir,
    )


# =========================================================
# MODULO AUXILIAR DE RECUPERACION DE AGUA
# =========================================================

def read_recovery_control(path: Path = RECOVERY_CONTROL_FILE) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"No existe recovery_control.txt: {path}")

    df = pd.read_csv(path, sep=r"[;\t]", engine="python")
    df.columns = [str(c).strip() for c in df.columns]
    if "variable" not in df.columns or "data" not in df.columns:
        raise ValueError("recovery_control.txt debe contener las columnas 'variable' y 'data'.")

    values = {
        str(row["variable"]).strip(): row["data"]
        for _, row in df.iterrows()
        if str(row.get("variable", "")).strip()
    }
    required = [
        "target_freshwater_L_s",
        "heating_temperature_C",
        "extraction_stage_id",
        "reinjection_stage_id",
    ]
    missing = [name for name in required if name not in values]
    if missing:
        raise ValueError(f"Faltan variables en recovery_control.txt: {', '.join(missing)}")

    control = {
        "target_freshwater_L_s": float(values["target_freshwater_L_s"]),
        "heating_temperature_C": float(values["heating_temperature_C"]),
        "extraction_stage_id": str(values["extraction_stage_id"]).strip(),
        "reinjection_stage_id": str(values["reinjection_stage_id"]).strip(),
    }

    if control["target_freshwater_L_s"] <= 0:
        raise ValueError("target_freshwater_L_s debe ser > 0.")

    return control


def validate_recovery_control(control: dict, stages: pd.DataFrame) -> dict:
    stage_ids = [str(s).strip() for s in stages["stage_id"].tolist()]
    stage_lookup = {sid.upper(): sid for sid in stage_ids}

    extraction_key = control["extraction_stage_id"].upper()
    reinjection_key = control["reinjection_stage_id"].upper()

    if extraction_key not in stage_lookup:
        raise ValueError(f"extraction_stage_id no existe en stages_base: {control['extraction_stage_id']}")
    if reinjection_key not in stage_lookup:
        raise ValueError(f"reinjection_stage_id no existe en stages_base: {control['reinjection_stage_id']}")

    stage_order = {sid.upper(): idx for idx, sid in enumerate(stage_ids)}
    if stage_order[reinjection_key] <= stage_order[extraction_key]:
        raise ValueError("reinjection_stage_id debe estar aguas abajo de extraction_stage_id.")

    checked = control.copy()
    checked["extraction_stage_id"] = stage_lookup[extraction_key]
    checked["reinjection_stage_id"] = stage_lookup[reinjection_key]
    return checked


def skipped_pond_stage_ids(stages: pd.DataFrame, extraction_stage_id: str, reinjection_stage_id: str) -> list[str]:
    stage_ids = [str(s) for s in stages["stage_id"].tolist()]
    lookup = {sid.upper(): idx for idx, sid in enumerate(stage_ids)}
    start = lookup[str(extraction_stage_id).upper()]
    end = lookup[str(reinjection_stage_id).upper()]
    skipped = []
    for _, row in stages.iloc[start:end].iterrows():
        if str(row.get("stage_type", "")).lower() == "pond":
            skipped.append(str(row["stage_id"]))
    return skipped


def derive_recovery_control_from_base(
    control: dict,
    base_summary: pd.DataFrame,
    stages: pd.DataFrame,
) -> dict:
    if base_summary is None or base_summary.empty:
        raise ValueError("No hay summary_results del caso base para calcular la conversion del modulo.")

    extraction_stage_id = str(control["extraction_stage_id"])
    reinjection_stage_id = str(control["reinjection_stage_id"])

    def base_q_in(stage_id: str) -> float:
        rows = base_summary[base_summary["stage_id"].astype(str).str.upper() == stage_id.upper()]
        if rows.empty:
            raise ValueError(f"No se encontro {stage_id} en summary_results del caso base.")
        value = pd.to_numeric(rows.iloc[0].get("Q_in_L_s"), errors="coerce")
        if pd.isna(value):
            raise ValueError(f"Q_in_L_s no es numerico para {stage_id} en el caso base.")
        return float(value)

    base_q_in_extraction = base_q_in(extraction_stage_id)
    base_q_in_reinjection = base_q_in(reinjection_stage_id)
    base_flow_drop = base_q_in_extraction - base_q_in_reinjection

    if base_q_in_extraction <= 0:
        raise ValueError("Q_in_L_s en la etapa de extraccion debe ser positivo.")
    if base_flow_drop <= 0:
        raise ValueError(
            "La conversion del modulo no puede calcularse porque Q_in no disminuye "
            f"entre {extraction_stage_id} y {reinjection_stage_id}."
        )

    module_conversion_fraction = base_flow_drop / base_q_in_extraction
    if not 0 < module_conversion_fraction < 1:
        raise ValueError(
            "La conversion calculada desde el caso base debe estar entre 0 y 1: "
            f"{module_conversion_fraction:.6g}"
        )

    q_module_l_s = float(control["target_freshwater_L_s"]) / module_conversion_fraction
    q_return_l_s = q_module_l_s - float(control["target_freshwater_L_s"])
    if q_return_l_s <= 0:
        raise ValueError("Q_return_L_s calculado debe ser positivo.")

    skipped_stages = skipped_pond_stage_ids(stages, extraction_stage_id, reinjection_stage_id)
    if not skipped_stages:
        raise ValueError("No hay piscinas saltadas entre extraccion y reinyeccion.")
    evap_reduction_per_stage = float(control["target_freshwater_L_s"]) / len(skipped_stages)
    insufficient_evap = []
    for stage_id in skipped_stages:
        rows = base_summary[base_summary["stage_id"].astype(str).str.upper() == stage_id.upper()]
        q_evap = pd.to_numeric(rows.iloc[0].get("Q_evap_L_s"), errors="coerce") if not rows.empty else None
        if q_evap is None or pd.isna(q_evap) or float(q_evap) < evap_reduction_per_stage:
            insufficient_evap.append(stage_id)
    if insufficient_evap:
        raise ValueError(
            "La reduccion uniforme de evaporacion supera la evaporacion base en: "
            + ", ".join(insufficient_evap)
        )

    checked = control.copy()
    checked.update({
        "module_conversion_fraction": module_conversion_fraction,
        "module_conversion_percent": 100.0 * module_conversion_fraction,
        "base_Q_in_extraction_L_s": base_q_in_extraction,
        "base_Q_in_reinjection_L_s": base_q_in_reinjection,
        "base_flow_drop_extraction_to_reinjection_L_s": base_flow_drop,
        "Q_module_L_s": q_module_l_s,
        "Q_return_L_s": q_return_l_s,
        "skipped_evap_stage_ids": "|".join(skipped_stages),
        "skipped_evap_stage_count": len(skipped_stages),
        "evap_reduction_per_skipped_stage_L_s": evap_reduction_per_stage,
    })
    return checked


def solution_liquid_l_s(solution: dict) -> float:
    return estimate_liquid_mass_kg(float(solution["water_kg"]), solution_molalities(solution))


def scale_solution_water(solution: dict, fraction: float) -> dict:
    if fraction < -1e-12:
        raise ValueError(f"La fraccion de escalado no puede ser negativa: {fraction}")
    scaled = solution.copy()
    scaled["water_kg"] = float(solution["water_kg"]) * max(float(fraction), 0.0)
    for optional_key in [
        "original_water_kg",
    ]:
        if optional_key in scaled and scaled[optional_key] is not None and pd.notna(scaled[optional_key]):
            scaled[optional_key] = float(scaled[optional_key]) * max(float(fraction), 0.0)
    return scaled


def mix_solutions(main_solution: dict, return_solution: dict, mixed_temp_C: float | None = None) -> dict:
    main_water = float(main_solution["water_kg"])
    return_water = float(return_solution["water_kg"])
    total_water = main_water + return_water
    if total_water <= 1e-12:
        raise ValueError("La mezcla no contiene agua suficiente para continuar.")

    mixed = main_solution.copy()
    mixed["water_kg"] = total_water
    mixed["temp_C"] = float(mixed_temp_C if mixed_temp_C is not None else main_solution["temp_C"])
    mixed["pH"] = (
        float(main_solution.get("pH", 7.0)) * main_water +
        float(return_solution.get("pH", 7.0)) * return_water
    ) / total_water

    for sp in OUTPUT_SPECIES:
        main_moles = get_solution_molality(main_solution, sp) * main_water
        return_moles = get_solution_molality(return_solution, sp) * return_water
        mixed[sp] = (main_moles + return_moles) / total_water

    return mixed


def build_module_stage_row(extraction_stage_row: pd.Series, control: dict) -> pd.Series:
    # Caja negra conceptual: PHREEQC concentra la corriente derivada retirando agua pura.
    module_row = extraction_stage_row.copy()
    module_row["stage_id"] = f"MODULE_{extraction_stage_row['stage_id']}"
    module_row["stage_type"] = "pond"
    module_row["Q_evap"] = float(control["target_freshwater_L_s"])
    module_row["Q_evap_kg_s"] = float(control["target_freshwater_L_s"]) * RHO_WATER_KG_L
    module_row["E_mm_d"] = 0.0
    module_row["surface_m2"] = 0.0
    module_row["reagent"] = ""
    module_row["reagent_param"] = 0.0
    module_row["temp_C"] = float(control["heating_temperature_C"])
    return module_row


def run_solution_volume_check(solution: dict, check_folder: Path, label: str) -> float:
    check_folder.mkdir(parents=True, exist_ok=True)
    safe_label = re.sub(r"[^A-Za-z0-9_]+", "_", label)
    input_file = check_folder / f"{safe_label}_volume_check.pqi"
    output_file = check_folder / f"{safe_label}_volume_check.pqo"
    selected_file = check_folder / f"{safe_label}_volume_check_selected.txt"
    input_text = "\n\n".join([
        build_selected_output([], selected_file.name),
        build_solution_block(1, solution, f"{safe_label}_VOLUME_CHECK"),
    ])
    input_file.write_text(input_text, encoding="utf-8")
    run_phreeqc(input_file, output_file, check_folder)
    props = parse_pqo_properties(output_file)
    volume_l = props.get("volume_L")
    if volume_l is None or pd.isna(volume_l) or float(volume_l) <= 0:
        raise ValueError(f"No se pudo calcular el volumen PHREEQC de {label}.")
    return float(volume_l)


def calibrate_solution_to_phreeqc_volume(
    solution: dict,
    target_volume_l_s: float,
    check_folder: Path,
    label: str,
) -> tuple[dict, dict]:
    base_volume_l_s = run_solution_volume_check(solution, check_folder, f"{label}_base")
    scale = float(target_volume_l_s) / base_volume_l_s
    calibrated = scale_solution_water(solution, scale)

    # A molalidad y temperatura constantes, el volumen PHREEQC escala linealmente con kg de agua.
    calibrated["module_input_volume_L_s"] = float(target_volume_l_s)
    calibrated["module_volume_check_base_L_s"] = base_volume_l_s
    calibrated["module_volume_calibration_factor"] = scale
    calibrated["module_input_water_kg_s"] = float(calibrated["water_kg"])

    return calibrated, {
        "module_input_volume_L_s": float(target_volume_l_s),
        "module_volume_check_base_L_s": base_volume_l_s,
        "module_volume_calibration_factor": scale,
        "module_input_water_kg_s": float(calibrated["water_kg"]),
    }


def module_failure_row(
    scenario: pd.Series,
    control: dict,
    q_available_l_s: float | None,
    q_module_l_s: float,
    q_main_l_s: float | None,
    error_message: str,
) -> dict:
    return {
        "scenario_id": scenario["scenario_id"],
        "month": scenario["month"],
        "extraction_stage_id": control["extraction_stage_id"],
        "reinjection_stage_id": control["reinjection_stage_id"],
        "heating_temperature_C": control["heating_temperature_C"],
        "Q_available_at_extraction_L_s": q_available_l_s,
        "Q_module_L_s": q_module_l_s,
        "Q_main_L_s": q_main_l_s,
        "target_freshwater_L_s": control["target_freshwater_L_s"],
        "module_conversion_fraction": control.get("module_conversion_fraction"),
        "module_conversion_percent": control.get("module_conversion_percent"),
        "base_Q_in_extraction_L_s": control.get("base_Q_in_extraction_L_s"),
        "base_Q_in_reinjection_L_s": control.get("base_Q_in_reinjection_L_s"),
        "base_flow_drop_extraction_to_reinjection_L_s": control.get("base_flow_drop_extraction_to_reinjection_L_s"),
        "skipped_evap_stage_ids": control.get("skipped_evap_stage_ids"),
        "skipped_evap_stage_count": control.get("skipped_evap_stage_count"),
        "evap_reduction_per_skipped_stage_L_s": control.get("evap_reduction_per_skipped_stage_L_s"),
        "bypass_fraction_percent": None if not q_available_l_s else 100.0 * q_module_l_s / q_available_l_s,
        "recovered_water_L_s": control["target_freshwater_L_s"],
        "Q_return_L_s": control.get("Q_return_L_s"),
        "Q_return_design_L_s": control.get("Q_return_L_s"),
        "Q_return_actual_L_s": None,
        "module_input_volume_L_s": None,
        "module_volume_check_base_L_s": None,
        "module_volume_calibration_factor": None,
        "module_input_water_kg_s": None,
        "module_phase": None,
        "module_mineral_family": None,
        "module_precipitated_mass_kg_s": 0.0,
        "module_precipitated_mass_t_d": 0.0,
        "module_accumulated_precipitated_mass_kg_s": 0.0,
        "module_solids_total_kg_s": 0.0,
        "retained_brine_kg_s": 0.0,
        "TSS_module_g_L": None,
        "TDS_return_g_L": None,
        "Li_return_mg_L": None,
        "saturation_index_phase": None,
        "precipitated_phases": "",
        "saturation_indexes_by_phase": "",
        "module_status": "FAIL",
        "module_error_message": error_message,
    }


def build_module_result_rows(
    scenario: pd.Series,
    control: dict,
    q_available_l_s: float,
    q_module_l_s: float,
    q_main_l_s: float,
    clean_row: dict,
    phase_df: pd.DataFrame,
    q_return_actual_l_s: float | None,
    calibration_info: dict | None = None,
) -> list[dict]:
    q_return_l_s = float(control["Q_return_L_s"])
    solids_total = float(clean_row.get("total_precipitated_mass_kg", 0.0) or 0.0)
    retained_brine = float(clean_row.get("retained_brine_kg", 0.0) or 0.0)
    si_text = str(clean_row.get("main_saturation_indexes", "") or "")
    si_lookup = parse_saturation_indexes(si_text)

    tds_return_g_l = None
    if clean_row.get("TDS") is not None and pd.notna(clean_row.get("TDS")):
        tds_return_g_l = float(clean_row["TDS"]) / 1000.0

    base = {
        "scenario_id": scenario["scenario_id"],
        "month": scenario["month"],
        "extraction_stage_id": control["extraction_stage_id"],
        "reinjection_stage_id": control["reinjection_stage_id"],
        "heating_temperature_C": control["heating_temperature_C"],
        "Q_available_at_extraction_L_s": q_available_l_s,
        "Q_module_L_s": q_module_l_s,
        "Q_main_L_s": q_main_l_s,
        "target_freshwater_L_s": control["target_freshwater_L_s"],
        "module_conversion_fraction": control.get("module_conversion_fraction"),
        "module_conversion_percent": control.get("module_conversion_percent"),
        "base_Q_in_extraction_L_s": control.get("base_Q_in_extraction_L_s"),
        "base_Q_in_reinjection_L_s": control.get("base_Q_in_reinjection_L_s"),
        "base_flow_drop_extraction_to_reinjection_L_s": control.get("base_flow_drop_extraction_to_reinjection_L_s"),
        "skipped_evap_stage_ids": control.get("skipped_evap_stage_ids"),
        "skipped_evap_stage_count": control.get("skipped_evap_stage_count"),
        "evap_reduction_per_skipped_stage_L_s": control.get("evap_reduction_per_skipped_stage_L_s"),
        "bypass_fraction_percent": 100.0 * q_module_l_s / q_available_l_s if q_available_l_s > 0 else None,
        "recovered_water_L_s": control["target_freshwater_L_s"],
        "Q_return_L_s": q_return_l_s,
        "Q_return_design_L_s": q_return_l_s,
        "Q_return_actual_L_s": q_return_actual_l_s,
        "module_input_volume_L_s": None if calibration_info is None else calibration_info.get("module_input_volume_L_s"),
        "module_volume_check_base_L_s": None if calibration_info is None else calibration_info.get("module_volume_check_base_L_s"),
        "module_volume_calibration_factor": None if calibration_info is None else calibration_info.get("module_volume_calibration_factor"),
        "module_input_water_kg_s": None if calibration_info is None else calibration_info.get("module_input_water_kg_s"),
        "module_solids_total_kg_s": solids_total,
        "retained_brine_kg_s": retained_brine,
        "TSS_module_g_L": 1000.0 * solids_total / q_return_l_s if q_return_l_s > 0 else None,
        "TDS_return_g_L": tds_return_g_l,
        "Li_return_mg_L": clean_row.get("Li_mg_L_est"),
        "precipitated_phases": clean_row.get("phases_precipitated", ""),
        "saturation_indexes_by_phase": si_text,
        "module_status": "OK",
        "module_error_message": "",
    }

    if phase_df.empty:
        row = base.copy()
        row.update({
            "module_phase": None,
            "module_mineral_family": None,
            "module_precipitated_mass_kg_s": 0.0,
            "module_precipitated_mass_t_d": 0.0,
            "module_accumulated_precipitated_mass_kg_s": 0.0,
            "saturation_index_phase": None,
        })
        return [row]

    rows = []
    accumulated_by_phase: dict[str, float] = {}
    for _, phase_row in phase_df.iterrows():
        phase = str(phase_row["phase_name"])
        mass = float(phase_row.get("phase_mass_kg", 0.0) or 0.0)
        accumulated_by_phase[phase] = accumulated_by_phase.get(phase, 0.0) + mass
        row = base.copy()
        row.update({
            "module_phase": phase,
            "module_mineral_family": mineral_family(phase),
            "module_precipitated_mass_kg_s": mass,
            "module_precipitated_mass_t_d": kg_s_to_t_d(mass),
            "module_accumulated_precipitated_mass_kg_s": accumulated_by_phase[phase],
            "saturation_index_phase": si_lookup.get(phase),
        })
        rows.append(row)
    return rows


def run_recovery_module(
    scenario: pd.Series,
    control: dict,
    extraction_stage_order: int,
    extraction_stage_row: pd.Series,
    extracted_solution: dict,
    q_available_l_s: float,
    q_module_l_s: float,
    q_main_l_s: float,
    module_runs_dir: Path,
    calibration_info: dict | None = None,
) -> tuple[dict, float | None, pd.DataFrame, list[dict]]:
    scenario_id = str(scenario["scenario_id"])
    module_folder = module_runs_dir / scenario_id / f"{extraction_stage_order:02d}_{control['extraction_stage_id']}_module"
    module_folder.mkdir(parents=True, exist_ok=True)

    module_solution = extracted_solution.copy()
    module_solution["temp_C"] = float(control["heating_temperature_C"])
    module_stage_row = build_module_stage_row(extraction_stage_row, control)

    module_stage_row = prepare_stage_row_for_run(module_stage_row, module_solution)
    react_row, pqo_props, input_file, output_file, selected_file = execute_stage_phreeqc(
        stage_row=module_stage_row,
        solution_in=module_solution,
        stage_folder=module_folder,
        input_name="module_input.pqi",
        output_name="module_output.pqo",
        selected_name="module_selected_output.txt",
    )

    phase_df = build_stage_phase_summary(react_row, module_stage_row, extraction_stage_order, scenario_id)
    clean_row, return_solution = build_stage_clean_row(
        react_row=react_row,
        stage_row=module_stage_row,
        stage_order=extraction_stage_order,
        scenario_id=scenario_id,
        pqo_props=pqo_props,
        phase_df=phase_df,
        solution_in=module_solution,
        input_file=input_file,
        output_file=output_file,
        selected_file=selected_file,
    )

    # La temperatura interna del modulo no se transfiere a la etapa de reinyeccion.
    return_solution["temp_C"] = float(extracted_solution["temp_C"])
    q_return_actual_l_s = brine_remaining_l_s(clean_row)
    rows = build_module_result_rows(
        scenario=scenario,
        control=control,
        q_available_l_s=q_available_l_s,
        q_module_l_s=q_module_l_s,
        q_main_l_s=q_main_l_s,
        clean_row=clean_row,
        phase_df=phase_df,
        q_return_actual_l_s=q_return_actual_l_s,
        calibration_info=calibration_info,
    )
    return return_solution, q_return_actual_l_s, phase_df, rows


def run_recovery_case(
    scenario: pd.Series,
    feed: dict,
    stages: pd.DataFrame,
    control: dict,
    batch_runs_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    scenario_id = str(scenario["scenario_id"])
    recovery_run_dir = batch_runs_dir / "recovery" / scenario_id
    module_runs_dir = batch_runs_dir / "module"
    recovery_run_dir.mkdir(parents=True, exist_ok=True)
    module_runs_dir.mkdir(parents=True, exist_ok=True)

    current_solution = build_initial_solution(feed)
    feed_check_folder = recovery_run_dir / "00_FEED_volume_check"
    current_solution["feed_phreeqc_volume_L_s"] = run_solution_volume_check(
        current_solution,
        feed_check_folder,
        "feed",
    )
    phase_tables = []
    clean_rows = [build_feed_row(current_solution, scenario_id)]
    module_rows = []
    scenario_status = "OK"
    cumulative_solids_kg = 0.0
    module_return_solution = None
    module_return_actual_l_s = None
    q_in_override_l_s = None
    latest_visible_q_out_l_s = None
    evap_adjustment_active = False
    bypass_fraction_percent = None
    module_evap_external_l_s = None
    module_recovered_water_l_s = None
    module_extra_recovered_water_l_s = None
    avoided_visible_evap_l_s = 0.0

    extraction_stage_id = str(control["extraction_stage_id"]).upper()
    reinjection_stage_id = str(control["reinjection_stage_id"]).upper()
    q_module_l_s = float(control["Q_module_L_s"])

    for stage_order, (_, stage_row) in enumerate(stages.iterrows(), start=1):
        stage_id_upper = str(stage_row["stage_id"]).upper()

        try:
            if stage_id_upper == extraction_stage_id:
                split_check_folder = module_runs_dir / scenario_id / f"{stage_order:02d}_{control['extraction_stage_id']}_split_check"
                q_available_l_s = run_solution_volume_check(
                    current_solution,
                    split_check_folder,
                    "available_at_extraction",
                )
                q_main_l_s = q_available_l_s - q_module_l_s

                if q_module_l_s > q_available_l_s:
                    message = (
                        f"Q_module_L_s={q_module_l_s:.6g} supera el caudal disponible "
                        f"en {stage_row['stage_id']} ({q_available_l_s:.6g} L/s)."
                    )
                    module_rows.append(module_failure_row(
                        scenario=scenario,
                        control=control,
                        q_available_l_s=q_available_l_s,
                        q_module_l_s=q_module_l_s,
                        q_main_l_s=q_main_l_s,
                        error_message=message,
                    ))
                    raise ValueError(message)

                if q_main_l_s <= 0:
                    message = f"Q_main_L_s no es positivo en {stage_row['stage_id']}: {q_main_l_s:.6g}"
                    module_rows.append(module_failure_row(
                        scenario=scenario,
                        control=control,
                        q_available_l_s=q_available_l_s,
                        q_module_l_s=q_module_l_s,
                        q_main_l_s=q_main_l_s,
                        error_message=message,
                    ))
                    raise ValueError(message)

                module_seed_solution = current_solution.copy()
                module_seed_solution["temp_C"] = float(control["heating_temperature_C"])
                extracted_solution, calibration_info = calibrate_solution_to_phreeqc_volume(
                    solution=module_seed_solution,
                    target_volume_l_s=q_module_l_s,
                    check_folder=split_check_folder,
                    label="module_inlet",
                )

                module_water_kg = float(extracted_solution["water_kg"])
                source_water_kg = float(current_solution["water_kg"])
                if module_water_kg >= source_water_kg:
                    message = (
                        f"El modulo requiere {module_water_kg:.6g} kg/s de agua, pero la entrada "
                        f"de {stage_row['stage_id']} solo contiene {source_water_kg:.6g} kg/s."
                    )
                    module_rows.append(module_failure_row(
                        scenario=scenario,
                        control=control,
                        q_available_l_s=q_available_l_s,
                        q_module_l_s=q_module_l_s,
                        q_main_l_s=q_main_l_s,
                        error_message=message,
                    ))
                    raise ValueError(message)

                current_solution = current_solution.copy()
                current_solution["water_kg"] = source_water_kg - module_water_kg
                q_in_override_l_s = q_main_l_s
                evap_adjustment_active = True
                bypass_fraction_percent = 100.0 * q_module_l_s / q_available_l_s

                try:
                    module_return_solution, module_return_actual_l_s, _, rows = run_recovery_module(
                        scenario=scenario,
                        control=control,
                        extraction_stage_order=stage_order,
                        extraction_stage_row=stage_row,
                        extracted_solution=extracted_solution,
                        q_available_l_s=q_available_l_s,
                        q_module_l_s=q_module_l_s,
                        q_main_l_s=q_main_l_s,
                        module_runs_dir=module_runs_dir,
                        calibration_info=calibration_info,
                    )
                    module_rows.extend(rows)
                except Exception as module_exc:
                    module_rows.append(module_failure_row(
                        scenario=scenario,
                        control=control,
                        q_available_l_s=q_available_l_s,
                        q_module_l_s=q_module_l_s,
                        q_main_l_s=q_main_l_s,
                        error_message=str(module_exc),
                    ))
                    raise RuntimeError(f"Fallo en modulo recovery: {module_exc}") from module_exc

            if stage_id_upper == reinjection_stage_id:
                if module_return_solution is None:
                    raise RuntimeError("No existe salmuera de retorno del modulo antes de la reinyeccion.")
                current_solution = mix_solutions(
                    main_solution=current_solution,
                    return_solution=module_return_solution,
                    mixed_temp_C=float(current_solution["temp_C"]),
                )
                q_before_reinjection_l_s = (
                    latest_visible_q_out_l_s
                    if latest_visible_q_out_l_s is not None
                    else solution_liquid_l_s(current_solution)
                )
                q_in_override_l_s = q_before_reinjection_l_s + float(module_return_actual_l_s or 0.0)
                module_evap_external_l_s = avoided_visible_evap_l_s
                module_recovered_water_l_s = float(control["target_freshwater_L_s"])
                module_extra_recovered_water_l_s = module_recovered_water_l_s - module_evap_external_l_s
                evap_adjustment_active = False
                bypass_fraction_percent = None

            stage_row_for_run = stage_row.copy()
            stage_row_for_run["Q_evap_external_L_s"] = float(module_evap_external_l_s or 0.0)
            stage_row_for_run["Q_module_recovered_water_L_s"] = float(module_recovered_water_l_s or 0.0)
            stage_row_for_run["Q_module_extra_recovered_water_L_s"] = float(module_extra_recovered_water_l_s or 0.0)
            if (
                evap_adjustment_active and
                str(stage_row_for_run.get("stage_type", "")).lower() == "pond"
            ):
                original_q_evap = float(stage_row_for_run.get("Q_evap", 0.0))
                evap_recovered_l_s = min(
                    float(control.get("evap_reduction_per_skipped_stage_L_s", 0.0) or 0.0),
                    original_q_evap,
                )
                scaled_q_evap = max(original_q_evap - evap_recovered_l_s, 0.0)
                stage_scale_factor = scaled_q_evap / original_q_evap if original_q_evap > 0 else 1.0
                evap_reduction_percent = (
                    100.0 * evap_recovered_l_s / original_q_evap
                    if original_q_evap > 0 else 0.0
                )
                avoided_visible_evap_l_s += evap_recovered_l_s
                stage_row_for_run["Q_evap_original_L_s"] = original_q_evap
                stage_row_for_run["Q_evap"] = scaled_q_evap
                stage_row_for_run["Q_evap_kg_s"] = scaled_q_evap * RHO_WATER_KG_L
                stage_row_for_run["recovery_evap_scale_factor"] = float(stage_scale_factor)
                stage_row_for_run["bypass_fraction_percent"] = float(bypass_fraction_percent or 0.0)
                stage_row_for_run["evap_recovered_by_module_L_s"] = evap_recovered_l_s
                stage_row_for_run["evap_reduction_percent"] = evap_reduction_percent
                stage_row_for_run["warning"] = " | ".join(
                    part for part in [
                        str(stage_row_for_run.get("warning", "") or "").strip(),
                        (
                            "Evaporacion reducida por agua recuperada del modulo: "
                            f"Q_evap={original_q_evap:.6g}-{evap_recovered_l_s:.6g} "
                            f"({evap_reduction_percent:.3g}%)"
                        ),
                    ]
                    if part
                )
            else:
                stage_row_for_run["Q_evap_original_L_s"] = float(stage_row_for_run.get("Q_evap", 0.0))
                stage_row_for_run["recovery_evap_scale_factor"] = 1.0
                stage_row_for_run["bypass_fraction_percent"] = None
                stage_row_for_run["evap_recovered_by_module_L_s"] = 0.0
                stage_row_for_run["evap_reduction_percent"] = 0.0
            module_evap_external_l_s = None
            module_recovered_water_l_s = None
            module_extra_recovered_water_l_s = None
            if stage_id_upper == reinjection_stage_id:
                avoided_visible_evap_l_s = 0.0

            phase_df, clean_row, next_solution = run_stage(
                stage_order=stage_order,
                stage_row=stage_row_for_run,
                solution_in=current_solution,
                scenario_id=scenario_id,
                scenario_run_dir=recovery_run_dir,
            )

            if not phase_df.empty:
                phase_tables.append(phase_df)

            cumulative_solids_kg += float(clean_row.get("total_precipitated_mass_kg", 0.0) or 0.0)
            clean_row["cumulative_solids_kg"] = cumulative_solids_kg
            if q_in_override_l_s is not None:
                clean_row["Q_in_L_s_override"] = q_in_override_l_s
                q_in_override_l_s = None
            clean_rows.append(clean_row)
            latest_visible_q_out_l_s = brine_remaining_l_s(clean_row)
            current_solution = next_solution

        except Exception as exc:
            scenario_status = "FAIL"
            clean_rows.append(build_error_row(stage_order, stage_row, scenario_id, current_solution, exc))
            break

    clean_df_all = pd.DataFrame(clean_rows)
    phases_df_all = pd.concat(phase_tables, ignore_index=True) if phase_tables else pd.DataFrame()
    module_df_all = pd.DataFrame(module_rows)

    clean_df_all = add_scenario_metadata(clean_df_all, scenario)
    phases_df_all = add_scenario_metadata(phases_df_all, scenario)
    module_df_all = add_scenario_metadata(module_df_all, scenario)

    if scenario_status == "OK":
        print(f"Caso recovery ejecutado correctamente: {scenario_id}")
    else:
        print(f"Caso recovery con fallo: {scenario_id}")

    return clean_df_all, phases_df_all, module_df_all


def augment_metadata_with_module(metadata: dict, module_df: pd.DataFrame) -> dict:
    enriched = metadata.copy()
    module_lookup = {}
    if module_df is not None and not module_df.empty:
        for sid, group in module_df.groupby("scenario_id", sort=False):
            statuses = set(group["module_status"].fillna("").astype(str))
            errors = [
                str(msg)
                for msg in group.get("module_error_message", pd.Series(dtype=object)).fillna("").astype(str).unique()
                if str(msg).strip()
            ]
            first = group.iloc[0]
            module_lookup[sid] = {
                "module_status": "FAIL" if "FAIL" in statuses else "OK",
                "module_error_message": " | ".join(errors),
                "Q_module_L_s": first.get("Q_module_L_s"),
                "target_freshwater_L_s": first.get("target_freshwater_L_s"),
                "module_conversion_percent": first.get("module_conversion_percent"),
                "bypass_fraction_percent": first.get("bypass_fraction_percent"),
                "Q_return_L_s": first.get("Q_return_L_s"),
            }

    scenarios = []
    for row in enriched.get("scenarios", []):
        merged = row.copy()
        merged.update(module_lookup.get(row.get("scenario_id"), {
            "module_status": "NOT_RUN",
            "module_error_message": "",
            "Q_module_L_s": None,
            "target_freshwater_L_s": None,
            "module_conversion_percent": None,
            "bypass_fraction_percent": None,
            "Q_return_L_s": None,
        }))
        if merged.get("module_status") == "FAIL":
            merged["scenario_status"] = "FAIL"
        scenarios.append(merged)
    enriched["scenarios"] = scenarios
    return enriched


def build_comparison_results(
    base_summary: pd.DataFrame,
    recovery_summary: pd.DataFrame,
    base_evaluation: pd.DataFrame,
    recovery_evaluation: pd.DataFrame,
    base_scaling: pd.DataFrame,
    recovery_scaling: pd.DataFrame,
) -> tuple[pd.DataFrame, list[str]]:
    rows = []
    warnings = []

    requested = {
        "summary_results": [
            "Q_in_L_s",
            "Q_out_L_s",
            "Q_evap_L_s",
            "Q_evap_accum_L_s",
            "accumulated_evaporated_water_percent",
            "Li_mg_L",
            "Li_g_L",
            "TDS_g_L",
            "total_solids_precipitated_kg_s",
            "accumulated_solids_kg_s",
            "accumulated_halite_precipitated_percent",
            "accumulated_sulfates_precipitated_percent",
        ],
        "scaling_risk_results": [
            "scaling_risk_indicator",
            "solid_water_score",
            "phase_penalty_score",
            "saturation_score",
            "chemical_treatment_penalty",
            "solid_water_contribution",
            "phase_penalty_contribution",
            "saturation_contribution",
            "chemical_treatment_contribution",
        ],
    }

    tables = [
        ("summary_results", base_summary, recovery_summary, requested["summary_results"]),
        ("scaling_risk_results", base_scaling, recovery_scaling, requested["scaling_risk_results"]),
    ]

    if base_evaluation is not None and not base_evaluation.empty:
        ratio_cols = [
            col for col in base_evaluation.columns
            if col not in {"scenario_id", "month", "stage_id", "stage_type"} and
            pd.to_numeric(base_evaluation[col], errors="coerce").notna().any()
        ]
        tables.append(("evaluation_results", base_evaluation, recovery_evaluation, ratio_cols))

    for group_name, base_df, recovery_df, variables in tables:
        if base_df is None or recovery_df is None or base_df.empty or recovery_df.empty:
            warnings.append(f"No se compara {group_name}: tabla vacia.")
            continue

        required_keys = {"scenario_id", "stage_id"}
        if not required_keys.issubset(base_df.columns) or not required_keys.issubset(recovery_df.columns):
            warnings.append(f"No se compara {group_name}: faltan claves scenario_id/stage_id.")
            continue

        for variable in variables:
            if variable not in base_df.columns or variable not in recovery_df.columns:
                warnings.append(f"Variable omitida en {group_name}: {variable}")
                continue

            base_part = base_df[["scenario_id", "stage_id", variable]].rename(columns={variable: "base_value"})
            recovery_part = recovery_df[["scenario_id", "stage_id", variable]].rename(columns={variable: "recovery_value"})
            merged = base_part.merge(recovery_part, on=["scenario_id", "stage_id"], how="outer")

            for _, row in merged.iterrows():
                base_value = pd.to_numeric(row.get("base_value"), errors="coerce")
                recovery_value = pd.to_numeric(row.get("recovery_value"), errors="coerce")
                if pd.isna(base_value) or pd.isna(recovery_value):
                    delta_abs = None
                    delta_pct = None
                else:
                    delta_abs = float(recovery_value) - float(base_value)
                    delta_pct = None if abs(float(base_value)) <= 1e-12 else 100.0 * delta_abs / abs(float(base_value))
                rows.append({
                    "scenario_id": row["scenario_id"],
                    "stage_id": row["stage_id"],
                    "variable_group": group_name,
                    "variable": variable,
                    "base_value": None if pd.isna(base_value) else float(base_value),
                    "recovery_value": None if pd.isna(recovery_value) else float(recovery_value),
                    "delta_absolute": delta_abs,
                    "delta_percent": delta_pct,
                })

    return pd.DataFrame(rows), warnings


def stage_interval_ids(stages: pd.DataFrame, extraction_stage_id: str, reinjection_stage_id: str) -> list[str]:
    stage_ids = [str(s) for s in stages["stage_id"].tolist()]
    lookup = {sid.upper(): idx for idx, sid in enumerate(stage_ids)}
    start = lookup[extraction_stage_id.upper()]
    end = lookup[reinjection_stage_id.upper()]
    return stage_ids[start:end + 1]


def build_distributed_effect_results(module_df: pd.DataFrame, stages: pd.DataFrame, control: dict) -> pd.DataFrame:
    columns = [
        "scenario_id", "distribution_case", "stage_id", "distribution_weight",
        "distributed_recovered_water_L_s", "distributed_solids_total_kg_s",
        "distributed_phase", "distributed_mineral_family", "distributed_phase_mass_kg_s",
        "distributed_phase_mass_t_d", "distributed_accumulated_water_L_s",
        "distributed_accumulated_solids_kg_s", "is_visual_postprocess", "note",
    ]
    if module_df is None or module_df.empty:
        return pd.DataFrame(columns=columns)

    interval = skipped_pond_stage_ids(stages, control["extraction_stage_id"], control["reinjection_stage_id"])
    if not interval:
        return pd.DataFrame(columns=columns)
    weight = 1.0 / len(interval)
    rows = []
    note = "Visualizacion distribuida; no simulacion PHREEQC."

    for sid, group in module_df.groupby("scenario_id", sort=False):
        first = group.iloc[0]
        recovered = float(first.get("recovered_water_L_s", 0.0) or 0.0)
        total_solids = float(first.get("module_solids_total_kg_s", 0.0) or 0.0)
        phase_rows = group[group["module_phase"].notna()].copy()
        if phase_rows.empty:
            phase_rows = pd.DataFrame([{
                "module_phase": None,
                "module_mineral_family": None,
                "module_precipitated_mass_kg_s": 0.0,
            }])

        for idx, stage_id in enumerate(interval, start=1):
            for _, phase_row in phase_rows.iterrows():
                phase_mass = float(phase_row.get("module_precipitated_mass_kg_s", 0.0) or 0.0)
                rows.append({
                    "scenario_id": sid,
                    "distribution_case": f"{control['extraction_stage_id']}->{control['reinjection_stage_id']}",
                    "stage_id": stage_id,
                    "distribution_weight": weight,
                    "distributed_recovered_water_L_s": recovered * weight,
                    "distributed_solids_total_kg_s": total_solids * weight,
                    "distributed_phase": phase_row.get("module_phase"),
                    "distributed_mineral_family": phase_row.get("module_mineral_family"),
                    "distributed_phase_mass_kg_s": phase_mass * weight,
                    "distributed_phase_mass_t_d": kg_s_to_t_d(phase_mass * weight),
                    "distributed_accumulated_water_L_s": recovered * weight * idx,
                    "distributed_accumulated_solids_kg_s": total_solids * weight * idx,
                    "is_visual_postprocess": True,
                    "note": note,
                })

    return pd.DataFrame(rows, columns=columns)


def merge_viewer_numeric_tables(summary_df: pd.DataFrame, evaluation_df: pd.DataFrame, scaling_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df is None or summary_df.empty:
        return pd.DataFrame()

    merged = summary_df.copy()
    keys = ["scenario_id", "stage_id"]
    for extra in [evaluation_df, scaling_df]:
        if extra is None or extra.empty or not set(keys).issubset(extra.columns):
            continue
        cols = [
            col for col in extra.columns
            if col in keys or col not in merged.columns
        ]
        merged = merged.merge(extra[cols], on=keys, how="left")
    return merged


def build_module_point_viewer_df(module_df: pd.DataFrame) -> pd.DataFrame:
    if module_df is None or module_df.empty:
        return pd.DataFrame()

    records = []
    for sid, group in module_df.groupby("scenario_id", sort=False):
        first = group.iloc[0]
        q_return_design = first.get("Q_return_L_s")
        q_return_actual = first.get("Q_return_actual_L_s")
        if q_return_actual is None or pd.isna(q_return_actual):
            q_return_actual = q_return_design
        records.append({
            "scenario_id": sid,
            "month": first.get("month"),
            "stage_id": first.get("extraction_stage_id"),
            "Q_in_L_s": first.get("Q_module_L_s"),
            "Q_out_L_s": q_return_actual,
            "Q_return_design_L_s": q_return_design,
            "Q_return_actual_L_s": q_return_actual,
            "Q_evap_L_s": first.get("recovered_water_L_s"),
            "Q_evap_accum_L_s": first.get("recovered_water_L_s"),
            "Li_mg_L": first.get("Li_return_mg_L"),
            "Li_g_L": None if pd.isna(first.get("Li_return_mg_L")) else float(first.get("Li_return_mg_L")) / 1000.0,
            "TDS_g_L": first.get("TDS_return_g_L"),
            "total_solids_precipitated_kg_s": first.get("module_solids_total_kg_s"),
            "accumulated_solids_kg_s": first.get("module_solids_total_kg_s"),
        })
    return pd.DataFrame(records)


def build_distributed_viewer_df(distributed_df: pd.DataFrame) -> pd.DataFrame:
    if distributed_df is None or distributed_df.empty:
        return pd.DataFrame()

    grouped = (
        distributed_df
        .groupby(["scenario_id", "stage_id"], sort=False)
        .agg({
            "distributed_recovered_water_L_s": "first",
            "distributed_solids_total_kg_s": "first",
            "distributed_accumulated_water_L_s": "first",
            "distributed_accumulated_solids_kg_s": "first",
        })
        .reset_index()
    )
    grouped = grouped.rename(columns={
        "distributed_recovered_water_L_s": "Q_evap_L_s",
        "distributed_accumulated_water_L_s": "Q_evap_accum_L_s",
        "distributed_solids_total_kg_s": "total_solids_precipitated_kg_s",
        "distributed_accumulated_solids_kg_s": "accumulated_solids_kg_s",
    })
    return grouped


def combine_case_tables(base_df: pd.DataFrame, recovery_df: pd.DataFrame) -> pd.DataFrame:
    parts = []
    for case_order, (case_name, df) in enumerate([("base", base_df), ("recovery", recovery_df)]):
        if df is None or df.empty:
            continue
        part = df.copy()
        insert_at = 1 if "scenario_id" in part.columns else 0
        if "simulation_case" in part.columns:
            part = part.drop(columns=["simulation_case"])
        part.insert(insert_at, "simulation_case", case_name)
        part["_case_order"] = case_order
        part["_row_order"] = range(len(part))
        parts.append(part)

    if not parts:
        return pd.DataFrame()

    combined = pd.concat(parts, ignore_index=True, sort=False)
    sort_cols = []
    if "scenario_id" in combined.columns:
        sort_cols.append("scenario_id")
    sort_cols.extend(["_case_order", "_row_order"])
    combined = combined.sort_values(sort_cols, kind="stable").reset_index(drop=True)
    return combined.drop(columns=["_case_order", "_row_order"])


def build_module_export_tables(module_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    main_cols = [
        "scenario_id",
        "month",
        "month_label",
        "month_f_m",
        "bypass_velocity_factor",
        "extraction_stage_id",
        "reinjection_stage_id",
        "heating_temperature_C",
        "Q_available_at_extraction_L_s",
        "Q_module_L_s",
        "Q_main_L_s",
        "target_freshwater_L_s",
        "module_conversion_fraction",
        "module_conversion_percent",
        "base_Q_in_extraction_L_s",
        "base_Q_in_reinjection_L_s",
        "base_flow_drop_extraction_to_reinjection_L_s",
        "skipped_evap_stage_ids",
        "skipped_evap_stage_count",
        "evap_reduction_per_skipped_stage_L_s",
        "bypass_fraction_percent",
        "recovered_water_L_s",
        "Q_return_L_s",
        "Q_return_design_L_s",
        "Q_return_actual_L_s",
        "module_input_volume_L_s",
        "module_volume_check_base_L_s",
        "module_volume_calibration_factor",
        "module_input_water_kg_s",
        "module_solids_total_kg_s",
        "retained_brine_kg_s",
        "TSS_module_g_L",
        "TDS_return_g_L",
        "Li_return_mg_L",
        "module_status",
        "module_error_message",
    ]
    phase_cols = [
        "scenario_id",
        "month",
        "month_label",
        "month_f_m",
        "bypass_velocity_factor",
        "extraction_stage_id",
        "reinjection_stage_id",
        "module_phase",
        "module_mineral_family",
        "module_precipitated_mass_kg_s",
        "module_precipitated_mass_t_d",
        "module_accumulated_precipitated_mass_kg_s",
        "saturation_index_phase",
        "precipitated_phases",
        "saturation_indexes_by_phase",
        "module_status",
        "module_error_message",
    ]

    if module_df is None or module_df.empty:
        return pd.DataFrame(columns=main_cols), pd.DataFrame(columns=phase_cols)

    cleaned = module_df.copy()
    if "module_solids_removed_kg_s" in cleaned.columns:
        cleaned = cleaned.drop(columns=["module_solids_removed_kg_s"])

    available_main_cols = [col for col in main_cols if col in cleaned.columns]
    available_phase_cols = [col for col in phase_cols if col in cleaned.columns]
    module_main = cleaned[available_main_cols].drop_duplicates(subset=["scenario_id"], keep="first")
    module_phases = cleaned[available_phase_cols].copy()
    return module_main.reset_index(drop=True), module_phases.reset_index(drop=True)


def write_recovery_compare_viewer(
    base_summary: pd.DataFrame,
    base_evaluation: pd.DataFrame,
    base_geochemical: pd.DataFrame,
    base_precipitation: pd.DataFrame,
    base_scaling: pd.DataFrame,
    recovery_summary: pd.DataFrame,
    recovery_evaluation: pd.DataFrame,
    recovery_geochemical: pd.DataFrame,
    recovery_precipitation: pd.DataFrame,
    recovery_scaling: pd.DataFrame,
    module_df: pd.DataFrame,
    distributed_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    metadata: dict,
    batch_results_dir: Path,
) -> None:
    plotly_src = copy_plotly_asset(batch_results_dir)
    plotly_tag = (
        f'<script src="{plotly_src}"></script>'
        if plotly_src
        else '<script>window.PLOTLY_MISSING = true;</script>'
    )

    base_curve = merge_viewer_numeric_tables(base_summary, base_evaluation, base_scaling)
    recovery_curve = merge_viewer_numeric_tables(recovery_summary, recovery_evaluation, recovery_scaling)
    module_curve = build_module_point_viewer_df(module_df)
    distributed_curve = build_distributed_viewer_df(distributed_df)
    module_main_df, module_phases_df = build_module_export_tables(module_df)
    combined_summary = combine_case_tables(base_summary, recovery_summary)
    combined_evaluation = combine_case_tables(base_evaluation, recovery_evaluation)
    combined_geochemical = combine_case_tables(base_geochemical, recovery_geochemical)
    combined_precipitation = combine_case_tables(base_precipitation, recovery_precipitation)
    combined_scaling = combine_case_tables(base_scaling, recovery_scaling)

    curve_tables = {
        "base": df_to_json_records(base_curve),
        "recovery": df_to_json_records(recovery_curve),
        "module_point": df_to_json_records(module_curve),
        "distributed": df_to_json_records(distributed_curve),
    }
    all_curve_cols = set()
    for df in [base_curve, recovery_curve, module_curve, distributed_curve]:
        if df is not None and not df.empty:
            all_curve_cols.update(df.columns)
    variables = sorted(
        col for col in all_curve_cols
        if col not in {"scenario_id", "month", "stage_id", "stage_type", "group_type"} and
        any(
            df is not None and not df.empty and col in df.columns and
            pd.to_numeric(df[col], errors="coerce").notna().any()
            for df in [base_curve, recovery_curve, module_curve, distributed_curve]
        )
    )

    payload = {
        "curve_tables": curve_tables,
        "plot_tables": {
            "base": {
                "summary_results": df_to_json_records(base_summary),
                "evaluation_results": df_to_json_records(base_evaluation),
                "geochemical_results": df_to_json_records(base_geochemical),
                "precipitation_by_phase": df_to_json_records(base_precipitation),
                "scaling_risk_results": df_to_json_records(base_scaling),
            },
            "recovery": {
                "summary_results": df_to_json_records(recovery_summary),
                "evaluation_results": df_to_json_records(recovery_evaluation),
                "geochemical_results": df_to_json_records(recovery_geochemical),
                "precipitation_by_phase": df_to_json_records(recovery_precipitation),
                "scaling_risk_results": df_to_json_records(recovery_scaling),
            },
            "module_point": {
                "summary_results": df_to_json_records(module_curve),
                "recovery_module_results": df_to_json_records(module_df),
                "module_phases": df_to_json_records(module_phases_df),
            },
            "distributed": {
                "summary_results": df_to_json_records(distributed_curve),
                "distributed_effect_results": df_to_json_records(distributed_df),
            },
        },
        "tables": {
            "summary_results": df_to_json_records(combined_summary),
            "evaluation_results": df_to_json_records(combined_evaluation),
            "geochemical_results": df_to_json_records(combined_geochemical),
            "precipitation_by_phase": df_to_json_records(combined_precipitation),
            "scaling_risk_results": df_to_json_records(combined_scaling),
            "recovery_module_results": df_to_json_records(module_main_df),
            "module_phases": df_to_json_records(module_phases_df),
            "comparison_results": df_to_json_records(comparison_df),
            "distributed_effect_results": df_to_json_records(distributed_df),
        },
        "variables": variables,
        "metadata": metadata,
        "recovery_methodology": RECOVERY_METHOD_TEXT,
        "scaling_risk_methodology": SCALING_RISK_METHOD_TEXT,
    }
    payload_json = json.dumps(make_json_safe(payload), ensure_ascii=False, allow_nan=False)

    html_template = r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Visor de resultados - Comparacion recovery</title>
  __PLOTLY_TAG__
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #17202a;
      --muted: #667085;
      --line: #d8dee8;
      --accent: #0969da;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      background: var(--bg);
      color: var(--ink);
    }
    header {
      padding: 16px 22px;
      border-bottom: 1px solid var(--line);
      background: var(--panel);
      position: sticky;
      top: 0;
      z-index: 10;
    }
    h1 { margin: 0 0 4px; font-size: 20px; }
    .sub { color: var(--muted); font-size: 13px; }
    main {
      display: grid;
      grid-template-columns: 320px 1fr;
      min-height: calc(100vh - 72px);
    }
    aside {
      padding: 16px;
      border-right: 1px solid var(--line);
      background: var(--panel);
      overflow: auto;
    }
    section { padding: 16px; overflow: auto; }
    label {
      display: block;
      margin: 12px 0 5px;
      font-size: 12px;
      font-weight: 700;
      color: #344054;
    }
    select, button, input {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px;
      background: #fff;
      color: var(--ink);
      font-size: 13px;
    }
    .checkbox-list {
      width: 100%;
      max-height: 180px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 6px;
      background: #fff;
    }
    .checkbox-list.compact { max-height: 130px; }
    .check-option {
      display: flex;
      align-items: center;
      gap: 8px;
      margin: 0;
      padding: 4px 2px;
      font-size: 12px;
      font-weight: 400;
      line-height: 1.2;
      color: var(--ink);
    }
    .check-option input {
      width: auto;
      margin: 0;
      flex: 0 0 auto;
    }
    .filter-actions {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 8px;
      margin-top: 6px;
    }
    .filter-actions button {
      margin-top: 0;
      padding: 6px 8px;
      font-size: 12px;
    }
    button {
      cursor: pointer;
      font-weight: 700;
      margin-top: 10px;
    }
    button.primary {
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }
    button.danger {
      color: #b42318;
      background: #fff;
    }
    .hint {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
      margin-top: 10px;
    }
    .stats {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
      gap: 10px;
      margin-bottom: 14px;
    }
    .stat {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
    }
    .stat .k { color: var(--muted); font-size: 12px; }
    .stat .v { font-size: 18px; font-weight: 700; margin-top: 4px; }
    .chart-card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-bottom: 14px;
      overflow: hidden;
    }
    .chart-head {
      display: grid;
      grid-template-columns: repeat(6, minmax(130px, 1fr)) 86px;
      gap: 8px;
      padding: 12px;
      border-bottom: 1px solid var(--line);
      align-items: end;
    }
    .chart-head label { margin-top: 0; }
    .plot { min-height: 430px; }
    .table-wrap {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: auto;
      max-height: 440px;
      margin-top: 14px;
    }
    .methodology {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-top: 14px;
      padding: 12px;
    }
    .methodology summary {
      cursor: pointer;
      font-weight: 700;
    }
    .methodology pre {
      white-space: pre-wrap;
      margin: 12px 0 0;
      color: #344054;
      font-family: Arial, Helvetica, sans-serif;
      font-size: 12px;
      line-height: 1.45;
    }
    table { border-collapse: collapse; width: 100%; font-size: 12px; }
    th, td { border-bottom: 1px solid #edf0f5; padding: 6px 8px; text-align: right; white-space: nowrap; }
    th { position: sticky; top: 0; background: #f9fafb; z-index: 1; color: #344054; }
    th:first-child, td:first-child { text-align: left; }
    .missing {
      padding: 14px;
      margin-bottom: 12px;
      border: 1px solid #fecdca;
      color: #b42318;
      background: #fffbfa;
      border-radius: 8px;
      display: none;
    }
  </style>
</head>
<body>
  <header>
    <h1>Visor de resultados del simulador</h1>
    <div class="sub" id="runInfo"></div>
  </header>
  <main>
    <aside>
      <button class="primary" id="addChartBtn">Anadir grafico</button>
      <label>Tabla para vista rapida</label>
      <select id="previewTable"></select>
      <label>Filtrar meses</label>
      <div id="globalMonths" class="checkbox-list compact"></div>
      <div class="filter-actions">
        <button type="button" data-check-target="globalMonths" data-check-value="true">Todos</button>
        <button type="button" data-check-target="globalMonths" data-check-value="false">Ninguno</button>
      </div>
      <label>Casos en graficos</label>
      <div id="globalScenarios" class="checkbox-list"></div>
      <div class="filter-actions">
        <button type="button" data-check-target="globalScenarios" data-check-value="true">Todos</button>
        <button type="button" data-check-target="globalScenarios" data-check-value="false">Ninguno</button>
      </div>
      <button id="applyPreviewBtn">Actualizar tabla</button>
      <p class="hint">
        En los graficos, el eje X son las etapas del tren y el eje Y es el parametro elegido.
        El desplegable Caso permite elegir base, recovery fisico, modulo en extraccion,
        distribuido visual o todos superpuestos.
      </p>
    </aside>
    <section>
      <div class="missing" id="plotlyMissing">
        No se encontro Plotly local. El visor se genero, pero los graficos no podran dibujarse hasta que exista viewer_assets/plotly-2.35.2.min.js.
      </div>
      <div class="stats" id="stats"></div>
      <div id="charts"></div>
      <details class="methodology">
        <summary>Metodologia modulo recovery</summary>
        <pre id="recoveryMethodology"></pre>
      </details>
      <details class="methodology">
        <summary>Metodologia scaling_risk_indicator</summary>
        <pre id="scalingRiskMethodology"></pre>
      </details>
      <div class="table-wrap"><table id="preview"></table></div>
    </section>
  </main>
  <script id="payload" type="application/json">__PAYLOAD_JSON__</script>
  <script>
    const STAGE_ORDER = ["PC1","PC2","PC3","PC4","PC5","PC6","PC7","PC8","LIM1","H1","H2","H3","K1","C1","L1"];
    const stageRank = Object.fromEntries(STAGE_ORDER.map((s, i) => [s, i]));
    const payload = JSON.parse(document.getElementById("payload").textContent);
    const tables = payload.tables || {};
    const plotTables = payload.plot_tables || {};
    const scenarioSource = payload.metadata?.base?.scenarios || payload.metadata?.recovery?.scenarios || [];
    const scenarioMap = Object.fromEntries(scenarioSource.map(s => [s.scenario_id, s]));
    const CASE_LABELS = {
      all: "Todos superpuestos",
      base: "Caso base",
      recovery: "Recovery fisico",
      module_point: "Modulo en extraccion",
      distributed: "Distribuido visual"
    };
    const CASE_OPTIONS = ["all", "base", "recovery", "module_point", "distributed"];
    const hiddenPreviewColumns = new Set(["stage_type"]);
    let chartCounter = 0;

    function num(v) {
      if (v === null || v === undefined || v === "") return null;
      const n = Number(String(v).replace(",", "."));
      return Number.isFinite(n) ? n : null;
    }
    function fmt(v) {
      const n = num(v);
      if (n === null) return v ?? "";
      return Math.abs(n) >= 1000 ? n.toFixed(2) : n.toPrecision(5);
    }
    function stageOf(row) {
      return row.stage_id || row.extraction_stage_id || row.reinjection_stage_id || "";
    }
    function annualGroup(sid) {
      const s = scenarioMap[sid] || {};
      return `ANNUAL_MEAN | T=${s.temperature ?? "-"} | E=${s.evap_factor ?? "-"} | Mg=${s.mg_removal ?? "-"} | R=${s.retention_r ?? "-"}`;
    }
    function setOptions(select, values, selectedAll = false) {
      const previous = select.value;
      select.innerHTML = "";
      values.forEach(v => {
        const opt = document.createElement("option");
        opt.value = v;
        opt.textContent = CASE_LABELS[v] || v;
        opt.selected = selectedAll;
        select.appendChild(opt);
      });
      if (values.includes(previous)) select.value = previous;
    }
    function checkedValues(containerId) {
      return [...document.querySelectorAll(`#${containerId} input[type="checkbox"]:checked`)].map(i => i.value);
    }
    function renderCheckboxGroup(containerId, values) {
      const container = document.getElementById(containerId);
      container.innerHTML = "";
      values.forEach(v => {
        const label = document.createElement("label");
        label.className = "check-option";
        const input = document.createElement("input");
        input.type = "checkbox";
        input.value = v;
        input.checked = true;
        input.addEventListener("change", refreshOutputs);
        const text = document.createElement("span");
        text.textContent = v;
        label.appendChild(input);
        label.appendChild(text);
        container.appendChild(label);
      });
    }
    function setAllCheckboxes(containerId, checked) {
      document.querySelectorAll(`#${containerId} input[type="checkbox"]`).forEach(input => {
        input.checked = checked;
      });
      refreshOutputs();
    }
    function caseList(caseValue) {
      return caseValue === "all" ? ["base", "recovery", "module_point", "distributed"] : [caseValue];
    }
    function tableNamesForCase(caseValue) {
      const names = new Set();
      caseList(caseValue).forEach(caseName => {
        Object.keys(plotTables[caseName] || {}).forEach(name => names.add(name));
      });
      return [...names];
    }
    function rowsForCaseTable(caseName, tableName) {
      return (plotTables[caseName] || {})[tableName] || [];
    }
    function filteredPlotRows(caseName, tableName) {
      const months = checkedValues("globalMonths");
      const sids = checkedValues("globalScenarios");
      return rowsForCaseTable(caseName, tableName).filter(r => {
        const hasMonth = r.month !== null && r.month !== undefined && r.month !== "";
        const hasScenario = r.scenario_id !== null && r.scenario_id !== undefined && r.scenario_id !== "";
        const mOk = !hasMonth || months.includes(String(r.month));
        const sOk = !hasScenario || sids.includes(String(r.scenario_id));
        return mOk && sOk;
      });
    }
    function filteredRows(tableName) {
      const rows = tables[tableName] || [];
      const months = checkedValues("globalMonths");
      const sids = checkedValues("globalScenarios");
      return rows.filter(r => {
        const hasMonth = r.month !== null && r.month !== undefined && r.month !== "";
        const hasScenario = r.scenario_id !== null && r.scenario_id !== undefined && r.scenario_id !== "";
        const mOk = !hasMonth || months.includes(String(r.month));
        const sOk = !hasScenario || sids.includes(String(r.scenario_id));
        return mOk && sOk;
      });
    }
    function numericColumnsFor(card) {
      const caseValue = card.querySelector(".case-select").value;
      const tableName = card.querySelector(".table-select").value;
      const cols = new Set();
      caseList(caseValue).forEach(caseName => {
        filteredPlotRows(caseName, tableName).forEach(row => {
          Object.keys(row).forEach(key => {
            if (!["scenario_id", "month", "stage_id", "stage_type", "group_type"].includes(key) && num(row[key]) !== null) {
              cols.add(key);
            }
          });
        });
      });
      return [...cols].sort();
    }
    function sortByStage(rows) {
      return [...rows].sort((a, b) => {
        const ar = stageRank[stageOf(a)] ?? 999;
        const br = stageRank[stageOf(b)] ?? 999;
        if (ar !== br) return ar - br;
        return String(stageOf(a)).localeCompare(String(stageOf(b)));
      });
    }
    function seriesValue(row, mode) {
      if (mode === "annual_mean") return annualGroup(row.scenario_id);
      if (mode === "mineral_phase") {
        return row.mineral_phase || row.module_phase || row.distributed_phase || "fase";
      }
      return String(row[mode] ?? "serie");
    }
    function aggregateValues(values, tableName, yCol, mode) {
      if (!values.length) return null;
      const useSum = tableName === "precipitation_by_phase" && mode !== "mineral_phase" && yCol.includes("mass");
      const total = values.reduce((a, c) => a + c, 0);
      return useSum ? total : total / values.length;
    }
    function groupRows(rows, yCol, mode, caseName, prefixCase, tableName) {
      const bySeries = new Map();
      if (mode === "annual_mean") {
        const buckets = new Map();
        rows.forEach(r => {
          const stage = stageOf(r);
          const y = num(r[yCol]);
          if (!stage || y === null) return;
          const series = seriesValue(r, mode);
          const key = `${series}||${stage}`;
          if (!buckets.has(key)) buckets.set(key, { series, stage_id: stage, values: [] });
          buckets.get(key).values.push(y);
        });
        buckets.forEach(b => {
          const series = prefixCase ? `${CASE_LABELS[caseName]} | ${b.series}` : b.series;
          if (!bySeries.has(series)) bySeries.set(series, []);
          bySeries.get(series).push({
            stage_id: b.stage_id,
            value: b.values.reduce((a, c) => a + c, 0) / b.values.length
          });
        });
        return bySeries;
      }

      const buckets = new Map();
      rows.forEach(r => {
        const stage = stageOf(r);
        const y = num(r[yCol]);
        if (!stage || y === null) return;
        const baseSeries = seriesValue(r, mode);
        const series = prefixCase ? `${CASE_LABELS[caseName]} | ${baseSeries}` : baseSeries;
        const key = `${series}||${stage}`;
        if (!buckets.has(key)) buckets.set(key, { series, stage_id: stage, values: [] });
        buckets.get(key).values.push(y);
      });
      buckets.forEach(b => {
        const value = aggregateValues(b.values, tableName, yCol, mode);
        if (value === null) return;
        if (!bySeries.has(b.series)) bySeries.set(b.series, []);
        bySeries.get(b.series).push({ stage_id: b.stage_id, value });
      });
      return bySeries;
    }
    function renderChart(card) {
      if (window.PLOTLY_MISSING || typeof Plotly === "undefined") return;
      const caseValue = card.querySelector(".case-select").value;
      const tableName = card.querySelector(".table-select").value;
      const yCol = card.querySelector(".y-select").value;
      const mode = card.querySelector(".series-select").value;
      const traces = [];
      caseList(caseValue).forEach(caseName => {
        const rows = filteredPlotRows(caseName, tableName);
        const grouped = groupRows(rows, yCol, mode, caseName, caseValue === "all", tableName);
        grouped.forEach((items, series) => {
          const sorted = sortByStage(items);
          traces.push({
            x: sorted.map(i => i.stage_id),
            y: sorted.map(i => i.value),
            type: "scatter",
            mode: "lines+markers",
            name: series,
          });
        });
      });
      const layout = {
        margin: { l: 70, r: 20, t: 40, b: 70 },
        title: `${yCol} por etapa`,
        xaxis: { title: "Etapa", categoryorder: "array", categoryarray: STAGE_ORDER },
        yaxis: { title: yCol },
        legend: { orientation: "h", y: -0.25 },
      };
      Plotly.newPlot(card.querySelector(".plot"), traces, layout, { responsive: true, displaylogo: false });
    }
    function updateTableOptions(card) {
      const caseValue = card.querySelector(".case-select").value;
      const tableSelect = card.querySelector(".table-select");
      const tablesForCase = tableNamesForCase(caseValue);
      setOptions(tableSelect, tablesForCase);
      if (!tableSelect.value && tablesForCase.length) tableSelect.value = tablesForCase[0];
      if (tablesForCase.includes("summary_results")) tableSelect.value = tableSelect.value || "summary_results";
      updateYOptions(card);
    }
    function updateYOptions(card) {
      const cols = numericColumnsFor(card);
      const y = card.querySelector(".y-select");
      const preferred = cols.includes("Li_mg_L") ? "Li_mg_L" : cols.includes("TDS_g_L") ? "TDS_g_L" : cols[0];
      setOptions(y, cols);
      y.value = preferred || "";
    }
    function addChart() {
      chartCounter += 1;
      const card = document.createElement("div");
      card.className = "chart-card";
      card.innerHTML = `
        <div class="chart-head">
          <div><label>Tabla</label><select class="table-select"></select></div>
          <div><label>Parametro Y</label><select class="y-select"></select></div>
          <div><label>Caso</label><select class="case-select"></select></div>
          <div><label>Series</label><select class="series-select"><option value="scenario_id">Escenario</option><option value="month">Mes</option><option value="annual_mean">Media anual</option><option value="mineral_phase">Fase mineral</option></select></div>
          <div><label>Actualizar</label><button class="update-btn">Redibujar</button></div>
          <div><label>Duplicar</label><button class="duplicate-btn">Duplicar</button></div>
          <div><button class="danger delete-btn">Quitar</button></div>
        </div>
        <div class="plot" id="plot-${chartCounter}"></div>`;
      document.getElementById("charts").appendChild(card);
      setOptions(card.querySelector(".case-select"), CASE_OPTIONS);
      card.querySelector(".case-select").value = "all";
      updateTableOptions(card);
      if (tableNamesForCase("all").includes("summary_results")) card.querySelector(".table-select").value = "summary_results";
      updateYOptions(card);
      card.querySelector(".case-select").addEventListener("change", () => { updateTableOptions(card); renderChart(card); });
      card.querySelector(".table-select").addEventListener("change", () => { updateYOptions(card); renderChart(card); });
      card.querySelector(".update-btn").addEventListener("click", () => renderChart(card));
      card.querySelector(".duplicate-btn").addEventListener("click", addChart);
      card.querySelector(".delete-btn").addEventListener("click", () => card.remove());
      card.querySelectorAll("select").forEach(s => s.addEventListener("change", () => renderChart(card)));
      renderChart(card);
    }
    function renderPreview() {
      const tableName = document.getElementById("previewTable").value;
      const rows = filteredRows(tableName).slice(0, 300);
      const table = document.getElementById("preview");
      if (!rows.length) { table.innerHTML = ""; return; }
      const cols = Object.keys(rows[0]).filter(c => !hiddenPreviewColumns.has(c));
      table.innerHTML = `<thead><tr>${cols.map(c => `<th>${c}</th>`).join("")}</tr></thead>` +
        `<tbody>${rows.map(r => `<tr>${cols.map(c => `<td>${num(r[c]) === null ? (r[c] ?? "") : fmt(r[c])}</td>`).join("")}</tr>`).join("")}</tbody>`;
    }
    function renderStats() {
      const meta = payload.metadata || {};
      const baseRows = (plotTables.base?.summary_results || []).filter(r => r.stage_id === "L1").map(r => num(r.Li_mg_L)).filter(v => v !== null);
      const recoveryRows = (plotTables.recovery?.summary_results || []).filter(r => r.stage_id === "L1").map(r => num(r.Li_mg_L)).filter(v => v !== null);
      const avgBaseLi = baseRows.length ? baseRows.reduce((a, c) => a + c, 0) / baseRows.length : null;
      const avgRecoveryLi = recoveryRows.length ? recoveryRows.reduce((a, c) => a + c, 0) / recoveryRows.length : null;
      const moduleRows = tables.recovery_module_results || [];
      const recovered = moduleRows.length ? num(moduleRows[0].target_freshwater_L_s) : null;
      const cards = [
        ["Run", meta.run_stamp || ""],
        ["Escenarios", scenarioSource.length],
        ["Li L1 base", avgBaseLi === null ? "" : `${avgBaseLi.toFixed(1)} mg/L`],
        ["Li L1 recovery", avgRecoveryLi === null ? "" : `${avgRecoveryLi.toFixed(1)} mg/L`],
        ["Agua recuperada", recovered === null ? "" : `${recovered.toFixed(2)} L/s`],
      ];
      document.getElementById("stats").innerHTML = cards.map(([k, v]) => `<div class="stat"><div class="k">${k}</div><div class="v">${v}</div></div>`).join("");
    }
    function refreshOutputs() {
      renderPreview();
      document.querySelectorAll(".chart-card").forEach(renderChart);
    }
    function init() {
      if (window.PLOTLY_MISSING || typeof Plotly === "undefined") {
        document.getElementById("plotlyMissing").style.display = "block";
      }
      document.getElementById("runInfo").textContent = `${payload.metadata?.run_stamp || ""} · PHREEQC: ${payload.metadata?.phreeqc_database || ""}`;
      setOptions(document.getElementById("previewTable"), Object.keys(tables));
      if (Object.keys(tables).includes("summary_results")) document.getElementById("previewTable").value = "summary_results";
      const allMonths = [...new Set(scenarioSource.map(s => String(s.month)))].sort((a, b) => Number(a) - Number(b));
      const allScenarios = scenarioSource.map(s => s.scenario_id);
      renderCheckboxGroup("globalMonths", allMonths);
      renderCheckboxGroup("globalScenarios", allScenarios);
      document.getElementById("recoveryMethodology").textContent = payload.recovery_methodology || "";
      document.getElementById("scalingRiskMethodology").textContent = payload.scaling_risk_methodology || "";
      document.getElementById("addChartBtn").addEventListener("click", addChart);
      document.getElementById("applyPreviewBtn").addEventListener("click", refreshOutputs);
      document.querySelectorAll("[data-check-target]").forEach(button => {
        button.addEventListener("click", () => {
          setAllCheckboxes(button.dataset.checkTarget, button.dataset.checkValue === "true");
        });
      });
      renderStats();
      addChart();
      renderPreview();
    }
    init();
  </script>
</body>
</html>
"""

    html = (
        html_template
        .replace("__PLOTLY_TAG__", plotly_tag)
        .replace("__PAYLOAD_JSON__", payload_json.replace("</", "<\\/"))
    )
    (batch_results_dir / "results_viewer.html").write_text(html, encoding="utf-8")


def export_recovery_compare_outputs(
    base_outputs: tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict],
    recovery_outputs: tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict],
    base_validation_df: pd.DataFrame,
    recovery_validation_df: pd.DataFrame,
    module_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    distributed_df: pd.DataFrame,
    comparison_warnings: list[str],
    recovery_control: dict,
    batch_results_dir: Path,
) -> None:
    base_summary, base_eval, base_geo, base_precip, base_scaling, base_metadata = base_outputs
    rec_summary, rec_eval, rec_geo, rec_precip, rec_scaling, rec_metadata = recovery_outputs

    base_scenario_metadata_df = pd.DataFrame(base_metadata.get("scenarios", []))
    rec_scenario_metadata_df = pd.DataFrame(rec_metadata.get("scenarios", []))
    summary_df = combine_case_tables(base_summary, rec_summary)
    evaluation_df = combine_case_tables(base_eval, rec_eval)
    geochemical_df = combine_case_tables(base_geo, rec_geo)
    precipitation_df = combine_case_tables(base_precip, rec_precip)
    scaling_risk_df = combine_case_tables(base_scaling, rec_scaling)
    scenario_metadata_df = combine_case_tables(base_scenario_metadata_df, rec_scenario_metadata_df)
    validation_df = combine_case_tables(base_validation_df, recovery_validation_df)
    module_main_df, module_phases_df = build_module_export_tables(module_df)
    comparison_warnings_df = pd.DataFrame({"warning": comparison_warnings})
    recovery_control_df = pd.DataFrame([
        {"variable": key, "data": value}
        for key, value in recovery_control.items()
    ])
    run_info_df = pd.DataFrame([{
        "run_stamp": rec_metadata.get("run_stamp"),
        "generated_at": rec_metadata.get("generated_at"),
        "root": rec_metadata.get("root"),
        "input_dir": rec_metadata.get("input_dir"),
        "phreeqc_exe": rec_metadata.get("phreeqc_exe"),
        "phreeqc_database": rec_metadata.get("phreeqc_database"),
    }])
    tables = {
        "summary_results": summary_df,
        "evaluation_results": evaluation_df,
        "precipitation_by_phase": precipitation_df,
        "geochemical_results": geochemical_df,
        "scaling_risk_results": scaling_risk_df,
        "scenario_metadata": scenario_metadata_df,
        "recovery_module_results": module_main_df,
        "module_phases": module_phases_df,
        "comparison_results": comparison_df,
        "distributed_effect_results": distributed_df,
        "comparison_warnings": comparison_warnings_df,
        "recovery_control": recovery_control_df,
        "validation": validation_df,
        "run_info": run_info_df,
    }

    for name, df in tables.items():
        df.to_csv(batch_results_dir / f"{name}.csv", index=False)

    workbook_path = batch_results_dir / "simulator_results.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(summary_df), "summary_results", blue_recovery_rows=True)
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(evaluation_df), "evaluation_results", blue_recovery_rows=True)
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(precipitation_df), "precipitation_by_phase", blue_recovery_rows=True)
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(geochemical_df), "geochemical_results", blue_recovery_rows=True)
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(scaling_risk_df), "scaling_risk_results", blue_recovery_rows=True)
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(scenario_metadata_df), "scenario_metadata", blue_recovery_rows=True)
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(module_main_df), "recovery_module_results", tab_color="5B9BD5")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(module_phases_df), "module_phases", tab_color="5B9BD5")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(comparison_df), "comparison_results")
        write_sheet_with_units(writer, add_blank_rows_between_scenarios(distributed_df), "distributed_effect_results", tab_color="5B9BD5")
        write_sheet_with_units(writer, comparison_warnings_df, "comparison_warnings")
        write_sheet_with_units(writer, recovery_control_df, "recovery_control")
        write_sheet_with_units(writer, validation_df, "validation", blue_recovery_rows=True)
        write_sheet_with_units(writer, run_info_df, "run_info")

    combined_metadata = {
        "run_stamp": rec_metadata.get("run_stamp"),
        "generated_at": rec_metadata.get("generated_at"),
        "root": str(ROOT),
        "input_dir": str(INPUTS),
        "phreeqc_exe": str(PHREEQC_EXE),
        "phreeqc_database": str(DATABASE),
        "recovery_control": recovery_control,
        "base": base_metadata,
        "recovery": rec_metadata,
        "comparison_warnings": comparison_warnings,
    }
    (batch_results_dir / "scenario_metadata.json").write_text(
        json.dumps(make_json_safe(combined_metadata), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    (batch_results_dir / "recovery_methodology.txt").write_text(
        RECOVERY_METHOD_TEXT,
        encoding="utf-8",
    )
    (batch_results_dir / "scaling_risk_methodology.txt").write_text(
        SCALING_RISK_METHOD_TEXT,
        encoding="utf-8",
    )

    write_recovery_compare_viewer(
        base_summary=base_summary,
        base_evaluation=base_eval,
        base_geochemical=base_geo,
        base_precipitation=base_precip,
        base_scaling=base_scaling,
        recovery_summary=rec_summary,
        recovery_evaluation=rec_eval,
        recovery_geochemical=rec_geo,
        recovery_precipitation=rec_precip,
        recovery_scaling=rec_scaling,
        module_df=module_df,
        distributed_df=distributed_df,
        comparison_df=comparison_df,
        metadata=combined_metadata,
        batch_results_dir=batch_results_dir,
    )


# =========================================================
# EJECUCION DE CASOS
# =========================================================

def build_initial_solution(feed: dict) -> dict:
    return {
        "temp_C": float(feed["temp_C"]),
        "pH": float(feed["pH"]),
        "water_kg": float(feed["water_kg"]),
        "Li": float(feed["Li"]),
        "K": float(feed["K"]),
        "Mg": float(feed["Mg"]),
        "B": float(feed["B"]),
        "Ca": float(feed["Ca"]),
        "Na": float(feed["Na"]),
        "Cl": float(feed["Cl"]),
        "HCO3": float(feed["HCO3"]),
        "S6": float(feed["S6"]),
        "original_water_kg": feed.get("original_water_kg"),
        "design_total_evap_L_s": feed.get("design_total_evap_L_s"),
    }


def add_scenario_metadata(df: pd.DataFrame, scenario: pd.Series) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    for col in [
        "month", "month_label", "month_f_m", "bypass_velocity_factor",
        "temperature", "evap_factor", "mg_removal", "retention_r",
    ]:
        df[col] = scenario[col]
    return df


def run_case(
    scenario: pd.Series,
    feed: dict,
    stages: pd.DataFrame,
    batch_runs_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    scenario_id = str(scenario["scenario_id"])
    scenario_run_dir = batch_runs_dir / scenario_id
    scenario_run_dir.mkdir(parents=True, exist_ok=True)

    current_solution = build_initial_solution(feed)
    feed_check_folder = scenario_run_dir / "00_FEED_volume_check"
    current_solution["feed_phreeqc_volume_L_s"] = run_solution_volume_check(
        current_solution,
        feed_check_folder,
        "feed",
    )
    phase_tables = []
    clean_rows = [build_feed_row(current_solution, scenario_id)]
    scenario_status = "OK"
    cumulative_solids_kg = 0.0

    for stage_order, (_, stage_row) in enumerate(stages.iterrows(), start=1):
        try:
            phase_df, clean_row, next_solution = run_stage(
                stage_order=stage_order,
                stage_row=stage_row,
                solution_in=current_solution,
                scenario_id=scenario_id,
                scenario_run_dir=scenario_run_dir,
            )

            if not phase_df.empty:
                phase_tables.append(phase_df)

            cumulative_solids_kg += float(clean_row.get("total_precipitated_mass_kg", 0.0) or 0.0)
            clean_row["cumulative_solids_kg"] = cumulative_solids_kg
            clean_rows.append(clean_row)
            current_solution = next_solution

        except Exception as exc:
            scenario_status = "FAIL"
            clean_rows.append(build_error_row(stage_order, stage_row, scenario_id, current_solution, exc))
            break

    clean_df_all = pd.DataFrame(clean_rows)
    phases_df_all = pd.concat(phase_tables, ignore_index=True) if phase_tables else pd.DataFrame()

    clean_df_all = add_scenario_metadata(clean_df_all, scenario)
    phases_df_all = add_scenario_metadata(phases_df_all, scenario)

    if scenario_status == "OK":
        print(f"Caso ejecutado correctamente: {scenario_id}")
    else:
        print(f"Caso con fallo: {scenario_id}")

    return clean_df_all, phases_df_all


def run_all_scenarios():
    validate_paths()

    feed = load_feed()
    stages = load_stages()
    months_control = load_months_control()
    scenarios_control = load_scenarios_control()
    recovery_control = validate_recovery_control(read_recovery_control(), stages)

    validate_feed(feed)
    validate_stages(stages)
    validate_months_control(months_control)
    validate_scenarios_control(scenarios_control)

    scenarios_generated = expand_scenarios_control(scenarios_control, months_control)

    RUNS.mkdir(parents=True, exist_ok=True)
    RESULTS.mkdir(parents=True, exist_ok=True)

    run_stamp = build_run_stamp()
    run_stamp_ret = f"{run_stamp}_ret_recovery_compare"

    batch_runs_dir = RUNS / run_stamp_ret
    batch_results_dir = RESULTS / run_stamp_ret

    batch_runs_dir.mkdir(parents=True, exist_ok=True)
    batch_results_dir.mkdir(parents=True, exist_ok=True)

    all_base_clean = []
    all_base_phases = []
    all_recovery_clean = []
    all_recovery_phases = []
    all_module_rows = []

    for _, scenario in scenarios_generated.iterrows():
        feed_mod, stages_mod = apply_scenario(feed, stages, months_control, scenario)

        base_clean_df, base_phases_df = run_case(
            scenario=scenario,
            feed=feed_mod,
            stages=stages_mod,
            batch_runs_dir=batch_runs_dir / "base",
        )
        base_summary_for_control = build_summary_results(base_clean_df)
        scenario_recovery_control = derive_recovery_control_from_base(
            control=recovery_control,
            base_summary=base_summary_for_control,
            stages=stages_mod,
        )

        recovery_clean_df, recovery_phases_df, module_df = run_recovery_case(
            scenario=scenario,
            feed=feed_mod,
            stages=stages_mod,
            control=scenario_recovery_control,
            batch_runs_dir=batch_runs_dir,
        )

        if not base_clean_df.empty:
            all_base_clean.append(base_clean_df)
        if not base_phases_df.empty:
            all_base_phases.append(base_phases_df)
        if not recovery_clean_df.empty:
            all_recovery_clean.append(recovery_clean_df)
        if not recovery_phases_df.empty:
            all_recovery_phases.append(recovery_phases_df)
        if not module_df.empty:
            all_module_rows.append(module_df)

    base_clean_all = pd.concat(all_base_clean, ignore_index=True) if all_base_clean else pd.DataFrame()
    base_phases_all = pd.concat(all_base_phases, ignore_index=True) if all_base_phases else pd.DataFrame()
    recovery_clean_all = pd.concat(all_recovery_clean, ignore_index=True) if all_recovery_clean else pd.DataFrame()
    recovery_phases_all = pd.concat(all_recovery_phases, ignore_index=True) if all_recovery_phases else pd.DataFrame()
    module_all = pd.concat(all_module_rows, ignore_index=True) if all_module_rows else pd.DataFrame()

    base_validation_df = build_validation_summary(base_clean_all, scenarios_generated)
    recovery_validation_df = build_validation_summary(recovery_clean_all, scenarios_generated)

    base_outputs = build_methodological_outputs(
        base_clean_all,
        base_phases_all,
        scenarios_generated,
        base_validation_df,
        f"{run_stamp_ret}_base",
    )
    recovery_outputs_raw = build_methodological_outputs(
        recovery_clean_all,
        recovery_phases_all,
        scenarios_generated,
        recovery_validation_df,
        f"{run_stamp_ret}_recovery",
    )
    recovery_outputs = (
        recovery_outputs_raw[0],
        recovery_outputs_raw[1],
        recovery_outputs_raw[2],
        recovery_outputs_raw[3],
        recovery_outputs_raw[4],
        augment_metadata_with_module(recovery_outputs_raw[5], module_all),
    )

    comparison_df, comparison_warnings = build_comparison_results(
        base_summary=base_outputs[0],
        recovery_summary=recovery_outputs[0],
        base_evaluation=base_outputs[1],
        recovery_evaluation=recovery_outputs[1],
        base_scaling=base_outputs[4],
        recovery_scaling=recovery_outputs[4],
    )
    distributed_df = build_distributed_effect_results(module_all, stages, recovery_control)

    export_recovery_compare_outputs(
        base_outputs=base_outputs,
        recovery_outputs=recovery_outputs,
        base_validation_df=base_validation_df,
        recovery_validation_df=recovery_validation_df,
        module_df=module_all,
        comparison_df=comparison_df,
        distributed_df=distributed_df,
        comparison_warnings=comparison_warnings,
        recovery_control=recovery_control,
        batch_results_dir=batch_results_dir,
    )

    print(f"Resultados comparativos base vs recovery generados en: {batch_results_dir}")


if __name__ == "__main__":
    run_all_scenarios()
